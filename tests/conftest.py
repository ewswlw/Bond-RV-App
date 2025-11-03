"""
Shared pytest fixtures for bond pipeline tests
"""
import pytest
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, time
import openpyxl
from openpyxl import Workbook


@pytest.fixture
def sample_dataframe():
    """Sample DataFrame with 10 bonds, valid data"""
    return pd.DataFrame({
        'CUSIP': ['037833CY4', '03524BAH9', '008911BL2', '06051GJG5', '89678ZAB2',
                  '06048WW63', '06051GJK6', '06051GJZ3', '06051GKE8', '06051GKJ7'],
        'Security': ['Bond A', 'Bond B', 'Bond C', 'Bond D', 'Bond E',
                     'Bond F', 'Bond G', 'Bond H', 'Bond I', 'Bond J'],
        'Benchmark Cusip': ['135087B45', '135087M68', '135087J39', '135087E67', '135087M27',
                            '135087B45', '135087M68', '135087J39', '135087E67', '135087M27'],
        'Custom_Sector': ['Non Financial Maple', 'Non Financial Maple', 'HY', 'IG', 'IG',
                          'HY', 'IG', 'Non Financial Maple', 'HY', 'IG'],
        'Pricing Date': [datetime(2020, 8, 13), datetime(2017, 5, 8), datetime(2021, 7, 27),
                        datetime(2019, 3, 15), datetime(2020, 4, 3),
                        datetime(2021, 1, 10), datetime(2019, 6, 20), datetime(2020, 9, 5),
                        datetime(2018, 11, 30), datetime(2022, 2, 14)],
        'Benchmark': ['Benchmark 1', 'Benchmark 2', 'Benchmark 3', 'Benchmark 4', 'Benchmark 5',
                      'Benchmark 1', 'Benchmark 2', 'Benchmark 3', 'Benchmark 4', 'Benchmark 5'],
        'Ticker': ['TICK1', 'TICK2', 'TICK3', 'TICK4', 'TICK5',
                   'TICK6', 'TICK7', 'TICK8', 'TICK9', 'TICK10'],
        'Currency': ['USD', 'USD', 'CAD', 'USD', 'CAD',
                     'USD', 'CAD', 'USD', 'CAD', 'USD'],
    })


@pytest.fixture
def sample_with_duplicates():
    """Sample DataFrame with duplicate Date+CUSIP"""
    df = pd.DataFrame({
        'Date': [datetime(2025, 10, 20)] * 6,
        'CUSIP': ['037833CY4', '037833CY4', '03524BAH9', '03524BAH9', '008911BL2', '008911BL2'],
        'Security': ['Bond A v1', 'Bond A v2', 'Bond B v1', 'Bond B v2', 'Bond C v1', 'Bond C v2'],
        'Benchmark Cusip': ['135087B45'] * 6,
        'Custom_Sector': ['IG'] * 6,
    })
    return df


@pytest.fixture
def sample_invalid_cusips():
    """Sample DataFrame with invalid CUSIPs"""
    return pd.DataFrame({
        'CUSIP': [
            '037833CY4',      # Valid
            '123',            # Too short
            '0636B108',       # 8 chars
            '6698Z3Z452',     # 10 chars
            '880789A#9',      # Invalid character
            'BBG01G27TPY1',   # Bloomberg ID
            '89678zab2',      # Lowercase (needs normalization)
            ' 06418GAD9 Corp', # Extra text
        ],
        'Security': [f'Bond {i}' for i in range(8)],
    })


@pytest.fixture
def sample_old_schema():
    """Sample DataFrame with old schema (59 columns)"""
    data = {
        'CUSIP': ['037833CY4', '03524BAH9'],
        'Security': ['Bond A', 'Bond B'],
        'Benchmark Cusip': ['135087B45', '135087M68'],
    }
    # Add 56 more columns to make 59 total
    for i in range(56):
        data[f'Column_{i}'] = ['Value'] * 2
    return pd.DataFrame(data)


@pytest.fixture
def sample_new_schema():
    """Sample DataFrame with new schema (75 columns)"""
    data = {
        'CUSIP': ['037833CY4', '03524BAH9'],
        'Security': ['Bond A', 'Bond B'],
        'Benchmark Cusip': ['135087B45', '135087M68'],
    }
    # Add 72 more columns to make 75 total
    for i in range(72):
        data[f'Column_{i}'] = ['Value'] * 2
    return pd.DataFrame(data)


@pytest.fixture
def temp_excel_file(tmp_path):
    """Create a temporary Excel file with sample data"""
    def _create_excel(filename, data_df, date_str='10.20.25'):
        """
        Create an Excel file with proper structure
        
        Args:
            filename: Name of the file (e.g., 'API 10.20.25.xlsx')
            data_df: DataFrame to write
            date_str: Date string for filename (MM.DD.YY format)
        """
        filepath = tmp_path / filename
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        
        # Row 1: Empty (matches real files)
        # Row 2: Headers
        headers = list(data_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col_idx, value=header)
        
        # Row 3+: Data
        for row_idx, row_data in enumerate(data_df.itertuples(index=False), start=3):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filepath)
        return filepath
    
    return _create_excel


@pytest.fixture
def sample_dates():
    """Sample dates for testing"""
    return {
        'valid': [
            ('API 10.20.25.xlsx', datetime(2025, 10, 20)),
            ('API 08.04.23.xlsx', datetime(2023, 8, 4)),
            ('API 12.29.23.xlsx', datetime(2023, 12, 29)),
            ('API 01.01.00.xlsx', datetime(2000, 1, 1)),
        ],
        'invalid': [
            'API 2025-10-20.xlsx',  # Wrong format
            'bonds_10_20_25.xlsx',  # Wrong format
            'API 13.32.25.xlsx',    # Invalid date
            'API 02.29.23.xlsx',    # Non-leap year
        ],
        'edge_cases': [
            ('API 02.29.24.xlsx', datetime(2024, 2, 29)),  # Leap year
            ('API 12.31.99.xlsx', datetime(1999, 12, 31)),  # Year 1999
        ]
    }


@pytest.fixture
def sample_cusips():
    """Sample CUSIPs for testing"""
    return {
        'valid': [
            '037833CY4',
            '89678ZAB2',
            '06051GJG5',
            '000000000',  # All zeros
            'AAAAAAAAA',  # All letters
        ],
        'invalid_length': [
            ('123', 3),
            ('0636B108', 8),
            ('6698Z3Z452', 10),
            ('BBG01G27TPY1', 12),
        ],
        'invalid_chars': [
            '880789A#9',   # Contains #
            '880789A 9',   # Contains space
            '880789A-9',   # Contains hyphen
        ],
        'needs_normalization': [
            ('89678zab2', '89678ZAB2'),
            (' 06418GAD9 Corp', '06418GAD9'),
            ('38141GYD0 CORP', '38141GYD0'),
        ]
    }


@pytest.fixture
def sample_na_values():
    """Sample NA values for testing"""
    return [
        '#N/A Field Not Applicable',
        '#N/A Invalid Security',
        'N/A',
        'nan',
        'NaN',
        '#N/A',
    ]


@pytest.fixture
def mock_logger(mocker):
    """Mock logger for testing"""
    return mocker.Mock()


@pytest.fixture
def clean_test_dirs(tmp_path):
    """Create clean test directories"""
    parquet_dir = tmp_path / "parquet"
    logs_dir = tmp_path / "logs"
    raw_data_dir = tmp_path / "raw_data"
    
    parquet_dir.mkdir()
    logs_dir.mkdir()
    raw_data_dir.mkdir()
    
    return {
        'parquet': parquet_dir,
        'logs': logs_dir,
        'raw_data': raw_data_dir,
    }


# ============================================================================
# Runs Pipeline Fixtures
# ============================================================================

@pytest.fixture
def sample_runs_dataframe():
    """Sample DataFrame with RUNS data (30 columns)"""
    from datetime import datetime, time
    return pd.DataFrame({
        'Reference Security': ['AM259052 Corp'] * 5,
        'Date': ['10/31/25', '10/31/25', '10/31/25', '10/31/25', '10/31/25'],
        'Time': ['15:45', '15:45', '15:46', '15:47', '16:00'],
        'Bid Workout Risk': [5.14] * 5,
        'Ticker': ['T'] * 5,
        'Dealer': ['BMO', 'BMO', 'BMO', 'RBC', 'TD'],
        'Source': ['RUN'] * 5,
        'Security': ['T 4.85 05/25/47'] * 5,
        'Bid Price': [91.608, 91.608, 91.609, 91.610, 91.615],
        'Ask Price': [92.226, 92.226, 92.227, 92.228, 92.233],
        'Bid Spread': [235.0, 235.0, 236.0, 237.0, 238.0],
        'Ask Spread': [230.0, 230.0, 231.0, 232.0, 233.0],
        'Benchmark': ['CAN 1 05/25/45'] * 5,
        'Reference Benchmark': ['QZ074427 Corp'] * 5,
        'Bid Size': [3000000.0] * 5,
        'Ask Size': [2000000.0] * 5,
        'Sector': ['HY'] * 5,
        'Bid Yield To Convention': [5.1] * 5,
        'Ask Yield To Convention': [5.05] * 5,
        'Bid Discount Margin': [235.0] * 5,
        'Ask Discount Margin': [230.0] * 5,
        'CUSIP': ['00206RDY5', '00206RDY5', '00206RDY5', '00206RDY5', '00206RGB2'],
        'Sender Name': ['BMOCM'] * 5,
        'Currency': ['CAD'] * 5,
        'Subject': ['BMOCM - ATT #1'] * 5,
        'Keyword': [''] * 5,
        'Bid Interpolated Spread to Government': [235.0] * 5,
        'Ask Interpolated Spread to Government': [230.0] * 5,
        'Bid Contributed Yield': [5.1] * 5,
        'Bid Z-spread': [235.0] * 5,
    })


@pytest.fixture
def sample_runs_with_duplicates():
    """Sample RUNS DataFrame with duplicate Date+Dealer+CUSIP (different times)"""
    return pd.DataFrame({
        'Date': [
            datetime(2025, 10, 31), datetime(2025, 10, 31), datetime(2025, 10, 31),
            datetime(2025, 10, 31), datetime(2025, 10, 31), datetime(2025, 10, 31)
        ],
        'Time': [
            time(7, 17), time(11, 58), time(11, 58),
            time(8, 0), time(8, 0), time(9, 0)
        ],
        'Dealer': ['BMO', 'BMO', 'BMO', 'RBC', 'RBC', 'RBC'],
        'CUSIP': ['00208DAB7', '00208DAB7', '00208DAB7', '00208DAB7', '00208DAB7', '00208DAB7'],
        'Security': ['ARXCN 3.465 03/10/31'] * 6,
        'Bid Price': [98.314, 98.429, 98.429, 98.300, 98.301, 98.302],
        'Ask Price': [98.504, 98.620, 98.620, 98.490, 98.491, 98.492],
        'Bid Spread': [106.0, 106.0, 106.0, 105.0, 105.0, 105.0],
        'Ask Spread': [102.0, 102.0, 102.0, 101.0, 101.0, 101.0],
    })


@pytest.fixture
def sample_runs_with_same_time_duplicates():
    """Sample RUNS DataFrame with same Date+Dealer+CUSIP+Time duplicates"""
    return pd.DataFrame({
        'Date': [datetime(2025, 10, 31), datetime(2025, 10, 31), datetime(2025, 10, 31)],
        'Time': [time(15, 45), time(15, 45), time(15, 45)],
        'Dealer': ['BMO', 'BMO', 'BMO'],
        'CUSIP': ['00206RDY5', '00206RDY5', '00206RDY5'],
        'Security': ['T 4.85 05/25/47'] * 3,
        'Bid Price': [91.608, 91.608, 91.610],  # Last one different
        'Ask Price': [92.226, 92.226, 92.228],
    })


@pytest.fixture
def sample_runs_dates():
    """Sample dates for RUNS testing"""
    return {
        'valid': [
            ('10/31/25', datetime(2025, 10, 31)),
            ('01/02/24', datetime(2024, 1, 2)),
            ('12/29/23', datetime(2023, 12, 29)),
            ('01/01/00', datetime(2000, 1, 1)),
        ],
        'invalid': [
            '2025-10-31',  # Wrong format
            '10-31-25',    # Wrong separator
            '31/10/25',    # Day/month swapped
            '13/32/25',    # Invalid month/day
            '',
            None,
        ],
        'edge_cases': [
            ('02/29/24', datetime(2024, 2, 29)),  # Leap year
            ('12/31/99', datetime(1999, 12, 31)),  # Year 1999
        ]
    }


@pytest.fixture
def sample_runs_times():
    """Sample times for RUNS testing"""
    return {
        'valid': [
            ('15:45', time(15, 45)),
            ('08:12', time(8, 12)),
            ('8:12', time(8, 12)),  # Also accepts without leading zero
            ('23:59', time(23, 59)),
            ('00:00', time(0, 0)),
        ],
        'invalid': [
            '15:45:30',  # Has seconds
            '25:00',     # Invalid hour
            '15:60',     # Invalid minute
            '15-45',     # Wrong separator
            '',
            None,
        ],
        'edge_cases': [
            ('00:00', time(0, 0)),
            ('23:59', time(23, 59)),
        ]
    }


@pytest.fixture
def temp_runs_excel_file(tmp_path):
    """Create a temporary RUNS Excel file with sample data"""
    def _create_runs_excel(filename, data_df):
        """
        Create a RUNS Excel file with proper structure
        
        Args:
            filename: Name of the file (e.g., 'RUNS 10.31.25.xlsx')
            data_df: DataFrame to write (with Date and Time as columns)
        """
        filepath = tmp_path / filename
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        
        # Row 1: Headers (header is row 1, 0-indexed)
        headers = list(data_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Row 2+: Data
        for row_idx, row_data in enumerate(data_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                # Convert datetime objects to strings for Excel
                if isinstance(value, datetime):
                    value = value.strftime('%m/%d/%y')
                elif isinstance(value, time):
                    value = value.strftime('%H:%M')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(filepath)
        return filepath
    
    return _create_runs_excel


@pytest.fixture
def sample_universe_parquet(tmp_path):
    """Create sample universe.parquet for orphan tracking tests"""
    universe_df = pd.DataFrame({
        'CUSIP': [
            '00206RDY5',
            '00206RGB2',
            '00208DAB7',
            '037833CY4',
            '03524BAH9',
        ],
        'Security': [
            'T 4.85 05/25/47',
            'T 4 11/25/25',
            'ARXCN 3.465 03/10/31',
            'Bond A',
            'Bond B',
        ],
    })
    
    parquet_path = tmp_path / 'parquet' / 'universe.parquet'
    parquet_path.parent.mkdir(parents=True, exist_ok=True)
    universe_df.to_parquet(parquet_path, index=False)
    
    return parquet_path
