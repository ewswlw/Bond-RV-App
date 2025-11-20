"""
Runs views script.

This module creates custom formatted tables from runs_today.csv data.
Outputs nicely formatted tables to portfolio_runs_view.txt and portfolio_runs_view.xlsx for portfolio monitoring.
Also generates universe RV views to uni_runs_view.txt and uni_runs_view.xlsx for universe-wide analysis.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional, Dict

import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Get script directory and build paths relative to it
SCRIPT_DIR = Path(__file__).parent.resolve()
RUNS_TODAY_CSV_PATH = SCRIPT_DIR.parent / "processed_data" / "runs_today.csv"
RUNS_PARQUET_PATH = SCRIPT_DIR.parent.parent / "bond_data" / "parquet" / "runs_timeseries.parquet"
OUTPUT_DIR = SCRIPT_DIR.parent / "processed_data"
OUTPUT_FILE = OUTPUT_DIR / "portfolio_runs_view.txt"
EXCEL_OUTPUT_FILE = OUTPUT_DIR / "portfolio_runs_view.xlsx"
UNI_OUTPUT_FILE = OUTPUT_DIR / "uni_runs_view.txt"
UNI_EXCEL_OUTPUT_FILE = OUTPUT_DIR / "uni_runs_view.xlsx"

# ============================================================================
# CONFIGURATION
# ============================================================================

# Column display names (shorter versions for table display)
COLUMN_DISPLAY_NAMES = {
    "Security": "Security",
    "QUANTITY": "Qty",
    "POSITION CR01": "Pos CR01",
    "Yrs (Cvn)": "Yrs Cvn",
    "Tight Bid >3mm": "TB >3mm",
    "Wide Offer >3mm": "WO >3mm",
    "Tight Bid": "TB",
    "Wide Offer": "WO",
    "Bid/Offer>3mm": "B/O>3mm",
    "Bid/Offer": "B/O",
    "Dealer @ Tight Bid >3mm": "Dealer TB>3mm",
    "Dealer @ Wide Offer >3mm": "Dealer WO>3mm",
    "Size @ Tight Bid >3mm": "Size TB>3mm",
    "Size @ Wide Offer >3mm": "Size WO>3mm",
    "CR01 @ Tight Bid": "CR01 TB",
    "CR01 @ Wide Offer": "CR01 WO",
    "# of Bids >3mm": "# Bids>3mm",
    "# of Offers >3mm": "# Offers>3mm",
    "DoD Chg Tight Bid >3mm": "DoD TB>3mm",
    "DoD Chg Wide Offer >3mm": "DoD WO>3mm",
    "DoD Chg Tight Bid": "DoD TB",
    "DoD Chg Wide Offer": "DoD WO",
    "DoD Chg Size @ Tight Bid >3mm": "DoD Size TB>3mm",
    "DoD Chg Size @ Wide Offer >3mm": "DoD Size WO>3mm",
    "MTD Chg Tight Bid": "MTD TB",
    "YTD Chg Tight Bid": "YTD TB",
    "1yr Chg Tight Bid": "1yr TB",
    "MTD Equity": "MTD Equity",
    "YTD Equity": "YTD Equity",
    "Retracement": "Retracement",
    "Custom_Sector": "Custom Sector",
}

# Column order for Portfolio Sorted By CR01 Risk table
PORTFOLIO_CR01_RISK_COLUMNS = [
    "Security",
    "QUANTITY",
    "POSITION CR01",
    "Yrs (Cvn)",
    "Tight Bid >3mm",
    "Wide Offer >3mm",
    "Tight Bid",
    "Wide Offer",
    "Bid/Offer>3mm",
    "Bid/Offer",
    "Dealer @ Tight Bid >3mm",
    "Dealer @ Wide Offer >3mm",
    "Size @ Tight Bid >3mm",
    "Size @ Wide Offer >3mm",
    "CR01 @ Tight Bid",
    "CR01 @ Wide Offer",
    "# of Bids >3mm",
    "# of Offers >3mm",
    "DoD Chg Tight Bid >3mm",
    "DoD Chg Wide Offer >3mm",
    "DoD Chg Tight Bid",
    "DoD Chg Wide Offer",
    "DoD Chg Size @ Tight Bid >3mm",
    "DoD Chg Size @ Wide Offer >3mm",
    "MTD Chg Tight Bid",
    "YTD Chg Tight Bid",
    "1yr Chg Tight Bid",
    "MTD Equity",
    "YTD Equity",
    "Retracement",
    "Custom_Sector",
]

# Column order for MTD table (excludes DoD columns)
PORTFOLIO_MTD_COLUMNS = [
    "Security",
    "QUANTITY",
    "POSITION CR01",
    "Yrs (Cvn)",
    "Tight Bid >3mm",
    "Wide Offer >3mm",
    "Tight Bid",
    "Wide Offer",
    "Bid/Offer>3mm",
    "Bid/Offer",
    "Dealer @ Tight Bid >3mm",
    "Dealer @ Wide Offer >3mm",
    "Size @ Tight Bid >3mm",
    "Size @ Wide Offer >3mm",
    "CR01 @ Tight Bid",
    "CR01 @ Wide Offer",
    "# of Bids >3mm",
    "# of Offers >3mm",
    "MTD Chg Tight Bid",
    "YTD Chg Tight Bid",
    "1yr Chg Tight Bid",
    "MTD Equity",
    "YTD Equity",
    "Retracement",
    "Custom_Sector",
]

# Column order for YTD table (excludes DoD and MTD columns)
PORTFOLIO_YTD_COLUMNS = [
    "Security",
    "QUANTITY",
    "POSITION CR01",
    "Yrs (Cvn)",
    "Tight Bid >3mm",
    "Wide Offer >3mm",
    "Tight Bid",
    "Wide Offer",
    "Bid/Offer>3mm",
    "Bid/Offer",
    "Dealer @ Tight Bid >3mm",
    "Dealer @ Wide Offer >3mm",
    "Size @ Tight Bid >3mm",
    "Size @ Wide Offer >3mm",
    "CR01 @ Tight Bid",
    "CR01 @ Wide Offer",
    "# of Bids >3mm",
    "# of Offers >3mm",
    "YTD Chg Tight Bid",
    "1yr Chg Tight Bid",
    "MTD Equity",
    "YTD Equity",
    "Retracement",
    "Custom_Sector",
]

# Column order for 1yr table (excludes DoD, MTD, and YTD columns)
PORTFOLIO_1YR_COLUMNS = [
    "Security",
    "QUANTITY",
    "POSITION CR01",
    "Yrs (Cvn)",
    "Tight Bid >3mm",
    "Wide Offer >3mm",
    "Tight Bid",
    "Wide Offer",
    "Bid/Offer>3mm",
    "Bid/Offer",
    "Dealer @ Tight Bid >3mm",
    "Dealer @ Wide Offer >3mm",
    "Size @ Tight Bid >3mm",
    "Size @ Wide Offer >3mm",
    "CR01 @ Tight Bid",
    "CR01 @ Wide Offer",
    "# of Bids >3mm",
    "# of Offers >3mm",
    "1yr Chg Tight Bid",
    "MTD Equity",
    "YTD Equity",
    "Retracement",
    "Custom_Sector",
]

# ============================================================================
# UNIVERSE RV VIEWS CONFIGURATION
# ============================================================================

# Custom_Sector values to filter out from universe tables
UNIVERSE_EXCLUDED_SECTORS = [
    "Asset Backed Subs",
    "Auto ABS",
    "Bail In",
    "CAD Govt",
    "CASH CAD",
    "CASH USD",
    "CDX",
    "CP",
    "Covered",
    "Dep Note",
    "Financial Hybrid",
    "HY",
    "Non Financial Hybrid",
    "Non Financial Hybrids",
    "USD Govt",
    "University",
    "Utility",
]

# Number of top and bottom rows to show for DoD moves table
UNIVERSE_TOP_BOTTOM_N = 20

# Sort column for Universe Sorted By DoD Moves With Size On Offer >3mm table
UNIVERSE_DOD_SORT_COLUMN = "DoD Chg Wide Offer >3mm"

# Column order for Universe Sorted By DoD Moves With Size On Offer >3mm table
UNIVERSE_DOD_MOVES_COLUMNS = [
    "Security",
    "Yrs (Cvn)",
    "Tight Bid >3mm",
    "Wide Offer >3mm",
    "Tight Bid",
    "Wide Offer",
    "Bid/Offer>3mm",
    "Bid/Offer",
    "Dealer @ Tight Bid >3mm",
    "Dealer @ Wide Offer >3mm",
    "Size @ Tight Bid >3mm",
    "Size @ Wide Offer >3mm",
    "CR01 @ Tight Bid",
    "CR01 @ Wide Offer",
    "# of Bids >3mm",
    "# of Offers >3mm",
    "DoD Chg Tight Bid >3mm",
    "DoD Chg Wide Offer >3mm",
    "DoD Chg Tight Bid",
    "DoD Chg Wide Offer",
    "DoD Chg Size @ Tight Bid >3mm",
    "DoD Chg Size @ Wide Offer >3mm",
    "MTD Chg Tight Bid",
    "YTD Chg Tight Bid",
    "1yr Chg Tight Bid",
    "MTD Equity",
    "YTD Equity",
    "Retracement",
    "Custom_Sector",
]

# Column order for Universe Sorted By MTD Moves With Size On Offer >3mm table
# Same as DoD Moves but excludes DoD columns (between # of Offers >3mm and MTD Chg Tight Bid)
UNIVERSE_MTD_MOVES_COLUMNS = [
    "Security",
    "Yrs (Cvn)",
    "Tight Bid >3mm",
    "Wide Offer >3mm",
    "Tight Bid",
    "Wide Offer",
    "Bid/Offer>3mm",
    "Bid/Offer",
    "Dealer @ Tight Bid >3mm",
    "Dealer @ Wide Offer >3mm",
    "Size @ Tight Bid >3mm",
    "Size @ Wide Offer >3mm",
    "CR01 @ Tight Bid",
    "CR01 @ Wide Offer",
    "# of Bids >3mm",
    "# of Offers >3mm",
    "MTD Chg Tight Bid",
    "YTD Chg Tight Bid",
    "1yr Chg Tight Bid",
    "MTD Equity",
    "YTD Equity",
    "Retracement",
    "Custom_Sector",
]

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def ensure_ascii(value: Optional[str]) -> str:
    """
    Convert text to ASCII by removing or replacing unsupported characters.
    
    Args:
        value: Input string that may contain non-ASCII characters.
    
    Returns:
        ASCII-safe string representation.
    """
    if value is None or pd.isna(value):
        return ""
    sanitized = str(value).encode("ascii", errors="ignore").decode("ascii")
    if sanitized:
        return sanitized
    return str(value).encode("ascii", errors="replace").decode("ascii")


def ensure_numeric_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert all numeric columns to proper numeric types after reading CSV.
    
    This ensures that filtering operations (e.g., DoD TB>3mm) work correctly.
    
    Args:
        df: DataFrame read from CSV (may have numeric columns as object/string types).
    
    Returns:
        DataFrame with all numeric columns converted to float64.
    """
    df_converted = df.copy()
    
    # List of all numeric columns that should be converted
    numeric_columns = [
        # Spread columns
        "Tight Bid >3mm",
        "Wide Offer >3mm",
        "Tight Bid",
        "Wide Offer",
        "Bid/Offer>3mm",
        "Bid/Offer",
        # Size columns
        "Size @ Tight Bid >3mm",
        "Size @ Wide Offer >3mm",
        "Cumm. Bid Size",
        "Cumm. Offer Size",
        "Bid Size RBC",
        "Offer Size RBC",
        # CR01 columns
        "CR01 @ Tight Bid",
        "CR01 @ Wide Offer",
        # RBC columns
        "Bid RBC",
        "Ask RBC",
        # Count columns
        "# of Bids >3mm",
        "# of Offers >3mm",
        # Portfolio columns
        "QUANTITY",
        "POSITION CR01",
        # Bond details columns
        "G Sprd",
        "Yrs (Cvn)",
        "vs BI",
        "vs BCE",
        "MTD Equity",
        "YTD Equity",
        "Retracement",
        "Yrs Since Issue",
        "Z Score",
        "Retracement2",
        # Risk column
        "Bid Workout Risk",
    ]
    
    # Convert each numeric column
    for col in numeric_columns:
        if col in df_converted.columns:
            df_converted[col] = pd.to_numeric(df_converted[col], errors="coerce")
    
    # Convert all change columns (DoD Chg *, MTD Chg *, YTD Chg *, 1yr Chg *)
    for col in df_converted.columns:
        if any(col.startswith(prefix) for prefix in ["DoD Chg ", "MTD Chg ", "YTD Chg ", "1yr Chg "]):
            df_converted[col] = pd.to_numeric(df_converted[col], errors="coerce")
    
    return df_converted


def format_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Format numeric columns with thousand separators (no decimals, rounded).
    Retracement column formatted as percentage with 2 decimals and "%" suffix.
    Yrs (Cvn) column formatted to 1 decimal place.
    
    Args:
        df: DataFrame to format.
    
    Returns:
        Formatted DataFrame with numeric columns as formatted strings.
    """
    df_formatted = df.copy()
    for col in df_formatted.columns:
        if pd.api.types.is_numeric_dtype(df_formatted[col]):
            if col == "Retracement":
                # Format as percentage with 2 decimals (multiply by 100: 1.0 = 100%)
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: f"{x * 100:.2f}%" if pd.notna(x) else ""
                )
            elif col == "Yrs (Cvn)":
                # Format to 1 decimal place
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: f"{x:.1f}" if pd.notna(x) else ""
                )
            else:
                # Format with thousand separators, no decimals (rounded)
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: f"{int(round(x)):,}" if pd.notna(x) else ""
                )
    return df_formatted


def get_reference_dates(runs_path: Path) -> dict:
    """
    Extract MTD and YTD reference dates from runs_timeseries.parquet.
    
    Args:
        runs_path: Path to runs_timeseries.parquet file.
    
    Returns:
        Dictionary with 'last_date', 'mtd_ref_date', 'ytd_ref_date', 'one_yr_ref_date'.
    """
    df = pd.read_parquet(runs_path)
    unique_dates = sorted(df["Date"].unique())
    
    if len(unique_dates) < 1:
        return {
            "last_date": None,
            "mtd_ref_date": None,
            "ytd_ref_date": None,
            "one_yr_ref_date": None,
        }
    
    last_date = unique_dates[-1]
    
    # MTD: First date of current month
    mtd_ref_date = pd.Timestamp(last_date.year, last_date.month, 1)
    mtd_dates = [d for d in unique_dates if d >= mtd_ref_date and d < last_date]
    mtd_ref_date = mtd_dates[0] if mtd_dates else None
    
    # YTD: First date of current year
    ytd_ref_date = pd.Timestamp(last_date.year, 1, 1)
    ytd_dates = [d for d in unique_dates if d >= ytd_ref_date and d < last_date]
    ytd_ref_date = ytd_dates[0] if ytd_dates else None
    
    # 1yr: Approximately 1 year ago (closest available date)
    one_year_ago = last_date - pd.DateOffset(years=1)
    one_yr_dates = [d for d in unique_dates if d <= one_year_ago]
    one_yr_ref_date = one_yr_dates[-1] if one_yr_dates else None
    
    return {
        "last_date": last_date,
        "mtd_ref_date": mtd_ref_date,
        "ytd_ref_date": ytd_ref_date,
        "one_yr_ref_date": one_yr_ref_date,
    }


def validate_runs_today_csv_is_current(
    csv_path: Path,
    parquet_path: Path,
    auto_regenerate: bool = True
) -> pd.Timestamp:
    """
    Validate that runs_today.csv is current with runs_timeseries.parquet.
    
    Checks if the CSV file exists and reflects the latest date in the parquet file.
    If CSV is stale or missing, optionally regenerates it.
    
    Args:
        csv_path: Path to runs_today.csv file.
        parquet_path: Path to runs_timeseries.parquet file.
        auto_regenerate: If True, automatically regenerate CSV if stale. If False, raises error.
    
    Returns:
        Last date from parquet file (pd.Timestamp).
    
    Raises:
        FileNotFoundError: If parquet file doesn't exist.
        ValueError: If CSV is stale and auto_regenerate is False.
    """
    # Check parquet file exists
    if not parquet_path.exists():
        raise FileNotFoundError(f"Parquet file not found: {parquet_path}")
    
    # Get last date from parquet
    ref_dates = get_reference_dates(parquet_path)
    parquet_last_date = ref_dates["last_date"]
    
    if parquet_last_date is None:
        raise ValueError(f"No dates found in parquet file: {parquet_path}")
    
    # Check if CSV exists
    csv_exists = csv_path.exists()
    
    # Check if CSV is stale (parquet modified more recently than CSV)
    if csv_exists:
        parquet_mtime = parquet_path.stat().st_mtime
        csv_mtime = csv_path.stat().st_mtime
        
        # If parquet is newer than CSV, CSV might be stale
        if parquet_mtime > csv_mtime:
            if auto_regenerate:
                print(f"\nWARNING: runs_today.csv is stale (parquet modified after CSV).")
                print(f"  Parquet last modified: {datetime.fromtimestamp(parquet_mtime)}")
                print(f"  CSV last modified: {datetime.fromtimestamp(csv_mtime)}")
                print(f"  Auto-regenerating runs_today.csv...")
                
                # Import and run runs_today.py to regenerate CSV
                import sys
                import subprocess
                runs_today_script = csv_path.parent.parent / "runs" / "runs_today.py"
                if runs_today_script.exists():
                    # Use current Python interpreter and set working directory to script location
                    result = subprocess.run(
                        [sys.executable, str(runs_today_script)],
                        cwd=str(runs_today_script.parent),
                        capture_output=True,
                        text=True
                    )
                    if result.returncode != 0:
                        raise RuntimeError(
                            f"Failed to regenerate runs_today.csv:\n{result.stderr}\n{result.stdout}"
                        )
                    print("  Successfully regenerated runs_today.csv")
                else:
                    raise FileNotFoundError(
                        f"Cannot auto-regenerate: runs_today.py not found at {runs_today_script}"
                    )
            else:
                raise ValueError(
                    f"runs_today.csv is stale. Parquet last date: {parquet_last_date}, "
                    f"but CSV was generated earlier. Please regenerate runs_today.csv first."
                )
    else:
        # CSV doesn't exist, regenerate if auto_regenerate is True
        if auto_regenerate:
            print(f"\nWARNING: runs_today.csv not found. Auto-generating...")
            import sys
            import subprocess
            runs_today_script = csv_path.parent.parent / "runs" / "runs_today.py"
            if runs_today_script.exists():
                # Use current Python interpreter and set working directory to script location
                result = subprocess.run(
                    [sys.executable, str(runs_today_script)],
                    cwd=str(runs_today_script.parent),
                    capture_output=True,
                    text=True
                )
                if result.returncode != 0:
                    raise RuntimeError(
                        f"Failed to generate runs_today.csv:\n{result.stderr}\n{result.stdout}"
                    )
                print("  Successfully generated runs_today.csv")
            else:
                raise FileNotFoundError(
                    f"Cannot auto-generate: runs_today.py not found at {runs_today_script}"
                )
        else:
            raise FileNotFoundError(
                f"runs_today.csv not found: {csv_path}. "
                f"Please generate it first by running runs_today.py"
            )
    
    return parquet_last_date


def format_table(df: pd.DataFrame, title: str, column_display_names: dict, summary_dict: dict = None) -> str:
    """
    Format DataFrame as a nicely formatted text table with center alignment.
    
    Args:
        df: DataFrame to format.
        title: Table title.
        column_display_names: Dictionary mapping original column names to display names.
        summary_dict: Optional dictionary with summary statistics to display at top (e.g., {"Cumulative CR01 TB": 12345}).
    
    Returns:
        Formatted table string with center-aligned columns.
    """
    if df.empty:
        return f"\n{'='*100}\n{title}\n{'='*100}\nNo data available.\n"
    
    # Create copy for formatting
    df_formatted = df.copy()
    
    # Format numeric columns (thousand separators, Retracement as percentage)
    df_formatted = format_numeric_columns(df_formatted)
    
    # Replace NaN with empty string (blank)
    df_formatted = df_formatted.fillna("")
    
    # Ensure ASCII-safe strings
    for col in df_formatted.columns:
        if df_formatted[col].dtype == 'object':
            df_formatted[col] = df_formatted[col].apply(
                lambda x: ensure_ascii(str(x)) if x != "" else ""
            )
    
    # Rename columns to display names
    display_columns = {}
    for col in df_formatted.columns:
        display_columns[col] = column_display_names.get(col, col)
    df_formatted = df_formatted.rename(columns=display_columns)
    
    # Convert all columns to string for consistent formatting
    for col in df_formatted.columns:
        df_formatted[col] = df_formatted[col].astype(str)
    
    # Calculate column widths (max of header length and max content length)
    col_widths = {}
    for col in df_formatted.columns:
        header_len = len(str(col))
        content_len = df_formatted[col].str.len().max() if len(df_formatted) > 0 else 0
        col_widths[col] = max(header_len, content_len, 3)  # Minimum width of 3
    
    # Build table with center alignment
    output_lines = []
    
    # Header row (center-aligned)
    header_parts = []
    for col in df_formatted.columns:
        header_parts.append(str(col).center(col_widths[col]))
    output_lines.append("  ".join(header_parts))
    
    # Separator row
    separator_parts = []
    for col in df_formatted.columns:
        separator_parts.append("-" * col_widths[col])
    output_lines.append("  ".join(separator_parts))
    
    # Data rows (center-aligned)
    for idx, row in df_formatted.iterrows():
        row_parts = []
        for col in df_formatted.columns:
            value = str(row[col])
            row_parts.append(value.center(col_widths[col]))
        output_lines.append("  ".join(row_parts))
    
    table_str = "\n".join(output_lines)
    
    # Build formatted output
    output = f"\n{'='*100}\n{title}\n{'='*100}\n"
    
    # Add summary statistics if provided
    if summary_dict:
        for key, value in summary_dict.items():
            # Format numeric values with thousand separators
            if isinstance(value, (int, float)):
                formatted_value = f"{int(round(value)):,}"
            else:
                formatted_value = str(value)
            output += f"{key}: {formatted_value}\n"
    
    output += f"Total rows: {len(df):,}\n"
    output += f"{'-'*100}\n"
    output += table_str
    output += f"\n{'='*100}\n"
    
    return output


def create_portfolio_cr01_risk_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Portfolio Sorted By CR01 Risk table.
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Sort by POSITION CR01 descending (largest to smallest)
    df_filtered = df_filtered.sort_values("POSITION CR01", ascending=False, na_position='last')
    
    # Select only required columns (in order)
    available_columns = [col for col in PORTFOLIO_CR01_RISK_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_portfolio_less_liquid_lines_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Portfolio Less Liquid Lines table (replica of CR01 Risk but only shows rows where TB >3mm is blank).
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to rows where Tight Bid >3mm is blank (NaN/null)
    tb_col = "Tight Bid >3mm"
    if tb_col in df_filtered.columns:
        df_filtered = df_filtered[df_filtered[tb_col].isna()].copy()
    
    # Sort by POSITION CR01 descending (largest to smallest)
    df_filtered = df_filtered.sort_values("POSITION CR01", ascending=False, na_position='last')
    
    # Select only required columns (same as CR01 Risk table)
    available_columns = [col for col in PORTFOLIO_CR01_RISK_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_portfolio_dod_bid_chg_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Portfolio Sorted By DoD Bid Chg With >3MM on Bid table.
    
    Filters to rows where Tight Bid >3mm has a value (non-blank) and DoD Chg Tight Bid >3mm is non-zero (positive or negative).
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to rows where Tight Bid >3mm has a value (non-blank)
    tb_col = "Tight Bid >3mm"
    if tb_col in df_filtered.columns:
        df_filtered = df_filtered[df_filtered[tb_col].notna()].copy()
    
    # Filter to rows where DoD Chg Tight Bid >3mm is non-zero (positive or negative) and non-blank
    dod_col = "DoD Chg Tight Bid >3mm"
    if dod_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[dod_col].notna() & (df_filtered[dod_col] != 0)
        ].copy()
    
    # Sort by DoD Chg Tight Bid >3mm descending (largest changes first)
    if dod_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(dod_col, ascending=False, na_position='last')
    
    # Select only required columns (in order)
    available_columns = [col for col in PORTFOLIO_CR01_RISK_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_portfolio_mtd_bid_chg_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Portfolio Sorted By MTD Bid Chg With >3MM on Bid table.
    
    Filters to rows where Tight Bid >3mm has a value (non-blank) and MTD Chg Tight Bid is non-zero (positive or negative).
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to rows where Tight Bid >3mm has a value (non-blank)
    tb_col = "Tight Bid >3mm"
    if tb_col in df_filtered.columns:
        df_filtered = df_filtered[df_filtered[tb_col].notna()].copy()
    
    # Filter to rows where MTD Chg Tight Bid is non-zero (positive or negative) and non-blank
    mtd_col = "MTD Chg Tight Bid"
    if mtd_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[mtd_col].notna() & (df_filtered[mtd_col] != 0)
        ].copy()
    
    # Sort by MTD Chg Tight Bid descending (largest changes first)
    if mtd_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(mtd_col, ascending=False, na_position='last')
    
    # Select only required columns (in order) - excludes DoD columns
    available_columns = [col for col in PORTFOLIO_MTD_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_portfolio_ytd_bid_chg_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Portfolio Sorted By YTD Bid Chg With >3MM on Bid table.
    
    Filters to rows where Tight Bid >3mm has a value (non-blank) and YTD Chg Tight Bid is non-zero (positive or negative).
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to rows where Tight Bid >3mm has a value (non-blank)
    tb_col = "Tight Bid >3mm"
    if tb_col in df_filtered.columns:
        df_filtered = df_filtered[df_filtered[tb_col].notna()].copy()
    
    # Filter to rows where YTD Chg Tight Bid is non-zero (positive or negative) and non-blank
    ytd_col = "YTD Chg Tight Bid"
    if ytd_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[ytd_col].notna() & (df_filtered[ytd_col] != 0)
        ].copy()
    
    # Sort by YTD Chg Tight Bid descending (largest changes first)
    if ytd_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(ytd_col, ascending=False, na_position='last')
    
    # Select only required columns (in order) - excludes DoD and MTD columns
    available_columns = [col for col in PORTFOLIO_YTD_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_portfolio_1yr_bid_chg_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Portfolio Sorted By 1yr Bid Chg With >3MM on Bid table.
    
    Filters to rows where Tight Bid >3mm has a value (non-blank) and 1yr Chg Tight Bid is non-zero (positive or negative).
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to rows where Tight Bid >3mm has a value (non-blank)
    tb_col = "Tight Bid >3mm"
    if tb_col in df_filtered.columns:
        df_filtered = df_filtered[df_filtered[tb_col].notna()].copy()
    
    # Filter to rows where 1yr Chg Tight Bid is non-zero (positive or negative) and non-blank
    one_yr_col = "1yr Chg Tight Bid"
    if one_yr_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[one_yr_col].notna() & (df_filtered[one_yr_col] != 0)
        ].copy()
    
    # Sort by 1yr Chg Tight Bid descending (largest changes first)
    if one_yr_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(one_yr_col, ascending=False, na_position='last')
    
    # Select only required columns (in order) - excludes DoD, MTD, and YTD columns
    available_columns = [col for col in PORTFOLIO_1YR_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_size_bids_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Size Bids table (exactly like CR01 Risk but sorted by CR01 TB).
    
    Filters to rows where:
    - QUANTITY > 0
    - CR01 @ Tight Bid >= 1000
    - POSITION CR01 >= 1,000
    
    Sorted by CR01 @ Tight Bid descending.
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to CR01 @ Tight Bid >= 1000
    cr01_tb_col = "CR01 @ Tight Bid"
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[cr01_tb_col].notna() & (df_filtered[cr01_tb_col] >= 1000)
        ].copy()
    
    # Filter to POSITION CR01 >= 1,000
    pos_cr01_col = "POSITION CR01"
    if pos_cr01_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[pos_cr01_col].notna() & (df_filtered[pos_cr01_col] >= 1000)
        ].copy()
    
    # Sort by CR01 @ Tight Bid descending (largest first)
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(cr01_tb_col, ascending=False, na_position='last')
    
    # Select only required columns (same as CR01 Risk table)
    available_columns = [col for col in PORTFOLIO_CR01_RISK_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_size_bids_struggling_names_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Size Bids Struggling Names table (same as Size Bids but filters to Retracement < 50%).
    
    Filters to rows where:
    - QUANTITY > 0
    - CR01 @ Tight Bid >= 1000
    - POSITION CR01 >= 1,000
    - Retracement < 0.5 (Retracement < 50%, where Retracement is stored as decimal)
    
    Sorted by CR01 @ Tight Bid descending.
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to CR01 @ Tight Bid >= 1000
    cr01_tb_col = "CR01 @ Tight Bid"
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[cr01_tb_col].notna() & (df_filtered[cr01_tb_col] >= 1000)
        ].copy()
    
    # Filter to POSITION CR01 >= 1,000
    pos_cr01_col = "POSITION CR01"
    if pos_cr01_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[pos_cr01_col].notna() & (df_filtered[pos_cr01_col] >= 1000)
        ].copy()
    
    # Filter to Retracement < 0.5 (Retracement < 50%)
    retracement_col = "Retracement"
    if retracement_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[retracement_col].notna() & (df_filtered[retracement_col] < 0.5)
        ].copy()
    
    # Sort by CR01 @ Tight Bid descending (largest first)
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(cr01_tb_col, ascending=False, na_position='last')
    
    # Select only required columns (same as CR01 Risk table)
    available_columns = [col for col in PORTFOLIO_CR01_RISK_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_size_bids_heavily_offered_lines_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Size Bids Heavily Offered Lines table (same as Size Bids but excludes Bail In and filters to # Offers>3mm > 2).
    
    Filters to rows where:
    - QUANTITY > 0
    - CR01 @ Tight Bid >= 1000
    - POSITION CR01 >= 1,000
    - Custom_Sector != "Bail In" (excludes Bail In sector)
    - # Offers>3mm > 2
    
    Sorted by CR01 @ Tight Bid descending.
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to CR01 @ Tight Bid >= 1000
    cr01_tb_col = "CR01 @ Tight Bid"
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[cr01_tb_col].notna() & (df_filtered[cr01_tb_col] >= 1000)
        ].copy()
    
    # Filter to POSITION CR01 >= 1,000
    pos_cr01_col = "POSITION CR01"
    if pos_cr01_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[pos_cr01_col].notna() & (df_filtered[pos_cr01_col] >= 1000)
        ].copy()
    
    # Filter out "Bail In" from Custom_Sector
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | (df_filtered[custom_sector_col] != "Bail In")
        ].copy()
    
    # Filter to # of Offers >3mm > 2
    offers_3mm_col = "# of Offers >3mm"
    if offers_3mm_col in df_filtered.columns:
        # Convert to numeric in case it's stored as string
        df_filtered[offers_3mm_col] = pd.to_numeric(df_filtered[offers_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[offers_3mm_col].notna() & (df_filtered[offers_3mm_col] > 2)
        ].copy()
    
    # Sort by CR01 @ Tight Bid descending (largest first)
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(cr01_tb_col, ascending=False, na_position='last')
    
    # Select only required columns (same as CR01 Risk table)
    available_columns = [col for col in PORTFOLIO_CR01_RISK_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_size_bids_minimal_bo_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Size Bids With Minimal Bid/Offer table (same as Size Bids but with Bid/Offer>3mm filter).
    
    Filters to rows where:
    - QUANTITY > 0
    - CR01 @ Tight Bid >= 1000
    - POSITION CR01 >= 1,000
    - Bid/Offer>3mm <= 3 (excludes > 3 or blank)
    
    Sorted by CR01 @ Tight Bid descending.
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to CR01 @ Tight Bid >= 1000
    cr01_tb_col = "CR01 @ Tight Bid"
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[cr01_tb_col].notna() & (df_filtered[cr01_tb_col] >= 1000)
        ].copy()
    
    # Filter to POSITION CR01 >= 1,000
    pos_cr01_col = "POSITION CR01"
    if pos_cr01_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[pos_cr01_col].notna() & (df_filtered[pos_cr01_col] >= 1000)
        ].copy()
    
    # Filter to Bid/Offer>3mm <= 3 (excludes > 3 or blank)
    bo_3mm_col = "Bid/Offer>3mm"
    if bo_3mm_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[bo_3mm_col].notna() & (df_filtered[bo_3mm_col] <= 3)
        ].copy()
    
    # Sort by CR01 @ Tight Bid descending (largest first)
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(cr01_tb_col, ascending=False, na_position='last')
    
    # Select only required columns (same as CR01 Risk table)
    available_columns = [col for col in PORTFOLIO_CR01_RISK_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_size_bids_minimal_bo_no_bail_in_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create Size Bids With Minimal Bid/Offer No Bail In table (same as Size Bids With Minimal Bid/Offer but excludes Bail In sector).
    
    Filters to rows where:
    - QUANTITY > 0
    - CR01 @ Tight Bid >= 1000
    - POSITION CR01 >= 1,000
    - Bid/Offer>3mm <= 3 (excludes > 3 or blank)
    - Custom_Sector != "Bail In" (excludes Bail In sector)
    
    Sorted by CR01 @ Tight Bid descending.
    
    Args:
        df: Input DataFrame from runs_today.csv.
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Filter to QUANTITY > 0
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to CR01 @ Tight Bid >= 1000
    cr01_tb_col = "CR01 @ Tight Bid"
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[cr01_tb_col].notna() & (df_filtered[cr01_tb_col] >= 1000)
        ].copy()
    
    # Filter to POSITION CR01 >= 1,000
    pos_cr01_col = "POSITION CR01"
    if pos_cr01_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[pos_cr01_col].notna() & (df_filtered[pos_cr01_col] >= 1000)
        ].copy()
    
    # Filter to Bid/Offer>3mm <= 3 (excludes > 3 or blank)
    bo_3mm_col = "Bid/Offer>3mm"
    if bo_3mm_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[bo_3mm_col].notna() & (df_filtered[bo_3mm_col] <= 3)
        ].copy()
    
    # Filter out "Bail In" from Custom_Sector
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | (df_filtered[custom_sector_col] != "Bail In")
        ].copy()
    
    # Sort by CR01 @ Tight Bid descending (largest first)
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(cr01_tb_col, ascending=False, na_position='last')
    
    # Select only required columns (same as CR01 Risk table)
    available_columns = [col for col in PORTFOLIO_CR01_RISK_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_size_bids_minimal_bo_by_dealer_table(df: pd.DataFrame, dealer: str) -> pd.DataFrame:
    """
    Create Size Bids With Minimal Bid/Offer table filtered by specific dealer.
    
    Filters to rows where:
    - QUANTITY > 0
    - CR01 @ Tight Bid >= 1000
    - POSITION CR01 >= 1,000
    - Bid/Offer>3mm <= 3 (excludes > 3 or blank)
    - Dealer @ Tight Bid >3mm == dealer
    
    Sorted by CR01 @ Tight Bid descending.
    
    Args:
        df: Input DataFrame from runs_today.csv.
        dealer: Dealer name to filter by (e.g., "TD", "RBC", "BMO").
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Start with Size Bids With Minimal Bid/Offer filters
    df_filtered = df[df["QUANTITY"] > 0].copy()
    
    # Filter to CR01 @ Tight Bid >= 1000
    cr01_tb_col = "CR01 @ Tight Bid"
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[cr01_tb_col].notna() & (df_filtered[cr01_tb_col] >= 1000)
        ].copy()
    
    # Filter to POSITION CR01 >= 1,000
    pos_cr01_col = "POSITION CR01"
    if pos_cr01_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[pos_cr01_col].notna() & (df_filtered[pos_cr01_col] >= 1000)
        ].copy()
    
    # Filter to Bid/Offer>3mm <= 3 (excludes > 3 or blank)
    bo_3mm_col = "Bid/Offer>3mm"
    if bo_3mm_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[bo_3mm_col].notna() & (df_filtered[bo_3mm_col] <= 3)
        ].copy()
    
    # Filter to specific dealer in "Dealer @ Tight Bid >3mm"
    dealer_col = "Dealer @ Tight Bid >3mm"
    if dealer_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[dealer_col].notna() & (df_filtered[dealer_col] == dealer)
        ].copy()
    
    # Sort by CR01 @ Tight Bid descending (largest first)
    if cr01_tb_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(cr01_tb_col, ascending=False, na_position='last')
    
    # Select only required columns (same as CR01 Risk table)
    available_columns = [col for col in PORTFOLIO_CR01_RISK_COLUMNS if col in df_filtered.columns]
    df_filtered = df_filtered[available_columns].copy()
    
    return df_filtered


def create_universe_dod_moves_table(
    df: pd.DataFrame,
    excluded_sectors: list[str] = None,
    sort_column: str = None,
    top_bottom_n: int = None,
    columns: list[str] = None
) -> pd.DataFrame:
    """
    Create Universe Sorted By DoD Moves With Size On Offer >3mm table.
    
    Filters out excluded Custom_Sector values, excludes rows where sort column is 0 or blank,
    and shows top N and bottom N rows by sort column (largest positive and most negative values).
    
    Args:
        df: Input DataFrame from runs_today.csv.
        excluded_sectors: List of Custom_Sector values to exclude (defaults to UNIVERSE_EXCLUDED_SECTORS).
        sort_column: Column to sort by (defaults to UNIVERSE_DOD_SORT_COLUMN).
        top_bottom_n: Number of top and bottom rows to show (defaults to UNIVERSE_TOP_BOTTOM_N).
        columns: List of columns to include (defaults to UNIVERSE_DOD_MOVES_COLUMNS).
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Use defaults from config if not provided
    if excluded_sectors is None:
        excluded_sectors = UNIVERSE_EXCLUDED_SECTORS
    if sort_column is None:
        sort_column = UNIVERSE_DOD_SORT_COLUMN
    if top_bottom_n is None:
        top_bottom_n = UNIVERSE_TOP_BOTTOM_N
    if columns is None:
        columns = UNIVERSE_DOD_MOVES_COLUMNS
    
    # Start with copy of input DataFrame
    df_filtered = df.copy()
    
    # Filter out excluded Custom_Sector values
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | 
            (~df_filtered[custom_sector_col].isin(excluded_sectors))
        ].copy()
    
    # Filter to rows that had "Wide Offer >3mm" on both last date and second-to-last date
    # This ensures DoD changes are meaningful (comparing apples to apples)
    wo_col = "Wide Offer >3mm"
    dod_wo_col = "DoD Chg Wide Offer >3mm"
    
    if wo_col in df_filtered.columns:
        # Filter to rows where Wide Offer >3mm has a value on the last date
        df_filtered = df_filtered[df_filtered[wo_col].notna()].copy()
    
    if dod_wo_col in df_filtered.columns:
        # Filter to rows where DoD Chg Wide Offer >3mm exists (implies it had value on both dates)
        df_filtered = df_filtered[df_filtered[dod_wo_col].notna()].copy()
    
    # Ensure sort column is numeric (convert to numeric, coerce errors to NaN)
    if sort_column in df_filtered.columns:
        df_filtered[sort_column] = pd.to_numeric(df_filtered[sort_column], errors='coerce')
        
        # Filter out rows where sort column is 0 or blank (NaN/null)
        # Use numeric comparison: not NaN and not equal to 0
        df_filtered = df_filtered[
            df_filtered[sort_column].notna() & (df_filtered[sort_column] != 0)
        ].copy()
    
    # Sort by sort column descending (largest to smallest)
    if sort_column in df_filtered.columns:
        df_filtered = df_filtered.sort_values(sort_column, ascending=False, na_position='last')
    
    # Get top N and bottom N rows
    total_rows = len(df_filtered)
    if total_rows == 0:
        # Return empty DataFrame with correct columns
        available_columns = [col for col in columns if col in df.columns]
        return pd.DataFrame(columns=available_columns)
    
    if total_rows <= (2 * top_bottom_n):
        # If we have less than or equal to 2*N rows, show all
        result_df = df_filtered.copy()
    else:
        # Get top N (largest positive values) - first N rows after descending sort
        top_n = df_filtered.head(top_bottom_n).copy()
        
        # Get bottom N (most negative values) - last N rows after descending sort
        # Since we sorted descending, tail() gives us the smallest values in ascending order
        # We need to reverse them to get descending order (most negative first)
        bottom_n = df_filtered.tail(top_bottom_n).copy()
        
        # Reverse the bottom N to get descending order (most negative first)
        # This ensures -5 comes before -4, -4 before -3, etc.
        bottom_n = bottom_n.iloc[::-1].reset_index(drop=True)
        
        # Combine top and bottom (top first, then bottom)
        result_df = pd.concat([top_n, bottom_n], ignore_index=True)
    
    # Select only required columns (in order)
    available_columns = [col for col in columns if col in result_df.columns]
    result_df = result_df[available_columns].copy()
    
    return result_df


def create_large_cr01_on_offer_table(
    df: pd.DataFrame,
    excluded_sectors: list[str] = None,
    columns: list[str] = None,
    cr01_threshold: float = 3000.0,
    bo_threshold: float = 3.0
) -> pd.DataFrame:
    """
    Create Large CR01 On Offer table.
    
    Filters out excluded Custom_Sector values, filters to rows where CR01 @ Wide Offer > threshold
    and Bid/Offer>3mm < threshold, and shows all matching rows (no top/bottom limit).
    
    Args:
        df: Input DataFrame from runs_today.csv.
        excluded_sectors: List of Custom_Sector values to exclude (defaults to UNIVERSE_EXCLUDED_SECTORS).
        columns: List of columns to include (defaults to UNIVERSE_DOD_MOVES_COLUMNS).
        cr01_threshold: Minimum CR01 @ Wide Offer value (defaults to 3000.0).
        bo_threshold: Maximum Bid/Offer>3mm value (defaults to 3.0).
    
    Returns:
        Filtered DataFrame with selected columns, sorted by Bid/Offer>3mm ascending (low to high).
    """
    # Use defaults from config if not provided
    if excluded_sectors is None:
        excluded_sectors = UNIVERSE_EXCLUDED_SECTORS
    if columns is None:
        columns = UNIVERSE_DOD_MOVES_COLUMNS
    
    # Start with copy of input DataFrame
    df_filtered = df.copy()
    
    # Filter out excluded Custom_Sector values
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | 
            (~df_filtered[custom_sector_col].isin(excluded_sectors))
        ].copy()
    
    # Filter to rows where CR01 @ Wide Offer > threshold
    cr01_wo_col = "CR01 @ Wide Offer"
    if cr01_wo_col in df_filtered.columns:
        # Convert to numeric and filter to > threshold
        df_filtered[cr01_wo_col] = pd.to_numeric(df_filtered[cr01_wo_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[cr01_wo_col].notna() & (df_filtered[cr01_wo_col] > cr01_threshold)
        ].copy()
    
    # Filter to rows where Bid/Offer>3mm < threshold
    bo_3mm_col = "Bid/Offer>3mm"
    if bo_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to < threshold
        df_filtered[bo_3mm_col] = pd.to_numeric(df_filtered[bo_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[bo_3mm_col].notna() & (df_filtered[bo_3mm_col] < bo_threshold)
        ].copy()
    
    # Sort by Bid/Offer>3mm ascending (low to high)
    if bo_3mm_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(bo_3mm_col, ascending=True, na_position='last')
    
    # Select only required columns (in order)
    available_columns = [col for col in columns if col in df_filtered.columns]
    result_df = df_filtered[available_columns].copy()
    
    return result_df


def create_large_cr01_on_offer_no_longs_table(
    df: pd.DataFrame,
    excluded_sectors: list[str] = None,
    columns: list[str] = None,
    cr01_threshold: float = 3000.0,
    bo_threshold: float = 3.0,
    yrs_cvn_threshold: float = 11.0
) -> pd.DataFrame:
    """
    Create Large CR01 On Offer, No Longs table.
    
    Filters out excluded Custom_Sector values, filters to rows where CR01 @ Wide Offer > threshold,
    Bid/Offer>3mm < threshold, and Yrs (Cvn) < threshold, and shows all matching rows (no top/bottom limit).
    
    Args:
        df: Input DataFrame from runs_today.csv.
        excluded_sectors: List of Custom_Sector values to exclude (defaults to UNIVERSE_EXCLUDED_SECTORS).
        columns: List of columns to include (defaults to UNIVERSE_DOD_MOVES_COLUMNS).
        cr01_threshold: Minimum CR01 @ Wide Offer value (defaults to 3000.0).
        bo_threshold: Maximum Bid/Offer>3mm value (defaults to 3.0).
        yrs_cvn_threshold: Maximum Yrs (Cvn) value (defaults to 11.0).
    
    Returns:
        Filtered DataFrame with selected columns, sorted by Bid/Offer>3mm ascending (low to high).
    """
    # Use defaults from config if not provided
    if excluded_sectors is None:
        excluded_sectors = UNIVERSE_EXCLUDED_SECTORS
    if columns is None:
        columns = UNIVERSE_DOD_MOVES_COLUMNS
    
    # Start with copy of input DataFrame
    df_filtered = df.copy()
    
    # Filter out excluded Custom_Sector values
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | 
            (~df_filtered[custom_sector_col].isin(excluded_sectors))
        ].copy()
    
    # Filter to rows where CR01 @ Wide Offer > threshold
    cr01_wo_col = "CR01 @ Wide Offer"
    if cr01_wo_col in df_filtered.columns:
        # Convert to numeric and filter to > threshold
        df_filtered[cr01_wo_col] = pd.to_numeric(df_filtered[cr01_wo_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[cr01_wo_col].notna() & (df_filtered[cr01_wo_col] > cr01_threshold)
        ].copy()
    
    # Filter to rows where Bid/Offer>3mm < threshold
    bo_3mm_col = "Bid/Offer>3mm"
    if bo_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to < threshold
        df_filtered[bo_3mm_col] = pd.to_numeric(df_filtered[bo_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[bo_3mm_col].notna() & (df_filtered[bo_3mm_col] < bo_threshold)
        ].copy()
    
    # Filter to rows where Yrs (Cvn) < threshold
    yrs_cvn_col = "Yrs (Cvn)"
    if yrs_cvn_col in df_filtered.columns:
        # Convert to numeric and filter to < threshold
        df_filtered[yrs_cvn_col] = pd.to_numeric(df_filtered[yrs_cvn_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[yrs_cvn_col].notna() & (df_filtered[yrs_cvn_col] < yrs_cvn_threshold)
        ].copy()
    
    # Sort by Bid/Offer>3mm ascending (low to high)
    if bo_3mm_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(bo_3mm_col, ascending=True, na_position='last')
    
    # Select only required columns (in order)
    available_columns = [col for col in columns if col in df_filtered.columns]
    result_df = df_filtered[available_columns].copy()
    
    return result_df


def create_tough_to_find_offers_table(
    df: pd.DataFrame,
    excluded_sectors: list[str] = None,
    columns: list[str] = None,
    cr01_threshold: float = 2000.0,
    bo_threshold: float = 4.0,
    offers_threshold: float = 2.0
) -> pd.DataFrame:
    """
    Create Tough To Find Offers table.
    
    Filters out excluded Custom_Sector values, filters to rows where CR01 @ Wide Offer > threshold,
    Bid/Offer>3mm < threshold, and # Offers>3mm < threshold, and shows all matching rows (no top/bottom limit).
    
    Args:
        df: Input DataFrame from runs_today.csv.
        excluded_sectors: List of Custom_Sector values to exclude (defaults to UNIVERSE_EXCLUDED_SECTORS).
        columns: List of columns to include (defaults to UNIVERSE_DOD_MOVES_COLUMNS).
        cr01_threshold: Minimum CR01 @ Wide Offer value (defaults to 2000.0).
        bo_threshold: Maximum Bid/Offer>3mm value (defaults to 4.0).
        offers_threshold: Maximum # Offers>3mm value (defaults to 2.0).
    
    Returns:
        Filtered DataFrame with selected columns, sorted by Bid/Offer>3mm ascending (low to high).
    """
    # Use defaults from config if not provided
    if excluded_sectors is None:
        excluded_sectors = UNIVERSE_EXCLUDED_SECTORS
    if columns is None:
        columns = UNIVERSE_DOD_MOVES_COLUMNS
    
    # Start with copy of input DataFrame
    df_filtered = df.copy()
    
    # Filter out excluded Custom_Sector values
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | 
            (~df_filtered[custom_sector_col].isin(excluded_sectors))
        ].copy()
    
    # Filter to rows where CR01 @ Wide Offer > threshold
    cr01_wo_col = "CR01 @ Wide Offer"
    if cr01_wo_col in df_filtered.columns:
        # Convert to numeric and filter to > threshold
        df_filtered[cr01_wo_col] = pd.to_numeric(df_filtered[cr01_wo_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[cr01_wo_col].notna() & (df_filtered[cr01_wo_col] > cr01_threshold)
        ].copy()
    
    # Filter to rows where Bid/Offer>3mm < threshold
    bo_3mm_col = "Bid/Offer>3mm"
    if bo_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to < threshold
        df_filtered[bo_3mm_col] = pd.to_numeric(df_filtered[bo_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[bo_3mm_col].notna() & (df_filtered[bo_3mm_col] < bo_threshold)
        ].copy()
    
    # Filter to rows where # Offers>3mm < threshold
    offers_3mm_col = "# of Offers >3mm"
    if offers_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to < threshold
        df_filtered[offers_3mm_col] = pd.to_numeric(df_filtered[offers_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[offers_3mm_col].notna() & (df_filtered[offers_3mm_col] < offers_threshold)
        ].copy()
    
    # Sort by Bid/Offer>3mm ascending (low to high)
    if bo_3mm_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(bo_3mm_col, ascending=True, na_position='last')
    
    # Select only required columns (in order)
    available_columns = [col for col in columns if col in df_filtered.columns]
    result_df = df_filtered[available_columns].copy()
    
    return result_df


def create_carry_bonds_table(
    df: pd.DataFrame,
    excluded_sectors: list[str] = None,
    columns: list[str] = None,
    cr01_threshold: float = 500.0,
    bo_threshold: float = 6.0,
    yrs_cvn_threshold: float = 2.0,
    wo_3mm_threshold: float = 60.0
) -> pd.DataFrame:
    """
    Create Carry Bonds table.
    
    Filters out excluded Custom_Sector values, filters to rows where CR01 @ Wide Offer > threshold,
    Bid/Offer>3mm < threshold, Yrs (Cvn) < threshold, and Wide Offer >3mm > threshold,
    and shows all matching rows (no top/bottom limit).
    
    Args:
        df: Input DataFrame from runs_today.csv.
        excluded_sectors: List of Custom_Sector values to exclude (defaults to UNIVERSE_EXCLUDED_SECTORS).
        columns: List of columns to include (defaults to UNIVERSE_DOD_MOVES_COLUMNS).
        cr01_threshold: Minimum CR01 @ Wide Offer value (defaults to 500.0).
        bo_threshold: Maximum Bid/Offer>3mm value (defaults to 6.0).
        yrs_cvn_threshold: Maximum Yrs (Cvn) value (defaults to 2.0).
        wo_3mm_threshold: Minimum Wide Offer >3mm value (defaults to 60.0).
    
    Returns:
        Filtered DataFrame with selected columns, sorted by Wide Offer >3mm descending (high to low).
    """
    # Use defaults from config if not provided
    if excluded_sectors is None:
        excluded_sectors = UNIVERSE_EXCLUDED_SECTORS
    if columns is None:
        columns = UNIVERSE_DOD_MOVES_COLUMNS
    
    # Start with copy of input DataFrame
    df_filtered = df.copy()
    
    # Filter out excluded Custom_Sector values
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | 
            (~df_filtered[custom_sector_col].isin(excluded_sectors))
        ].copy()
    
    # Filter to rows where CR01 @ Wide Offer > threshold
    cr01_wo_col = "CR01 @ Wide Offer"
    if cr01_wo_col in df_filtered.columns:
        # Convert to numeric and filter to > threshold
        df_filtered[cr01_wo_col] = pd.to_numeric(df_filtered[cr01_wo_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[cr01_wo_col].notna() & (df_filtered[cr01_wo_col] > cr01_threshold)
        ].copy()
    
    # Filter to rows where Bid/Offer>3mm < threshold
    bo_3mm_col = "Bid/Offer>3mm"
    if bo_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to < threshold
        df_filtered[bo_3mm_col] = pd.to_numeric(df_filtered[bo_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[bo_3mm_col].notna() & (df_filtered[bo_3mm_col] < bo_threshold)
        ].copy()
    
    # Filter to rows where Yrs (Cvn) < threshold
    yrs_cvn_col = "Yrs (Cvn)"
    if yrs_cvn_col in df_filtered.columns:
        # Convert to numeric and filter to < threshold
        df_filtered[yrs_cvn_col] = pd.to_numeric(df_filtered[yrs_cvn_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[yrs_cvn_col].notna() & (df_filtered[yrs_cvn_col] < yrs_cvn_threshold)
        ].copy()
    
    # Filter to rows where Wide Offer >3mm > threshold
    wo_3mm_col = "Wide Offer >3mm"
    if wo_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to > threshold
        df_filtered[wo_3mm_col] = pd.to_numeric(df_filtered[wo_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[wo_3mm_col].notna() & (df_filtered[wo_3mm_col] > wo_3mm_threshold)
        ].copy()
    
    # Sort by Wide Offer >3mm descending (high to low)
    if wo_3mm_col in df_filtered.columns:
        df_filtered = df_filtered.sort_values(wo_3mm_col, ascending=False, na_position='last')
    
    # Select only required columns (in order)
    available_columns = [col for col in columns if col in df_filtered.columns]
    result_df = df_filtered[available_columns].copy()
    
    return result_df


def create_carry_bonds_sorted_by_mtd_table(
    df: pd.DataFrame,
    excluded_sectors: list[str] = None,
    columns: list[str] = None,
    cr01_threshold: float = 500.0,
    bo_threshold: float = 6.0,
    yrs_cvn_threshold: float = 2.0,
    wo_3mm_threshold: float = 60.0
) -> pd.DataFrame:
    """
    Create Carry Bonds Sorted by MTD Moves table.
    
    Filters out excluded Custom_Sector values, filters to rows where CR01 @ Wide Offer > threshold,
    Bid/Offer>3mm < threshold, Yrs (Cvn) < threshold, and Wide Offer >3mm > threshold,
    and shows all matching rows (no top/bottom limit).
    Excludes DoD columns (between # of Offers >3mm and MTD Chg Tight Bid).
    Sorted by MTD Chg Tight Bid descending (largest to smallest).
    
    Args:
        df: Input DataFrame from runs_today.csv.
        excluded_sectors: List of Custom_Sector values to exclude (defaults to UNIVERSE_EXCLUDED_SECTORS).
        columns: List of columns to include (defaults to UNIVERSE_MTD_MOVES_COLUMNS).
        cr01_threshold: Minimum CR01 @ Wide Offer value (defaults to 500.0).
        bo_threshold: Maximum Bid/Offer>3mm value (defaults to 6.0).
        yrs_cvn_threshold: Maximum Yrs (Cvn) value (defaults to 2.0).
        wo_3mm_threshold: Minimum Wide Offer >3mm value (defaults to 60.0).
    
    Returns:
        Filtered DataFrame with selected columns, sorted by MTD Chg Tight Bid descending (largest to smallest).
    """
    # Use defaults from config if not provided
    if excluded_sectors is None:
        excluded_sectors = UNIVERSE_EXCLUDED_SECTORS
    if columns is None:
        columns = UNIVERSE_MTD_MOVES_COLUMNS
    
    # Start with copy of input DataFrame
    df_filtered = df.copy()
    
    # Filter out excluded Custom_Sector values
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | 
            (~df_filtered[custom_sector_col].isin(excluded_sectors))
        ].copy()
    
    # Filter to rows where CR01 @ Wide Offer > threshold
    cr01_wo_col = "CR01 @ Wide Offer"
    if cr01_wo_col in df_filtered.columns:
        # Convert to numeric and filter to > threshold
        df_filtered[cr01_wo_col] = pd.to_numeric(df_filtered[cr01_wo_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[cr01_wo_col].notna() & (df_filtered[cr01_wo_col] > cr01_threshold)
        ].copy()
    
    # Filter to rows where Bid/Offer>3mm < threshold
    bo_3mm_col = "Bid/Offer>3mm"
    if bo_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to < threshold
        df_filtered[bo_3mm_col] = pd.to_numeric(df_filtered[bo_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[bo_3mm_col].notna() & (df_filtered[bo_3mm_col] < bo_threshold)
        ].copy()
    
    # Filter to rows where Yrs (Cvn) < threshold
    yrs_cvn_col = "Yrs (Cvn)"
    if yrs_cvn_col in df_filtered.columns:
        # Convert to numeric and filter to < threshold
        df_filtered[yrs_cvn_col] = pd.to_numeric(df_filtered[yrs_cvn_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[yrs_cvn_col].notna() & (df_filtered[yrs_cvn_col] < yrs_cvn_threshold)
        ].copy()
    
    # Filter to rows where Wide Offer >3mm > threshold
    wo_3mm_col = "Wide Offer >3mm"
    if wo_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to > threshold
        df_filtered[wo_3mm_col] = pd.to_numeric(df_filtered[wo_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[wo_3mm_col].notna() & (df_filtered[wo_3mm_col] > wo_3mm_threshold)
        ].copy()
    
    # Sort by MTD Chg Tight Bid descending (largest to smallest)
    mtd_tb_col = "MTD Chg Tight Bid"
    if mtd_tb_col in df_filtered.columns:
        df_filtered[mtd_tb_col] = pd.to_numeric(df_filtered[mtd_tb_col], errors='coerce')
        df_filtered = df_filtered.sort_values(mtd_tb_col, ascending=False, na_position='last')
    
    # Select only required columns (in order)
    available_columns = [col for col in columns if col in df_filtered.columns]
    result_df = df_filtered[available_columns].copy()
    
    return result_df


def create_universe_dod_moves_wo_table(
    df: pd.DataFrame,
    excluded_sectors: list[str] = None,
    sort_column: str = None,
    top_bottom_n: int = None,
    columns: list[str] = None
) -> pd.DataFrame:
    """
    Create Universe Sorted By DoD Moves table.
    
    Filters out excluded Custom_Sector values, excludes rows where sort column is 0 or blank,
    and shows top N and bottom N rows by sort column (largest positive and most negative values).
    Uses regular Wide Offer (not Wide Offer >3mm) for filtering and sorting.
    
    Args:
        df: Input DataFrame from runs_today.csv.
        excluded_sectors: List of Custom_Sector values to exclude (defaults to UNIVERSE_EXCLUDED_SECTORS).
        sort_column: Column to sort by (defaults to "DoD Chg Wide Offer").
        top_bottom_n: Number of top and bottom rows to show (defaults to UNIVERSE_TOP_BOTTOM_N).
        columns: List of columns to include (defaults to UNIVERSE_DOD_MOVES_COLUMNS).
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Use defaults from config if not provided
    if excluded_sectors is None:
        excluded_sectors = UNIVERSE_EXCLUDED_SECTORS
    if sort_column is None:
        sort_column = "DoD Chg Wide Offer"
    if top_bottom_n is None:
        top_bottom_n = UNIVERSE_TOP_BOTTOM_N
    if columns is None:
        columns = UNIVERSE_DOD_MOVES_COLUMNS
    
    # Start with copy of input DataFrame
    df_filtered = df.copy()
    
    # Filter out excluded Custom_Sector values
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | 
            (~df_filtered[custom_sector_col].isin(excluded_sectors))
        ].copy()
    
    # Filter to rows that had "Wide Offer" on both last date and second-to-last date
    # This ensures DoD changes are meaningful (comparing apples to apples)
    wo_col = "Wide Offer"
    dod_wo_col = "DoD Chg Wide Offer"
    
    if wo_col in df_filtered.columns:
        # Filter to rows where Wide Offer has a value on the last date
        df_filtered = df_filtered[df_filtered[wo_col].notna()].copy()
    
    if dod_wo_col in df_filtered.columns:
        # Filter to rows where DoD Chg Wide Offer exists (implies it had value on both dates)
        df_filtered = df_filtered[df_filtered[dod_wo_col].notna()].copy()
    
    # Ensure sort column is numeric (convert to numeric, coerce errors to NaN)
    if sort_column in df_filtered.columns:
        df_filtered[sort_column] = pd.to_numeric(df_filtered[sort_column], errors='coerce')
        
        # Filter out rows where sort column is 0 or blank (NaN/null)
        # Use numeric comparison: not NaN and not equal to 0
        df_filtered = df_filtered[
            df_filtered[sort_column].notna() & (df_filtered[sort_column] != 0)
        ].copy()
    
    # Sort by sort column descending (largest to smallest)
    if sort_column in df_filtered.columns:
        df_filtered = df_filtered.sort_values(sort_column, ascending=False, na_position='last')
    
    # Get top N and bottom N rows
    total_rows = len(df_filtered)
    if total_rows == 0:
        # Return empty DataFrame with correct columns
        available_columns = [col for col in columns if col in df.columns]
        return pd.DataFrame(columns=available_columns)
    
    if total_rows <= (2 * top_bottom_n):
        # If we have less than or equal to 2*N rows, show all
        result_df = df_filtered.copy()
    else:
        # Get top N (largest positive values) - first N rows after descending sort
        top_n = df_filtered.head(top_bottom_n).copy()
        
        # Get bottom N (most negative values) - last N rows after descending sort
        # Since we sorted descending, tail() gives us the smallest values in ascending order
        # We need to reverse them to get descending order (most negative first)
        bottom_n = df_filtered.tail(top_bottom_n).copy()
        
        # Reverse the bottom N to get descending order (most negative first)
        # This ensures -5 comes before -4, -4 before -3, etc.
        bottom_n = bottom_n.iloc[::-1].reset_index(drop=True)
        
        # Combine top and bottom (top first, then bottom)
        result_df = pd.concat([top_n, bottom_n], ignore_index=True)
    
    # Select only required columns (in order)
    available_columns = [col for col in columns if col in result_df.columns]
    result_df = result_df[available_columns].copy()
    
    return result_df


def create_universe_mtd_moves_table(
    df: pd.DataFrame,
    excluded_sectors: list[str] = None,
    sort_column: str = None,
    top_bottom_n: int = None,
    columns: list[str] = None
) -> pd.DataFrame:
    """
    Create Universe Sorted By MTD Moves With Size On Offer >3mm table.
    
    Filters out excluded Custom_Sector values, filters to rows where Wide Offer >3mm > 0,
    excludes rows where sort column is 0 or blank, and shows top N and bottom N rows by sort column.
    Excludes DoD columns (between # of Offers >3mm and MTD Chg Tight Bid).
    
    Args:
        df: Input DataFrame from runs_today.csv.
        excluded_sectors: List of Custom_Sector values to exclude (defaults to UNIVERSE_EXCLUDED_SECTORS).
        sort_column: Column to sort by (defaults to "MTD Chg Tight Bid").
        top_bottom_n: Number of top and bottom rows to show (defaults to UNIVERSE_TOP_BOTTOM_N).
        columns: List of columns to include (defaults to UNIVERSE_MTD_MOVES_COLUMNS).
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Use defaults from config if not provided
    if excluded_sectors is None:
        excluded_sectors = UNIVERSE_EXCLUDED_SECTORS
    if sort_column is None:
        sort_column = "MTD Chg Tight Bid"
    if top_bottom_n is None:
        top_bottom_n = UNIVERSE_TOP_BOTTOM_N
    if columns is None:
        columns = UNIVERSE_MTD_MOVES_COLUMNS
    
    # Start with copy of input DataFrame
    df_filtered = df.copy()
    
    # Filter out excluded Custom_Sector values
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | 
            (~df_filtered[custom_sector_col].isin(excluded_sectors))
        ].copy()
    
    # Filter to rows where Wide Offer >3mm > 0
    wo_3mm_col = "Wide Offer >3mm"
    if wo_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to > 0
        df_filtered[wo_3mm_col] = pd.to_numeric(df_filtered[wo_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[wo_3mm_col].notna() & (df_filtered[wo_3mm_col] > 0)
        ].copy()
    
    # Ensure sort column is numeric (convert to numeric, coerce errors to NaN)
    if sort_column in df_filtered.columns:
        df_filtered[sort_column] = pd.to_numeric(df_filtered[sort_column], errors='coerce')
        
        # Filter out rows where sort column is 0 or blank (NaN/null)
        # Use numeric comparison: not NaN and not equal to 0
        df_filtered = df_filtered[
            df_filtered[sort_column].notna() & (df_filtered[sort_column] != 0)
        ].copy()
    
    # Sort by sort column descending (largest to smallest)
    if sort_column in df_filtered.columns:
        df_filtered = df_filtered.sort_values(sort_column, ascending=False, na_position='last')
    
    # Get top N and bottom N rows
    total_rows = len(df_filtered)
    if total_rows == 0:
        # Return empty DataFrame with correct columns
        available_columns = [col for col in columns if col in df.columns]
        return pd.DataFrame(columns=available_columns)
    
    if total_rows <= (2 * top_bottom_n):
        # If we have less than or equal to 2*N rows, show all
        result_df = df_filtered.copy()
    else:
        # Get top N (largest positive values) - first N rows after descending sort
        top_n = df_filtered.head(top_bottom_n).copy()
        
        # Get bottom N (most negative values) - last N rows after descending sort
        # Since we sorted descending, tail() gives us the smallest values in ascending order
        # We need to reverse them to get descending order (most negative first)
        bottom_n = df_filtered.tail(top_bottom_n).copy()
        
        # Reverse the bottom N to get descending order (most negative first)
        # This ensures -5 comes before -4, -4 before -3, etc.
        bottom_n = bottom_n.iloc[::-1].reset_index(drop=True)
        
        # Combine top and bottom (top first, then bottom)
        result_df = pd.concat([top_n, bottom_n], ignore_index=True)
    
    # Select only required columns (in order)
    available_columns = [col for col in columns if col in result_df.columns]
    result_df = result_df[available_columns].copy()
    
    return result_df


def create_universe_ytd_moves_table(
    df: pd.DataFrame,
    excluded_sectors: list[str] = None,
    sort_column: str = None,
    top_bottom_n: int = None,
    columns: list[str] = None
) -> pd.DataFrame:
    """
    Create Universe Sorted By YTD Moves With Size On Offer >3mm table.
    
    Filters out excluded Custom_Sector values, filters to rows where Wide Offer >3mm > 0,
    excludes rows where sort column is 0 or blank, and shows top N and bottom N rows by sort column.
    Excludes DoD columns (between # of Offers >3mm and MTD Chg Tight Bid).
    
    Args:
        df: Input DataFrame from runs_today.csv.
        excluded_sectors: List of Custom_Sector values to exclude (defaults to UNIVERSE_EXCLUDED_SECTORS).
        sort_column: Column to sort by (defaults to "YTD Chg Tight Bid").
        top_bottom_n: Number of top and bottom rows to show (defaults to UNIVERSE_TOP_BOTTOM_N).
        columns: List of columns to include (defaults to UNIVERSE_MTD_MOVES_COLUMNS).
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Use defaults from config if not provided
    if excluded_sectors is None:
        excluded_sectors = UNIVERSE_EXCLUDED_SECTORS
    if sort_column is None:
        sort_column = "YTD Chg Tight Bid"
    if top_bottom_n is None:
        top_bottom_n = UNIVERSE_TOP_BOTTOM_N
    if columns is None:
        columns = UNIVERSE_MTD_MOVES_COLUMNS
    
    # Start with copy of input DataFrame
    df_filtered = df.copy()
    
    # Filter out excluded Custom_Sector values
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | 
            (~df_filtered[custom_sector_col].isin(excluded_sectors))
        ].copy()
    
    # Filter to rows where Wide Offer >3mm > 0
    wo_3mm_col = "Wide Offer >3mm"
    if wo_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to > 0
        df_filtered[wo_3mm_col] = pd.to_numeric(df_filtered[wo_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[wo_3mm_col].notna() & (df_filtered[wo_3mm_col] > 0)
        ].copy()
    
    # Ensure sort column is numeric (convert to numeric, coerce errors to NaN)
    if sort_column in df_filtered.columns:
        df_filtered[sort_column] = pd.to_numeric(df_filtered[sort_column], errors='coerce')
        
        # Filter out rows where sort column is 0 or blank (NaN/null)
        # Use numeric comparison: not NaN and not equal to 0
        df_filtered = df_filtered[
            df_filtered[sort_column].notna() & (df_filtered[sort_column] != 0)
        ].copy()
    
    # Sort by sort column descending (largest to smallest)
    if sort_column in df_filtered.columns:
        df_filtered = df_filtered.sort_values(sort_column, ascending=False, na_position='last')
    
    # Get top N and bottom N rows
    total_rows = len(df_filtered)
    if total_rows == 0:
        # Return empty DataFrame with correct columns
        available_columns = [col for col in columns if col in df.columns]
        return pd.DataFrame(columns=available_columns)
    
    if total_rows <= (2 * top_bottom_n):
        # If we have less than or equal to 2*N rows, show all
        result_df = df_filtered.copy()
    else:
        # Get top N (largest positive values) - first N rows after descending sort
        top_n = df_filtered.head(top_bottom_n).copy()
        
        # Get bottom N (most negative values) - last N rows after descending sort
        # Since we sorted descending, tail() gives us the smallest values in ascending order
        # We need to reverse them to get descending order (most negative first)
        bottom_n = df_filtered.tail(top_bottom_n).copy()
        
        # Reverse the bottom N to get descending order (most negative first)
        # This ensures -5 comes before -4, -4 before -3, etc.
        bottom_n = bottom_n.iloc[::-1].reset_index(drop=True)
        
        # Combine top and bottom (top first, then bottom)
        result_df = pd.concat([top_n, bottom_n], ignore_index=True)
    
    # Select only required columns (in order)
    available_columns = [col for col in columns if col in result_df.columns]
    result_df = result_df[available_columns].copy()
    
    return result_df


def create_universe_1yr_moves_table(
    df: pd.DataFrame,
    excluded_sectors: list[str] = None,
    sort_column: str = None,
    top_bottom_n: int = None,
    columns: list[str] = None
) -> pd.DataFrame:
    """
    Create Universe Sorted By 1yr Moves With Size On Offer >3mm table.
    
    Filters out excluded Custom_Sector values, filters to rows where Wide Offer >3mm > 0,
    excludes rows where sort column is 0 or blank, and shows top N and bottom N rows by sort column.
    Excludes DoD columns (between # of Offers >3mm and MTD Chg Tight Bid).
    
    Args:
        df: Input DataFrame from runs_today.csv.
        excluded_sectors: List of Custom_Sector values to exclude (defaults to UNIVERSE_EXCLUDED_SECTORS).
        sort_column: Column to sort by (defaults to "1yr Chg Tight Bid").
        top_bottom_n: Number of top and bottom rows to show (defaults to UNIVERSE_TOP_BOTTOM_N).
        columns: List of columns to include (defaults to UNIVERSE_MTD_MOVES_COLUMNS).
    
    Returns:
        Filtered and sorted DataFrame with selected columns.
    """
    # Use defaults from config if not provided
    if excluded_sectors is None:
        excluded_sectors = UNIVERSE_EXCLUDED_SECTORS
    if sort_column is None:
        sort_column = "1yr Chg Tight Bid"
    if top_bottom_n is None:
        top_bottom_n = UNIVERSE_TOP_BOTTOM_N
    if columns is None:
        columns = UNIVERSE_MTD_MOVES_COLUMNS
    
    # Start with copy of input DataFrame
    df_filtered = df.copy()
    
    # Filter out excluded Custom_Sector values
    custom_sector_col = "Custom_Sector"
    if custom_sector_col in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[custom_sector_col].isna() | 
            (~df_filtered[custom_sector_col].isin(excluded_sectors))
        ].copy()
    
    # Filter to rows where Wide Offer >3mm > 0
    wo_3mm_col = "Wide Offer >3mm"
    if wo_3mm_col in df_filtered.columns:
        # Convert to numeric and filter to > 0
        df_filtered[wo_3mm_col] = pd.to_numeric(df_filtered[wo_3mm_col], errors='coerce')
        df_filtered = df_filtered[
            df_filtered[wo_3mm_col].notna() & (df_filtered[wo_3mm_col] > 0)
        ].copy()
    
    # Ensure sort column is numeric (convert to numeric, coerce errors to NaN)
    if sort_column in df_filtered.columns:
        df_filtered[sort_column] = pd.to_numeric(df_filtered[sort_column], errors='coerce')
        
        # Filter out rows where sort column is 0 or blank (NaN/null)
        # Use numeric comparison: not NaN and not equal to 0
        df_filtered = df_filtered[
            df_filtered[sort_column].notna() & (df_filtered[sort_column] != 0)
        ].copy()
    
    # Sort by sort column descending (largest to smallest)
    if sort_column in df_filtered.columns:
        df_filtered = df_filtered.sort_values(sort_column, ascending=False, na_position='last')
    
    # Get top N and bottom N rows
    total_rows = len(df_filtered)
    if total_rows == 0:
        # Return empty DataFrame with correct columns
        available_columns = [col for col in columns if col in df.columns]
        return pd.DataFrame(columns=available_columns)
    
    if total_rows <= (2 * top_bottom_n):
        # If we have less than or equal to 2*N rows, show all
        result_df = df_filtered.copy()
    else:
        # Get top N (largest positive values) - first N rows after descending sort
        top_n = df_filtered.head(top_bottom_n).copy()
        
        # Get bottom N (most negative values) - last N rows after descending sort
        # Since we sorted descending, tail() gives us the smallest values in ascending order
        # We need to reverse them to get descending order (most negative first)
        bottom_n = df_filtered.tail(top_bottom_n).copy()
        
        # Reverse the bottom N to get descending order (most negative first)
        # This ensures -5 comes before -4, -4 before -3, etc.
        bottom_n = bottom_n.iloc[::-1].reset_index(drop=True)
        
        # Combine top and bottom (top first, then bottom)
        result_df = pd.concat([top_n, bottom_n], ignore_index=True)
    
    # Select only required columns (in order)
    available_columns = [col for col in columns if col in result_df.columns]
    result_df = result_df[available_columns].copy()
    
    return result_df


def write_excel_file(
    output_path: Path,
    tables: Dict[str, Dict],
    timestamp: str,
    last_date,
    mtd_ref_date,
    ytd_ref_date
) -> None:
    # Reset used sheet names for each file generation
    write_excel_file._used_sheet_names = set()
    """
    Write all portfolio tables to Excel file with formatted tables on separate sheets.
    
    Args:
        output_path: Path to Excel output file.
        tables: Dictionary mapping table titles to dicts with 'df' (DataFrame) and 'summary' (dict).
        timestamp: Timestamp string for header.
        last_date: Last date from runs data.
        mtd_ref_date: MTD reference date.
        ytd_ref_date: YTD reference date.
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for table_title, table_data in tables.items():
            df = table_data['df']
            summary_dict = table_data.get('summary', {})
            
            if df.empty:
                continue  # Skip empty DataFrames
            
            # Get sheet name (max 31 characters for Excel, use full title where possible)
            # Remove invalid characters for Excel sheet names: / \ ? * [ ] :
            sanitized_title = table_title.replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '').replace(':', '-')
            
            # Initialize dealer_match for dealer-specific tables
            dealer_match = ''
            if 'Where' in sanitized_title:
                # Extract dealer name from "Where {Dealer} Is..."
                parts = sanitized_title.split('Where ')
                if len(parts) > 1:
                    dealer_part = parts[1].split(' Is')[0] if ' Is' in parts[1] else ''
                    dealer_match = dealer_part.strip() if dealer_part else ''
            
            # For dealer tables, use shorter unique names
            if 'Where' in sanitized_title and ('Best Bid' in sanitized_title or 'Wide Offer' in sanitized_title):
                # Extract dealer name and create short unique sheet name
                if dealer_match:
                    if 'Best Bid' in sanitized_title:
                        # Format: "Min BO - {Dealer}" (shorter to ensure uniqueness)
                        sheet_name = f"Min BO - {dealer_match}"[:31]
                    elif 'Wide Offer' in sanitized_title:
                        # Format: "CR01 WO - {Dealer}" for Large CR01 On Offer dealer tables
                        sheet_name = f"CR01 WO - {dealer_match}"[:31]
                    else:
                        sheet_name = sanitized_title[:31] if len(sanitized_title) <= 31 else sanitized_title[:28] + "..."
                else:
                    sheet_name = sanitized_title[:31] if len(sanitized_title) <= 31 else sanitized_title[:28] + "..."
            else:
                sheet_name = sanitized_title[:31] if len(sanitized_title) <= 31 else sanitized_title[:28] + "..."
            
            # Track used sheet names to ensure uniqueness
            if not hasattr(write_excel_file, '_used_sheet_names'):
                write_excel_file._used_sheet_names = set()
            
            # If sheet name already used, append dealer name or number
            original_sheet_name = sheet_name
            counter = 1
            while sheet_name in write_excel_file._used_sheet_names:
                if 'Where' in sanitized_title and dealer_match:
                    # Try even shorter name
                    if 'Best Bid' in sanitized_title:
                        sheet_name = f"BO - {dealer_match}"[:31]
                        if sheet_name in write_excel_file._used_sheet_names:
                            sheet_name = f"{dealer_match} Best Bid"[:31]
                    elif 'Wide Offer' in sanitized_title:
                        sheet_name = f"WO - {dealer_match}"[:31]
                        if sheet_name in write_excel_file._used_sheet_names:
                            sheet_name = f"{dealer_match} WO"[:31]
                    else:
                        sheet_name = f"{original_sheet_name[:27]}_{counter}"[:31]
                else:
                    # Append number for other duplicates
                    sheet_name = f"{original_sheet_name[:27]}_{counter}"[:31]
                counter += 1
            
            write_excel_file._used_sheet_names.add(sheet_name)
            
            # Keep original DataFrame with numeric types for Excel
            # We'll apply formatting via openpyxl cell formatting (not string conversion)
            df_excel = df.copy()
            
            # Replace NaN with None (Excel will display as empty)
            df_excel = df_excel.where(pd.notna(df_excel), None)
            
            # Ensure ASCII-safe strings for text columns only
            for col in df_excel.columns:
                if df_excel[col].dtype == 'object':
                    df_excel[col] = df_excel[col].apply(
                        lambda x: ensure_ascii(str(x)) if x is not None and x != "" else None
                    )
            
            # Calculate start row for DataFrame (after summary statistics if any)
            summary_rows = len(summary_dict) if summary_dict else 0
            start_row = summary_rows
            
            # Write DataFrame to Excel sheet (starting after summary rows)
            # Keep numeric types - we'll format via openpyxl
            df_excel.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
            
            # Get the worksheet
            worksheet = writer.sheets[sheet_name]
            
            # Apply Excel number formatting to numeric columns
            # Find numeric columns and apply appropriate formats
            for col_idx, col_name in enumerate(df_excel.columns, start=1):
                if pd.api.types.is_numeric_dtype(df_excel[col_name]):
                    # Determine format based on column name
                    if col_name == "Retracement":
                        # Percentage format: 0.5 -> 50.00%
                        number_format = "0.00%"
                    elif col_name == "Yrs (Cvn)":
                        # One decimal place
                        number_format = "0.0"
                    else:
                        # Thousand separators, no decimals (rounded)
                        number_format = "#,##0"
                    
                    # Apply format to all data cells in this column
                    # Data starts at start_row + 2 (start_row is 0-indexed for DataFrame, +1 for header row, +1 for 1-indexed Excel)
                    data_start_row = start_row + 2  # Header row is at start_row + 1, data starts at start_row + 2
                    data_end_row = start_row + len(df_excel) + 1  # Last data row
                    for row_idx in range(data_start_row, data_end_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        # Apply format if cell has a numeric value
                        if cell.value is not None:
                            try:
                                # Check if it's a number (int, float, or numeric string)
                                float(cell.value)
                                cell.number_format = number_format
                            except (ValueError, TypeError):
                                # Not a number, skip formatting
                                pass
            
            # Write summary statistics FIRST (before creating table)
            # This ensures summary rows are in place before table validation
            if summary_dict:
                # Write summary statistics starting at row 1
                for idx, (key, value) in enumerate(summary_dict.items(), start=1):
                    # Write summary text to first cell
                    # Keep numeric values as numbers, format display via Excel formatting
                    cell = worksheet.cell(row=idx, column=1)
                    if isinstance(value, (int, float)):
                        # Store as number with thousand separator format
                        cell.value = value
                        cell.number_format = "#,##0"
                        # Update cell to show label + formatted value
                        # We'll use a formula or keep it as text with formatted number
                        # Actually, let's keep it simple: text label + formatted number
                        formatted_value = f"{int(round(value)):,}"
                        cell.value = f"{key}: {formatted_value}"
                    else:
                        cell.value = f"{key}: {str(value)}"
            
            # Calculate start row for table (header row)
            start_row_table = summary_rows + 1 if summary_rows > 0 else 1
            
            # Create Excel table AFTER summary rows are written
            # Table name must be unique and valid (no spaces, starts with letter)
            sanitized_name = table_title.replace(' ', '_').replace('-', '_').replace('.', '_').replace(':', '_')
            sanitized_name = ''.join(c if c.isalnum() or c == '_' else '' for c in sanitized_name)
            if sanitized_name and not sanitized_name[0].isalpha():
                sanitized_name = 'T_' + sanitized_name
            table_name = f"Table_{sanitized_name}"[:255]  # Excel table name max length is 255
            
            # Table range includes header + data (summary rows are above the table)
            # Excel tables: range includes header row + all data rows
            table_start_row = start_row_table  # Header row (1-indexed)
            # End row = header row + number of data rows
            # Note: len(df_excel) is number of data rows, header is at table_start_row
            table_end_row = table_start_row + len(df_excel)  # Last data row
            table_range = f"A{table_start_row}:{get_column_letter(len(df_excel.columns))}{table_end_row}"
            
            # Create table - ensure range is valid and doesn't include summary rows
            # Verify table range doesn't overlap with summary rows
            if table_start_row > summary_rows:
                table = Table(displayName=table_name, ref=table_range)
                
                # Apply table style with banded rows and filters
                style = TableStyleInfo(
                    name="TableStyleMedium9",  # Medium style with banded rows
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,  # Banded rows
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                
                # Add table to worksheet
                worksheet.add_table(table)
            else:
                print(f"Warning: Skipping table for '{sheet_name}' - invalid table range")
            
            # Find Security column index for freeze panes
            security_col_idx = None
            for idx, col in enumerate(df_excel.columns, start=1):
                if col == "Security":
                    security_col_idx = idx
                    break
            
            # Freeze panes after Security column (if Security column exists)
            # Freeze at row after summary rows + header row
            if security_col_idx is not None:
                # Freeze after Security column, starting at first data row (after summary + header)
                freeze_row = start_row_table + 1  # +1 for first data row (header is at start_row_table)
                freeze_cell = f"{get_column_letter(security_col_idx + 1)}{freeze_row}"
                worksheet.freeze_panes = freeze_cell
            
            # Auto-fit column widths
            for idx, col in enumerate(df_excel.columns, start=1):
                column_letter = get_column_letter(idx)
                # Calculate max width (content + padding)
                # For numeric columns, estimate width based on formatted display
                if pd.api.types.is_numeric_dtype(df_excel[col]):
                    # Estimate width for formatted numbers
                    sample_values = df_excel[col].dropna().head(10)
                    if len(sample_values) > 0:
                        if col == "Retracement":
                            max_length = max(len(f"{v*100:.2f}%") for v in sample_values if pd.notna(v))
                        elif col == "Yrs (Cvn)":
                            max_length = max(len(f"{v:.1f}") for v in sample_values if pd.notna(v))
                        else:
                            max_length = max(len(f"{int(round(v)):,}") for v in sample_values if pd.notna(v))
                    else:
                        max_length = len(str(col))
                else:
                    max_length = max(
                        len(str(col)),  # Header length
                        df_excel[col].astype(str).map(len).max() if len(df_excel) > 0 else 0
                    )
                # Set width with some padding (min 10, max 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"\nExcel file written to: {output_path}")
    print(f"Total sheets: {sum(1 for t in tables.values() if not t['df'].empty)}")


def main() -> None:
    """Main execution function."""
    print("="*100)
    print("RUNS VIEWS - Portfolio Runs View Generator")
    print("="*100)
    
    # Step 0: Validate runs_today.csv is current with parquet (auto-regenerate if stale)
    print("\n[STEP 0] Validating runs_today.csv is current with runs_timeseries.parquet...")
    parquet_last_date = validate_runs_today_csv_is_current(
        RUNS_TODAY_CSV_PATH,
        RUNS_PARQUET_PATH,
        auto_regenerate=True
    )
    print(f"  Parquet last date: {parquet_last_date}")
    
    # Step 1: Load runs_today.csv
    print("\n[STEP 1] Loading runs_today.csv...")
    df = pd.read_csv(RUNS_TODAY_CSV_PATH)
    
    # Convert all numeric columns to proper types (ensures filtering works correctly)
    df = ensure_numeric_types(df)
    
    print(f"Loaded {len(df):,} rows, {len(df.columns)} columns")
    
    # Step 2: Get reference dates from runs_timeseries.parquet
    print("\n[STEP 2] Extracting reference dates from runs_timeseries.parquet...")
    ref_dates = get_reference_dates(RUNS_PARQUET_PATH)
    last_date = ref_dates["last_date"]
    mtd_ref_date = ref_dates["mtd_ref_date"]
    ytd_ref_date = ref_dates["ytd_ref_date"]
    
    print(f"Last date: {last_date}")
    print(f"MTD reference date: {mtd_ref_date}")
    print(f"YTD reference date: {ytd_ref_date}")
    
    # Step 3: Create Portfolio Sorted By CR01 Risk table
    print("\n[STEP 3] Creating Portfolio Sorted By CR01 Risk table...")
    
    # Create the table first
    portfolio_cr01_df = create_portfolio_cr01_risk_table(df)
    print(f"Filtered to {len(portfolio_cr01_df):,} rows with QUANTITY > 0")
    
    # Calculate Total CR01 from the actual displayed table rows (sum of POSITION CR01)
    pos_cr01_col = "POSITION CR01"
    total_cr01 = 0
    if pos_cr01_col in portfolio_cr01_df.columns:
        total_cr01 = portfolio_cr01_df[pos_cr01_col].sum()
    print(f"Total CR01: {int(round(total_cr01)):,}")
    
    # Step 3.5: Create Portfolio Less Liquid Lines table
    print("\n[STEP 3.5] Creating Portfolio Less Liquid Lines table...")
    
    # Create the table first
    portfolio_less_liquid_df = create_portfolio_less_liquid_lines_table(df)
    print(f"Filtered to {len(portfolio_less_liquid_df):,} rows with QUANTITY > 0 and TB >3mm is blank")
    
    # Calculate Total CR01 from the actual displayed table rows (sum of POSITION CR01)
    total_cr01_less_liquid = 0
    if pos_cr01_col in portfolio_less_liquid_df.columns:
        total_cr01_less_liquid = portfolio_less_liquid_df[pos_cr01_col].sum()
    print(f"Total CR01: {int(round(total_cr01_less_liquid)):,}")
    
    # Step 4: Create Portfolio Sorted By DoD Bid Chg table
    print("\n[STEP 4] Creating Portfolio Sorted By DoD Bid Chg With >3MM on Bid table...")
    portfolio_dod_bid_df = create_portfolio_dod_bid_chg_table(df)
    print(f"Filtered to {len(portfolio_dod_bid_df):,} rows with QUANTITY > 0, TB >3mm has value, and DoD TB>3mm non-zero")
    
    # Step 5: Create Portfolio Sorted By MTD Bid Chg table
    print("\n[STEP 5] Creating Portfolio Sorted By MTD Bid Chg With >3MM on Bid table...")
    portfolio_mtd_bid_df = create_portfolio_mtd_bid_chg_table(df)
    print(f"Filtered to {len(portfolio_mtd_bid_df):,} rows with QUANTITY > 0, TB >3mm has value, and MTD TB non-zero")
    
    # Step 6: Create Portfolio Sorted By YTD Bid Chg table
    print("\n[STEP 6] Creating Portfolio Sorted By YTD Bid Chg With >3MM on Bid table...")
    portfolio_ytd_bid_df = create_portfolio_ytd_bid_chg_table(df)
    print(f"Filtered to {len(portfolio_ytd_bid_df):,} rows with QUANTITY > 0, TB >3mm has value, and YTD TB non-zero")
    
    # Step 7: Create Portfolio Sorted By 1yr Bid Chg table
    print("\n[STEP 7] Creating Portfolio Sorted By 1yr Bid Chg With >3MM on Bid table...")
    portfolio_1yr_bid_df = create_portfolio_1yr_bid_chg_table(df)
    print(f"Filtered to {len(portfolio_1yr_bid_df):,} rows with QUANTITY > 0, TB >3mm has value, and 1yr TB non-zero")
    
    # Step 8: Create Size Bids table
    print("\n[STEP 8] Creating Size Bids table...")
    size_bids_df = create_size_bids_table(df)
    print(f"Filtered to {len(size_bids_df):,} rows with QUANTITY > 0, CR01 TB >= 1000, and POSITION CR01 >= 1,000")
    
    # Calculate cumulative CR01 TB for Size Bids table
    cumulative_cr01_tb = 0
    cr01_tb_col = "CR01 @ Tight Bid"
    if cr01_tb_col in size_bids_df.columns:
        cumulative_cr01_tb = size_bids_df[cr01_tb_col].sum()
    print(f"Cumulative CR01 TB: {int(round(cumulative_cr01_tb)):,}")
    
    # Step 9: Create Size Bids Struggling Names table
    print("\n[STEP 9] Creating Size Bids Struggling Names table...")
    size_bids_struggling_names_df = create_size_bids_struggling_names_table(df)
    print(f"Filtered to {len(size_bids_struggling_names_df):,} rows with QUANTITY > 0, CR01 TB >= 1000, POSITION CR01 >= 1,000, and Retracement < 50%")
    
    # Calculate cumulative CR01 TB for Size Bids Struggling Names table
    cumulative_cr01_tb_struggling_names = 0
    if cr01_tb_col in size_bids_struggling_names_df.columns:
        cumulative_cr01_tb_struggling_names = size_bids_struggling_names_df[cr01_tb_col].sum()
    print(f"Cumulative CR01 TB: {int(round(cumulative_cr01_tb_struggling_names)):,}")
    
    # Step 10: Create Size Bids Heavily Offered Lines table
    print("\n[STEP 10] Creating Size Bids Heavily Offered Lines table...")
    size_bids_heavily_offered_lines_df = create_size_bids_heavily_offered_lines_table(df)
    print(f"Filtered to {len(size_bids_heavily_offered_lines_df):,} rows with QUANTITY > 0, CR01 TB >= 1000, POSITION CR01 >= 1,000, Custom_Sector != 'Bail In', and # of Offers >3mm > 2")
    
    # Calculate cumulative CR01 TB for Size Bids Heavily Offered Lines table
    cumulative_cr01_tb_heavily_offered_lines = 0
    if cr01_tb_col in size_bids_heavily_offered_lines_df.columns:
        cumulative_cr01_tb_heavily_offered_lines = size_bids_heavily_offered_lines_df[cr01_tb_col].sum()
    print(f"Cumulative CR01 TB: {int(round(cumulative_cr01_tb_heavily_offered_lines)):,}")
    
    # Step 11: Create Size Bids With Minimal Bid/Offer table
    print("\n[STEP 11] Creating Size Bids With Minimal Bid/Offer table...")
    size_bids_minimal_bo_df = create_size_bids_minimal_bo_table(df)
    print(f"Filtered to {len(size_bids_minimal_bo_df):,} rows with QUANTITY > 0, CR01 TB >= 1000, POSITION CR01 >= 1,000, and Bid/Offer>3mm <= 3")
    
    # Calculate cumulative CR01 TB for Size Bids With Minimal Bid/Offer table
    cumulative_cr01_tb_minimal_bo = 0
    if cr01_tb_col in size_bids_minimal_bo_df.columns:
        cumulative_cr01_tb_minimal_bo = size_bids_minimal_bo_df[cr01_tb_col].sum()
    print(f"Cumulative CR01 TB: {int(round(cumulative_cr01_tb_minimal_bo)):,}")
    
    # Step 12: Create Size Bids With Minimal Bid/Offer No Bail In table
    print("\n[STEP 12] Creating Size Bids With Minimal Bid/Offer No Bail In table...")
    size_bids_minimal_bo_no_bail_in_df = create_size_bids_minimal_bo_no_bail_in_table(df)
    print(f"Filtered to {len(size_bids_minimal_bo_no_bail_in_df):,} rows with QUANTITY > 0, CR01 TB >= 1000, POSITION CR01 >= 1,000, Bid/Offer>3mm <= 3, and Custom_Sector != 'Bail In'")
    
    # Calculate cumulative CR01 TB for Size Bids With Minimal Bid/Offer No Bail In table
    cumulative_cr01_tb_minimal_bo_no_bail_in = 0
    if cr01_tb_col in size_bids_minimal_bo_no_bail_in_df.columns:
        cumulative_cr01_tb_minimal_bo_no_bail_in = size_bids_minimal_bo_no_bail_in_df[cr01_tb_col].sum()
    print(f"Cumulative CR01 TB: {int(round(cumulative_cr01_tb_minimal_bo_no_bail_in)):,}")
    
    # Step 13: Create Size Bids With Minimal Bid/Offer tables by dealer
    print("\n[STEP 13] Creating Size Bids With Minimal Bid/Offer tables by dealer...")
    dealer_col = "Dealer @ Tight Bid >3mm"
    dealer_tables = {}
    dealer_cumulative_cr01 = {}
    
    if dealer_col in size_bids_minimal_bo_df.columns:
        # Get unique dealers (excluding NaN)
        unique_dealers = size_bids_minimal_bo_df[dealer_col].dropna().unique()
        unique_dealers = sorted([d for d in unique_dealers if pd.notna(d)])
        
        print(f"Found {len(unique_dealers)} unique dealers: {', '.join(unique_dealers)}")
        
        for dealer in unique_dealers:
            dealer_df = create_size_bids_minimal_bo_by_dealer_table(df, dealer)
            dealer_tables[dealer] = dealer_df
            
            # Calculate cumulative CR01 TB for this dealer
            dealer_cumulative = 0
            if cr01_tb_col in dealer_df.columns:
                dealer_cumulative = dealer_df[cr01_tb_col].sum()
            dealer_cumulative_cr01[dealer] = dealer_cumulative
            
            print(f"  {dealer}: {len(dealer_df):,} rows, Cumulative CR01 TB: {int(round(dealer_cumulative)):,}")
    else:
        print("Warning: 'Dealer @ Tight Bid >3mm' column not found, skipping dealer-specific tables")
    
    # Step 14: Format and write output
    print("\n[STEP 14] Formatting and writing output...")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Prepare tables dictionary for Excel export
    excel_tables = {}
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        # Write header
        f.write("PORTFOLIO RUNS VIEW\n")
        f.write("="*100 + "\n")
        f.write(f"Last Refreshed: {timestamp}\n")
        f.write(f"\nReference Dates:\n")
        f.write(f"  Last Date: {last_date}\n")
        if mtd_ref_date:
            f.write(f"  MTD Reference Date: {mtd_ref_date.strftime('%Y-%m-%d')}\n")
        else:
            f.write(f"  MTD Reference Date: N/A\n")
        if ytd_ref_date:
            f.write(f"  YTD Reference Date: {ytd_ref_date.strftime('%Y-%m-%d')}\n")
        else:
            f.write(f"  YTD Reference Date: N/A\n")
        f.write("="*100 + "\n")
        
        # Write Portfolio Sorted By CR01 Risk table with Total CR01 summary
        summary_dict_cr01 = {"Total CR01": total_cr01}
        table_title_cr01 = "Portfolio Sorted By CR01 Risk"
        excel_tables[table_title_cr01] = {
            'df': portfolio_cr01_df,
            'summary': summary_dict_cr01
        }
        filter_desc_cr01 = format_filter_description([
            "QUANTITY > 0",
            "Sorted by POSITION CR01 descending (largest to smallest)"
        ])
        f.write(filter_desc_cr01 + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            portfolio_cr01_df,
            table_title_cr01,
            COLUMN_DISPLAY_NAMES,
            summary_dict=summary_dict_cr01
        )
        f.write(table_str)
        
        # Write Portfolio Less Liquid Lines table with Total CR01 summary
        summary_dict_less_liquid = {"Total CR01": total_cr01_less_liquid}
        table_title_less_liquid = "Portfolio Less Liquid Lines"
        excel_tables[table_title_less_liquid] = {
            'df': portfolio_less_liquid_df,
            'summary': summary_dict_less_liquid
        }
        filter_desc_less_liquid = format_filter_description([
            "QUANTITY > 0",
            "Tight Bid >3mm is blank (NaN/null)",
            "Sorted by POSITION CR01 descending (largest to smallest)"
        ])
        f.write(filter_desc_less_liquid + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            portfolio_less_liquid_df,
            table_title_less_liquid,
            COLUMN_DISPLAY_NAMES,
            summary_dict=summary_dict_less_liquid
        )
        f.write(table_str)
        
        # Write Portfolio Sorted By DoD Bid Chg table
        table_title_dod = "Portfolio Sorted By DoD Bid Chg With >3MM on Bid"
        excel_tables[table_title_dod] = {'df': portfolio_dod_bid_df, 'summary': {}}
        filter_desc_dod = format_filter_description([
            "QUANTITY > 0",
            "Tight Bid >3mm has a value (non-blank)",
            "DoD Chg Tight Bid >3mm is non-zero (positive or negative) and non-blank",
            "Sorted by DoD Chg Tight Bid >3mm descending (largest changes first)"
        ])
        f.write(filter_desc_dod + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            portfolio_dod_bid_df,
            table_title_dod,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str)
        
        # Write Portfolio Sorted By MTD Bid Chg table
        table_title_mtd = "Portfolio Sorted By MTD Bid Chg With >3MM on Bid"
        excel_tables[table_title_mtd] = {'df': portfolio_mtd_bid_df, 'summary': {}}
        filter_desc_mtd = format_filter_description([
            "QUANTITY > 0",
            "Tight Bid >3mm has a value (non-blank)",
            "MTD Chg Tight Bid is non-zero (positive or negative) and non-blank",
            "Sorted by MTD Chg Tight Bid descending (largest changes first)"
        ])
        f.write(filter_desc_mtd + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            portfolio_mtd_bid_df,
            table_title_mtd,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str)
        
        # Write Portfolio Sorted By YTD Bid Chg table
        table_title_ytd = "Portfolio Sorted By YTD Bid Chg With >3MM on Bid"
        excel_tables[table_title_ytd] = {'df': portfolio_ytd_bid_df, 'summary': {}}
        filter_desc_ytd = format_filter_description([
            "QUANTITY > 0",
            "Tight Bid >3mm has a value (non-blank)",
            "YTD Chg Tight Bid is non-zero (positive or negative) and non-blank",
            "Sorted by YTD Chg Tight Bid descending (largest changes first)"
        ])
        f.write(filter_desc_ytd + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            portfolio_ytd_bid_df,
            table_title_ytd,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str)
        
        # Write Portfolio Sorted By 1yr Bid Chg table
        table_title_1yr = "Portfolio Sorted By 1yr Bid Chg With >3MM on Bid"
        excel_tables[table_title_1yr] = {'df': portfolio_1yr_bid_df, 'summary': {}}
        filter_desc_1yr = format_filter_description([
            "QUANTITY > 0",
            "Tight Bid >3mm has a value (non-blank)",
            "1yr Chg Tight Bid is non-zero (positive or negative) and non-blank",
            "Sorted by 1yr Chg Tight Bid descending (largest changes first)"
        ])
        f.write(filter_desc_1yr + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            portfolio_1yr_bid_df,
            table_title_1yr,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str)
        
        # Write Size Bids table with cumulative CR01 TB summary
        table_title_size_bids = "Size Bids"
        summary_dict = {"Cumulative CR01 TB": cumulative_cr01_tb}
        excel_tables[table_title_size_bids] = {'df': size_bids_df, 'summary': summary_dict}
        filter_desc_size_bids = format_filter_description([
            "QUANTITY > 0",
            "CR01 @ Tight Bid >= 1000",
            "POSITION CR01 >= 1,000",
            "Sorted by CR01 @ Tight Bid descending (largest first)"
        ])
        f.write(filter_desc_size_bids + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            size_bids_df,
            table_title_size_bids,
            COLUMN_DISPLAY_NAMES,
            summary_dict=summary_dict
        )
        f.write(table_str)
        
        # Write Size Bids Struggling Names table with cumulative CR01 TB summary
        table_title_struggling = "Size Bids Struggling Names"
        summary_dict_struggling_names = {"Cumulative CR01 TB": cumulative_cr01_tb_struggling_names}
        excel_tables[table_title_struggling] = {'df': size_bids_struggling_names_df, 'summary': summary_dict_struggling_names}
        filter_desc_struggling = format_filter_description([
            "QUANTITY > 0",
            "CR01 @ Tight Bid >= 1000",
            "POSITION CR01 >= 1,000",
            "Retracement < 0.5 (Retracement < 50%)",
            "Sorted by CR01 @ Tight Bid descending (largest first)"
        ])
        f.write(filter_desc_struggling + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            size_bids_struggling_names_df,
            table_title_struggling,
            COLUMN_DISPLAY_NAMES,
            summary_dict=summary_dict_struggling_names
        )
        f.write(table_str)
        
        # Write Size Bids Heavily Offered Lines table with cumulative CR01 TB summary
        table_title_heavily_offered = "Size Bids Heavily Offered Lines"
        summary_dict_heavily_offered_lines = {"Cumulative CR01 TB": cumulative_cr01_tb_heavily_offered_lines}
        excel_tables[table_title_heavily_offered] = {'df': size_bids_heavily_offered_lines_df, 'summary': summary_dict_heavily_offered_lines}
        filter_desc_heavily_offered = format_filter_description([
            "QUANTITY > 0",
            "CR01 @ Tight Bid >= 1000",
            "POSITION CR01 >= 1,000",
            "Custom_Sector != \"Bail In\" (excludes Bail In sector)",
            "# of Offers >3mm > 2",
            "Sorted by CR01 @ Tight Bid descending (largest first)"
        ])
        f.write(filter_desc_heavily_offered + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            size_bids_heavily_offered_lines_df,
            table_title_heavily_offered,
            COLUMN_DISPLAY_NAMES,
            summary_dict=summary_dict_heavily_offered_lines
        )
        f.write(table_str)
        
        # Write Size Bids With Minimal Bid/Offer table with cumulative CR01 TB summary
        table_title_minimal_bo = "Size Bids With Minimal Bid/Offer"
        summary_dict_minimal_bo = {"Cumulative CR01 TB": cumulative_cr01_tb_minimal_bo}
        excel_tables[table_title_minimal_bo] = {'df': size_bids_minimal_bo_df, 'summary': summary_dict_minimal_bo}
        filter_desc_minimal_bo = format_filter_description([
            "QUANTITY > 0",
            "CR01 @ Tight Bid >= 1000",
            "POSITION CR01 >= 1,000",
            "Bid/Offer>3mm <= 3 (excludes > 3 or blank)",
            "Sorted by CR01 @ Tight Bid descending (largest first)"
        ])
        f.write(filter_desc_minimal_bo + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            size_bids_minimal_bo_df,
            table_title_minimal_bo,
            COLUMN_DISPLAY_NAMES,
            summary_dict=summary_dict_minimal_bo
        )
        f.write(table_str)
        
        # Write Size Bids With Minimal Bid/Offer No Bail In table with cumulative CR01 TB summary
        table_title_minimal_bo_no_bail = "Size Bids With Minimal Bid/Offer No Bail In"
        summary_dict_minimal_bo_no_bail_in = {"Cumulative CR01 TB": cumulative_cr01_tb_minimal_bo_no_bail_in}
        excel_tables[table_title_minimal_bo_no_bail] = {'df': size_bids_minimal_bo_no_bail_in_df, 'summary': summary_dict_minimal_bo_no_bail_in}
        filter_desc_minimal_bo_no_bail = format_filter_description([
            "QUANTITY > 0",
            "CR01 @ Tight Bid >= 1000",
            "POSITION CR01 >= 1,000",
            "Bid/Offer>3mm <= 3 (excludes > 3 or blank)",
            "Custom_Sector != \"Bail In\" (excludes Bail In sector)",
            "Sorted by CR01 @ Tight Bid descending (largest first)"
        ])
        f.write(filter_desc_minimal_bo_no_bail + "\n")
        f.write("="*100 + "\n\n")
        table_str = format_table(
            size_bids_minimal_bo_no_bail_in_df,
            table_title_minimal_bo_no_bail,
            COLUMN_DISPLAY_NAMES,
            summary_dict=summary_dict_minimal_bo_no_bail_in
        )
        f.write(table_str)
        
        # Write Size Bids With Minimal Bid/Offer tables by dealer
        for dealer in sorted(dealer_tables.keys()):
            dealer_df = dealer_tables[dealer]
            dealer_cumulative = dealer_cumulative_cr01[dealer]
            table_title_dealer = f"Size Bids With Minimal Bid/Offer: Where {dealer} Is Best Bid"
            summary_dict_dealer = {"Cumulative CR01 TB": dealer_cumulative}
            excel_tables[table_title_dealer] = {'df': dealer_df, 'summary': summary_dict_dealer}
            filter_desc_dealer = format_filter_description([
                "QUANTITY > 0",
                "CR01 @ Tight Bid >= 1000",
                "POSITION CR01 >= 1,000",
                "Bid/Offer>3mm <= 3 (excludes > 3 or blank)",
                f"Dealer @ Tight Bid >3mm = {dealer}",
                "Sorted by CR01 @ Tight Bid descending (largest first)"
            ])
            f.write(filter_desc_dealer + "\n")
            f.write("="*100 + "\n\n")
            table_str = format_table(
                dealer_df,
                table_title_dealer,
                COLUMN_DISPLAY_NAMES,
                summary_dict=summary_dict_dealer
            )
            f.write(table_str)
        
        f.write("\n" + "="*100 + "\n")
        f.write("END OF REPORT\n")
    
    print(f"\nOutput written to: {OUTPUT_FILE}")
    print(f"Total rows in Portfolio CR01 Risk table: {len(portfolio_cr01_df):,}")
    print(f"Total rows in Portfolio Less Liquid Lines table: {len(portfolio_less_liquid_df):,}")
    print(f"Total rows in Portfolio DoD Bid Chg table: {len(portfolio_dod_bid_df):,}")
    print(f"Total rows in Portfolio MTD Bid Chg table: {len(portfolio_mtd_bid_df):,}")
    print(f"Total rows in Portfolio YTD Bid Chg table: {len(portfolio_ytd_bid_df):,}")
    print(f"Total rows in Portfolio 1yr Bid Chg table: {len(portfolio_1yr_bid_df):,}")
    print(f"Total rows in Size Bids table: {len(size_bids_df):,}")
    print(f"Total rows in Size Bids Struggling Names table: {len(size_bids_struggling_names_df):,}")
    print(f"Total rows in Size Bids Heavily Offered Lines table: {len(size_bids_heavily_offered_lines_df):,}")
    print(f"Total rows in Size Bids With Minimal Bid/Offer table: {len(size_bids_minimal_bo_df):,}")
    print(f"Total rows in Size Bids With Minimal Bid/Offer No Bail In table: {len(size_bids_minimal_bo_no_bail_in_df):,}")
    
    # Step 15: Write Excel file
    print("\n[STEP 15] Writing Excel file...")
    
    # Check if at least one table has data (openpyxl requires at least one visible sheet)
    has_data = any(not table_data['df'].empty for table_data in excel_tables.values())
    
    if has_data:
        write_excel_file(
            EXCEL_OUTPUT_FILE,
            excel_tables,
            timestamp,
            last_date,
            mtd_ref_date,
            ytd_ref_date
        )
        print(f"Excel file written to: {EXCEL_OUTPUT_FILE}")
    else:
        print("Warning: All tables are empty. Skipping Excel file generation.")
        print("(openpyxl requires at least one visible sheet)")
    
    print("\nDone!")


def format_filter_description(filter_parts: list[str]) -> str:
    """
    Format filter description as multi-line text for better readability.
    
    Args:
        filter_parts: List of filter description parts (each part is a sentence or clause).
    
    Returns:
        Multi-line formatted filter description string.
    """
    # Join parts with newlines and proper indentation
    # Use ASCII-safe "-" instead of Unicode bullet points
    formatted = "Filters Applied:\n"
    for part in filter_parts:
        # Add 2 spaces indentation for each filter line
        formatted += f"  - {part.strip()}\n"
    return formatted.rstrip()  # Remove trailing newline


def generate_universe_views() -> None:
    """
    Generate universe RV views from runs_today.csv.
    
    Creates custom formatted tables for universe-wide analysis (not filtered by portfolio).
    Outputs to uni_runs_view.txt and uni_runs_view.xlsx.
    """
    print("="*100)
    print("UNIVERSE RV VIEWS - Universe Runs View Generator")
    print("="*100)
    
    # Step 0: Validate runs_today.csv is current with parquet (auto-regenerate if stale)
    print("\n[STEP 0] Validating runs_today.csv is current with runs_timeseries.parquet...")
    parquet_last_date = validate_runs_today_csv_is_current(
        RUNS_TODAY_CSV_PATH,
        RUNS_PARQUET_PATH,
        auto_regenerate=True
    )
    print(f"  Parquet last date: {parquet_last_date}")
    
    # Step 1: Load runs_today.csv
    print("\n[STEP 1] Loading runs_today.csv...")
    df = pd.read_csv(RUNS_TODAY_CSV_PATH)
    
    # Convert all numeric columns to proper types (ensures filtering works correctly)
    df = ensure_numeric_types(df)
    
    print(f"Loaded {len(df):,} rows, {len(df.columns)} columns")
    
    # Step 2: Get reference dates from runs_timeseries.parquet
    print("\n[STEP 2] Extracting reference dates from runs_timeseries.parquet...")
    ref_dates = get_reference_dates(RUNS_PARQUET_PATH)
    last_date = ref_dates["last_date"]
    mtd_ref_date = ref_dates["mtd_ref_date"]
    ytd_ref_date = ref_dates["ytd_ref_date"]
    
    print(f"Last date: {last_date}")
    print(f"MTD reference date: {mtd_ref_date}")
    print(f"YTD reference date: {ytd_ref_date}")
    
    # Step 3: Create Universe Sorted By DoD Moves With Size On Offer >3mm table
    print("\n[STEP 3] Creating Universe Sorted By DoD Moves With Size On Offer >3mm table...")
    universe_dod_df = create_universe_dod_moves_table(df)
    
    # Build filter description for >3mm table
    excluded_sectors_str = ", ".join(UNIVERSE_EXCLUDED_SECTORS)
    filter_description_3mm = format_filter_description([
        f"Excluded Custom_Sector values: {excluded_sectors_str}",
        f"Only includes rows where Wide Offer >3mm had a value on both last date and second-to-last date",
        f"Excluded rows where {UNIVERSE_DOD_SORT_COLUMN} is 0 or blank",
        f"Showing top {UNIVERSE_TOP_BOTTOM_N} and bottom {UNIVERSE_TOP_BOTTOM_N} by {UNIVERSE_DOD_SORT_COLUMN}"
    ])
    
    print(f"Filtered to {len(universe_dod_df):,} rows")
    print(f"Filters: {filter_description_3mm}")
    
    # Step 3.5: Create Universe Sorted By DoD Moves table (sorted by DoD WO)
    print("\n[STEP 3.5] Creating Universe Sorted By DoD Moves table...")
    universe_dod_wo_df = create_universe_dod_moves_wo_table(df)
    
    # Build filter description for regular WO table
    filter_description_wo = format_filter_description([
        f"Excluded Custom_Sector values: {excluded_sectors_str}",
        f"Only includes rows where Wide Offer had a value on both last date and second-to-last date",
        f"Excluded rows where DoD Chg Wide Offer is 0 or blank",
        f"Showing top {UNIVERSE_TOP_BOTTOM_N} and bottom {UNIVERSE_TOP_BOTTOM_N} by DoD Chg Wide Offer"
    ])
    
    print(f"Filtered to {len(universe_dod_wo_df):,} rows")
    print(f"Filters: {filter_description_wo}")
    
    # Step 3.75: Create Universe Sorted By MTD Moves With Size On Offer >3mm table
    print("\n[STEP 3.75] Creating Universe Sorted By MTD Moves With Size On Offer >3mm table...")
    universe_mtd_df = create_universe_mtd_moves_table(df)
    
    # Build filter description for MTD table
    filter_description_mtd = format_filter_description([
        f"Excluded Custom_Sector values: {excluded_sectors_str}",
        f"Only includes rows where Wide Offer >3mm > 0",
        f"Excluded rows where MTD Chg Tight Bid is 0 or blank",
        f"Showing top {UNIVERSE_TOP_BOTTOM_N} and bottom {UNIVERSE_TOP_BOTTOM_N} by MTD Chg Tight Bid"
    ])
    
    print(f"Filtered to {len(universe_mtd_df):,} rows")
    print(f"Filters: {filter_description_mtd}")
    
    # Step 3.8: Create Universe Sorted By YTD Moves With Size On Offer >3mm table
    print("\n[STEP 3.8] Creating Universe Sorted By YTD Moves With Size On Offer >3mm table...")
    universe_ytd_df = create_universe_ytd_moves_table(df)
    
    # Build filter description for YTD table
    filter_description_ytd = format_filter_description([
        f"Excluded Custom_Sector values: {excluded_sectors_str}",
        f"Only includes rows where Wide Offer >3mm > 0",
        f"Excluded rows where YTD Chg Tight Bid is 0 or blank",
        f"Showing top {UNIVERSE_TOP_BOTTOM_N} and bottom {UNIVERSE_TOP_BOTTOM_N} by YTD Chg Tight Bid"
    ])
    
    print(f"Filtered to {len(universe_ytd_df):,} rows")
    print(f"Filters: {filter_description_ytd}")
    
    # Step 3.85: Create Universe Sorted By 1yr Moves With Size On Offer >3mm table
    print("\n[STEP 3.85] Creating Universe Sorted By 1yr Moves With Size On Offer >3mm table...")
    universe_1yr_df = create_universe_1yr_moves_table(df)
    
    # Build filter description for 1yr table
    filter_description_1yr = format_filter_description([
        f"Excluded Custom_Sector values: {excluded_sectors_str}",
        f"Only includes rows where Wide Offer >3mm > 0",
        f"Excluded rows where 1yr Chg Tight Bid is 0 or blank",
        f"Showing top {UNIVERSE_TOP_BOTTOM_N} and bottom {UNIVERSE_TOP_BOTTOM_N} by 1yr Chg Tight Bid"
    ])
    
    print(f"Filtered to {len(universe_1yr_df):,} rows")
    print(f"Filters: {filter_description_1yr}")
    
    # Step 3.9: Create Large CR01 On Offer table
    print("\n[STEP 3.9] Creating Large CR01 On Offer table...")
    large_cr01_df = create_large_cr01_on_offer_table(df)
    
    # Build filter description for Large CR01 On Offer table
    filter_description_large_cr01 = format_filter_description([
        f"Excluded Custom_Sector values: {excluded_sectors_str}",
        f"CR01 @ Wide Offer > 3000",
        f"Bid/Offer>3mm < 3",
        f"Showing all matching rows (sorted by Bid/Offer>3mm ascending, low to high)"
    ])
    
    print(f"Filtered to {len(large_cr01_df):,} rows")
    print(f"Filters: {filter_description_large_cr01}")
    
    # Step 3.92: Create Large CR01 On Offer tables by dealer
    print("\n[STEP 3.92] Creating Large CR01 On Offer tables by dealer...")
    dealer_wo_col = "Dealer @ Wide Offer >3mm"
    large_cr01_by_dealer = {}
    
    if dealer_wo_col in large_cr01_df.columns:
        # Get unique dealers (excluding NaN/null values)
        unique_dealers = large_cr01_df[dealer_wo_col].dropna().unique()
        unique_dealers = sorted([d for d in unique_dealers if pd.notna(d) and str(d).strip() != ''])
        
        print(f"Found {len(unique_dealers)} unique dealers: {', '.join(unique_dealers)}")
        
        for dealer in unique_dealers:
            # Filter to rows where this dealer is the Dealer @ Wide Offer >3mm
            dealer_df = large_cr01_df[
                large_cr01_df[dealer_wo_col].notna() & 
                (large_cr01_df[dealer_wo_col] == dealer)
            ].copy()
            
            if len(dealer_df) > 0:
                large_cr01_by_dealer[dealer] = dealer_df
                print(f"  {dealer}: {len(dealer_df)} rows")
    else:
        print(f"Warning: Column '{dealer_wo_col}' not found in DataFrame")
    
    # Step 3.95: Create Large CR01 On Offer, No Longs table
    print("\n[STEP 3.95] Creating Large CR01 On Offer, No Longs table...")
    large_cr01_no_longs_df = create_large_cr01_on_offer_no_longs_table(df)
    
    # Build filter description for Large CR01 On Offer, No Longs table
    filter_description_large_cr01_no_longs = format_filter_description([
        f"Excluded Custom_Sector values: {excluded_sectors_str}",
        f"CR01 @ Wide Offer > 3000",
        f"Bid/Offer>3mm < 3",
        f"Yrs (Cvn) < 11",
        f"Showing all matching rows (sorted by Bid/Offer>3mm ascending, low to high)"
    ])
    
    print(f"Filtered to {len(large_cr01_no_longs_df):,} rows")
    print(f"Filters: {filter_description_large_cr01_no_longs}")
    
    # Step 3.98: Create Tough To Find Offers table
    print("\n[STEP 3.98] Creating Tough To Find Offers table...")
    tough_to_find_offers_df = create_tough_to_find_offers_table(df)
    
    # Build filter description for Tough To Find Offers table
    filter_description_tough_to_find = format_filter_description([
        f"Excluded Custom_Sector values: {excluded_sectors_str}",
        f"CR01 @ Wide Offer > 2000",
        f"Bid/Offer>3mm < 4",
        f"# Offers>3mm < 2",
        f"Showing all matching rows (sorted by Bid/Offer>3mm ascending, low to high)"
    ])
    
    print(f"Filtered to {len(tough_to_find_offers_df):,} rows")
    print(f"Filters: {filter_description_tough_to_find}")
    
    # Step 3.99: Create Carry Bonds table
    print("\n[STEP 3.99] Creating Carry Bonds table...")
    carry_bonds_df = create_carry_bonds_table(df)
    
    # Build filter description for Carry Bonds table
    filter_description_carry_bonds = format_filter_description([
        f"Excluded Custom_Sector values: {excluded_sectors_str}",
        f"CR01 @ Wide Offer > 500",
        f"Bid/Offer>3mm < 6",
        f"Yrs (Cvn) < 2",
        f"Wide Offer >3mm > 60",
        f"Showing all matching rows (sorted by Wide Offer >3mm descending, high to low)"
    ])
    
    print(f"Filtered to {len(carry_bonds_df):,} rows")
    print(f"Filters: {filter_description_carry_bonds}")
    
    # Step 3.995: Create Carry Bonds Sorted by MTD Moves table
    print("\n[STEP 3.995] Creating Carry Bonds Sorted by MTD Moves table...")
    carry_bonds_mtd_df = create_carry_bonds_sorted_by_mtd_table(df)
    
    # Build filter description for Carry Bonds Sorted by MTD Moves table
    filter_description_carry_bonds_mtd = format_filter_description([
        f"Excluded Custom_Sector values: {excluded_sectors_str}",
        f"CR01 @ Wide Offer > 500",
        f"Bid/Offer>3mm < 6",
        f"Yrs (Cvn) < 2",
        f"Wide Offer >3mm > 60",
        f"Showing all matching rows (sorted by MTD Chg Tight Bid descending, largest to smallest)"
    ])
    
    print(f"Filtered to {len(carry_bonds_mtd_df):,} rows")
    print(f"Filters: {filter_description_carry_bonds_mtd}")
    
    # Step 4: Format and write output
    print("\n[STEP 4] Formatting and writing output...")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Prepare tables dictionary for Excel export
    excel_tables = {}
    
    with open(UNI_OUTPUT_FILE, 'w', encoding='utf-8') as f:
        # Write header
        f.write("UNIVERSE RV VIEWS\n")
        f.write("="*100 + "\n")
        f.write(f"Last Refreshed: {timestamp}\n")
        f.write(f"\nReference Dates:\n")
        f.write(f"  Last Date: {last_date}\n")
        if mtd_ref_date:
            f.write(f"  MTD Reference Date: {mtd_ref_date.strftime('%Y-%m-%d')}\n")
        else:
            f.write(f"  MTD Reference Date: N/A\n")
        if ytd_ref_date:
            f.write(f"  YTD Reference Date: {ytd_ref_date.strftime('%Y-%m-%d')}\n")
        else:
            f.write(f"  YTD Reference Date: N/A\n")
        f.write(f"\n{filter_description_3mm}\n")
        f.write("="*100 + "\n")
        
        # Write Universe Sorted By DoD Moves With Size On Offer >3mm table
        table_title_dod_3mm = "Universe Sorted By DoD Moves With Size On Offer >3mm"
        excel_tables[table_title_dod_3mm] = {'df': universe_dod_df, 'summary': {}}
        table_str_3mm = format_table(
            universe_dod_df,
            table_title_dod_3mm,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str_3mm)
        
        f.write("\n" + "="*100 + "\n")
        f.write(f"\n{filter_description_wo}\n")
        f.write("="*100 + "\n")
        
        # Write Universe Sorted By DoD Moves table (sorted by DoD WO)
        table_title_dod_wo = "Universe Sorted By DoD Moves"
        excel_tables[table_title_dod_wo] = {'df': universe_dod_wo_df, 'summary': {}}
        table_str_wo = format_table(
            universe_dod_wo_df,
            table_title_dod_wo,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str_wo)
        
        f.write("\n" + "="*100 + "\n")
        f.write(f"\n{filter_description_mtd}\n")
        f.write("="*100 + "\n")
        
        # Write Universe Sorted By MTD Moves With Size On Offer >3mm table
        table_title_mtd = "Universe Sorted By MTD Moves With Size On Offer >3mm"
        excel_tables[table_title_mtd] = {'df': universe_mtd_df, 'summary': {}}
        table_str_mtd = format_table(
            universe_mtd_df,
            table_title_mtd,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str_mtd)
        
        f.write("\n" + "="*100 + "\n")
        f.write(f"\n{filter_description_ytd}\n")
        f.write("="*100 + "\n")
        
        # Write Universe Sorted By YTD Moves With Size On Offer >3mm table
        table_title_ytd = "Universe Sorted By YTD Moves With Size On Offer >3mm"
        excel_tables[table_title_ytd] = {'df': universe_ytd_df, 'summary': {}}
        table_str_ytd = format_table(
            universe_ytd_df,
            table_title_ytd,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str_ytd)
        
        f.write("\n" + "="*100 + "\n")
        f.write(f"\n{filter_description_1yr}\n")
        f.write("="*100 + "\n")
        
        # Write Universe Sorted By 1yr Moves With Size On Offer >3mm table
        table_title_1yr = "Universe Sorted By 1yr Moves With Size On Offer >3mm"
        excel_tables[table_title_1yr] = {'df': universe_1yr_df, 'summary': {}}
        table_str_1yr = format_table(
            universe_1yr_df,
            table_title_1yr,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str_1yr)
        
        f.write("\n" + "="*100 + "\n")
        f.write(f"\n{filter_description_large_cr01}\n")
        f.write("="*100 + "\n")
        
        # Write Large CR01 On Offer table
        table_title_large_cr01 = "Large CR01 On Offer"
        excel_tables[table_title_large_cr01] = {'df': large_cr01_df, 'summary': {}}
        table_str_large_cr01 = format_table(
            large_cr01_df,
            table_title_large_cr01,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str_large_cr01)
        
        # Write Large CR01 On Offer tables by dealer
        dealer_wo_col = "Dealer @ Wide Offer >3mm"
        if dealer_wo_col in large_cr01_df.columns:
            # Get unique dealers (excluding NaN/null values)
            unique_dealers = large_cr01_df[dealer_wo_col].dropna().unique()
            unique_dealers = sorted([d for d in unique_dealers if pd.notna(d) and str(d).strip() != ''])
            
            for dealer in unique_dealers:
                # Filter to rows where this dealer is the Dealer @ Wide Offer >3mm
                dealer_df = large_cr01_df[
                    large_cr01_df[dealer_wo_col].notna() & 
                    (large_cr01_df[dealer_wo_col] == dealer)
                ].copy()
                
                if len(dealer_df) > 0:
                    f.write("\n" + "="*100 + "\n")
                    dealer_filter_description = format_filter_description([
                        f"Excluded Custom_Sector values: {excluded_sectors_str}",
                        f"CR01 @ Wide Offer > 3000",
                        f"Bid/Offer>3mm < 3",
                        f"Dealer @ Wide Offer >3mm = {dealer}",
                        f"Showing all matching rows (sorted by Bid/Offer>3mm ascending, low to high)"
                    ])
                    f.write(f"\n{dealer_filter_description}\n")
                    f.write("="*100 + "\n")
                    
                    # Write dealer-specific table
                    table_title_dealer = f"Large CR01 On Offer: Where {dealer} Is The Wide Offer"
                    excel_tables[table_title_dealer] = {'df': dealer_df, 'summary': {}}
                    table_str_dealer = format_table(
                        dealer_df,
                        table_title_dealer,
                        COLUMN_DISPLAY_NAMES
                    )
                    f.write(table_str_dealer)
        
        f.write("\n" + "="*100 + "\n")
        f.write(f"\n{filter_description_large_cr01_no_longs}\n")
        f.write("="*100 + "\n")
        
        # Write Large CR01 On Offer, No Longs table
        table_title_large_cr01_no_longs = "Large CR01 On Offer, No Longs"
        excel_tables[table_title_large_cr01_no_longs] = {'df': large_cr01_no_longs_df, 'summary': {}}
        table_str_large_cr01_no_longs = format_table(
            large_cr01_no_longs_df,
            table_title_large_cr01_no_longs,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str_large_cr01_no_longs)
        
        f.write("\n" + "="*100 + "\n")
        f.write(f"\n{filter_description_tough_to_find}\n")
        f.write("="*100 + "\n")
        
        # Write Tough To Find Offers table
        table_title_tough_to_find = "Tough To Find Offers"
        excel_tables[table_title_tough_to_find] = {'df': tough_to_find_offers_df, 'summary': {}}
        table_str_tough_to_find = format_table(
            tough_to_find_offers_df,
            table_title_tough_to_find,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str_tough_to_find)
        
        f.write("\n" + "="*100 + "\n")
        f.write(f"\n{filter_description_carry_bonds}\n")
        f.write("="*100 + "\n")
        
        # Write Carry Bonds table
        table_title_carry_bonds = "Carry Bonds"
        excel_tables[table_title_carry_bonds] = {'df': carry_bonds_df, 'summary': {}}
        table_str_carry_bonds = format_table(
            carry_bonds_df,
            table_title_carry_bonds,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str_carry_bonds)
        
        f.write("\n" + "="*100 + "\n")
        f.write(f"\n{filter_description_carry_bonds_mtd}\n")
        f.write("="*100 + "\n")
        
        # Write Carry Bonds Sorted by MTD Moves table
        table_title_carry_bonds_mtd = "Carry Bonds Sorted by MTD Moves"
        excel_tables[table_title_carry_bonds_mtd] = {'df': carry_bonds_mtd_df, 'summary': {}}
        table_str_carry_bonds_mtd = format_table(
            carry_bonds_mtd_df,
            table_title_carry_bonds_mtd,
            COLUMN_DISPLAY_NAMES
        )
        f.write(table_str_carry_bonds_mtd)
        
        f.write("\n" + "="*100 + "\n")
        f.write("END OF REPORT\n")
    
    print(f"\nOutput written to: {UNI_OUTPUT_FILE}")
    print(f"Total rows in Universe DoD Moves With Size On Offer >3mm table: {len(universe_dod_df):,}")
    print(f"Total rows in Universe DoD Moves table: {len(universe_dod_wo_df):,}")
    print(f"Total rows in Universe MTD Moves With Size On Offer >3mm table: {len(universe_mtd_df):,}")
    print(f"Total rows in Universe YTD Moves With Size On Offer >3mm table: {len(universe_ytd_df):,}")
    print(f"Total rows in Universe 1yr Moves With Size On Offer >3mm table: {len(universe_1yr_df):,}")
    print(f"Total rows in Large CR01 On Offer table: {len(large_cr01_df):,}")
    # Count dealer-specific tables
    dealer_wo_col = "Dealer @ Wide Offer >3mm"
    if dealer_wo_col in large_cr01_df.columns:
        unique_dealers = large_cr01_df[dealer_wo_col].dropna().unique()
        unique_dealers = sorted([d for d in unique_dealers if pd.notna(d) and str(d).strip() != ''])
        for dealer in unique_dealers:
            dealer_df = large_cr01_df[
                large_cr01_df[dealer_wo_col].notna() & 
                (large_cr01_df[dealer_wo_col] == dealer)
            ].copy()
            if len(dealer_df) > 0:
                print(f"Total rows in Large CR01 On Offer: Where {dealer} Is The Wide Offer table: {len(dealer_df):,}")
    print(f"Total rows in Large CR01 On Offer, No Longs table: {len(large_cr01_no_longs_df):,}")
    print(f"Total rows in Tough To Find Offers table: {len(tough_to_find_offers_df):,}")
    print(f"Total rows in Carry Bonds table: {len(carry_bonds_df):,}")
    print(f"Total rows in Carry Bonds Sorted by MTD Moves table: {len(carry_bonds_mtd_df):,}")
    
    # Step 5: Write Excel file
    print("\n[STEP 5] Writing Excel file...")
    
    # Check if at least one table has data (openpyxl requires at least one visible sheet)
    has_data = any(not table_data['df'].empty for table_data in excel_tables.values())
    
    if has_data:
        write_excel_file(
            UNI_EXCEL_OUTPUT_FILE,
            excel_tables,
            timestamp,
            last_date,
            mtd_ref_date,
            ytd_ref_date
        )
        print(f"Excel file written to: {UNI_EXCEL_OUTPUT_FILE}")
    else:
        print("Warning: All tables are empty. Skipping Excel file generation.")
        print("(openpyxl requires at least one visible sheet)")
    
    print("\nDone!")


if __name__ == "__main__":
    # Generate both portfolio and universe views
    main()  # Generates portfolio_runs_view.txt and portfolio_runs_view.xlsx
    print("\n" + "="*100 + "\n")
    generate_universe_views()  # Generates uni_runs_view.txt and uni_runs_view.xlsx

