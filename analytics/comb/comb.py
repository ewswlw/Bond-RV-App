"""
Consolidated pair analytics script.

This module consolidates all pair analytics scripts into a single modular file.
Runs all analysis types and outputs results to Excel file, console, comb.txt, and comb_validation.txt.
"""

from __future__ import annotations

from dataclasses import dataclass
from itertools import combinations, product
import math
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from datetime import datetime
import io

import numpy as np
import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ============================================================================
# PATHS (auto-configured based on script location)
# ============================================================================

# Get script directory and build paths relative to it
SCRIPT_DIR = Path(__file__).parent.resolve()
BQL_PARQUET_PATH = SCRIPT_DIR.parent.parent / "bond_data" / "parquet" / "bql.parquet"
HISTORICAL_PARQUET_PATH = SCRIPT_DIR.parent.parent / "bond_data" / "parquet" / "historical_bond_details.parquet"
UNIVERSE_PARQUET_PATH = SCRIPT_DIR.parent.parent / "bond_data" / "parquet" / "universe.parquet"
RUNS_TODAY_CSV_PATH = SCRIPT_DIR.parent / "processed_data" / "runs_today.csv"
PORTFOLIO_PARQUET_PATH = SCRIPT_DIR.parent.parent / "bond_data" / "parquet" / "historical_portfolio.parquet"
OUTPUT_DIR = SCRIPT_DIR.parent / "processed_data"

# ============================================================================
# CONFIGURATION SECTIONS
# ============================================================================

# ----------------------------------------------------------------------------
# FILTERS CONFIGURATION
# ----------------------------------------------------------------------------
FILTERS = {
    "recent_date_percent": 0.75,  # Percentage of recent dates to filter (0.75 = 75%)
    "currency_filter": "CAD",      # Filter to CAD CUSIPs (or None for all)
    "max_yrs_cvn_diff": 0.8,       # Maximum absolute difference in Yrs (Cvn) for term_comb
    "max_yrs_cvn_diff_cad_usd": 2.0,  # Maximum absolute difference for CAD/USD pairs
    "cr01_tight_bid_threshold": 2000,  # CR01 @ Tight Bid threshold
    "cr01_wide_offer_threshold": 2000,  # CR01 @ Wide Offer threshold
    "bid_offer_threshold": 3,      # Bid/Offer>3mm threshold (< 3 for decent bid offer)
}

# ----------------------------------------------------------------------------
# TARGET BOND CONFIGURATION
# ----------------------------------------------------------------------------
TARGET_BOND = {
    "ticker": "TCN",  # Ticker of target bond(s) - takes precedence over security
    "security": None,  # Security name of target bond (fallback if ticker not found)
}

# ----------------------------------------------------------------------------
# DISPLAY CONFIGURATION
# ----------------------------------------------------------------------------
DISPLAY = {
    "top_n_pairs": 80,  # Number of top pairs to display in console/txt files
}

# ----------------------------------------------------------------------------
# TABLE TITLES CONFIGURATION
# ----------------------------------------------------------------------------
TABLE_TITLES = {
    "all_comb": "All Combinations Pair Analytics",
    "term_comb": "Term Combinations Pair Analytics (Yrs Cvn Diff <= 0.8)",
    "ticker_comb": "Ticker Combinations Pair Analytics",
    "custom_sector": "Custom Sector Combinations Pair Analytics",
    "custom_bond_comb": "Custom Bond Combinations Pair Analytics",
    "custom_bond_vs_holdings": "Custom Bond vs Holdings Pair Analytics",
    "cad_cheap_vs_usd": "CAD Cheap vs USD Pair Analytics",
    "cad_rich_vs_usd": "CAD Rich vs USD Pair Analytics",
    "executable_cr01_vs_holdings": "Executable CR01 vs Holdings Pair Analytics",
    "executable_cr01_decent_bid_offer_vs_holdings": "Executable CR01 Decent Bid Offer vs Holdings Pair Analytics",
    "all_combos_vs_holdings": "All Combos vs Holdings Pair Analytics",
}

# ----------------------------------------------------------------------------
# EXCEL TAB NAMES CONFIGURATION (max 31 characters for Excel sheet names)
# ----------------------------------------------------------------------------
EXCEL_TAB_NAMES = {
    "all_comb": "All Combinations",
    "term_comb": "Term Combinations",
    "ticker_comb": "Ticker Combinations",
    "custom_sector": "Custom Sector",
    "custom_bond_comb": "Custom Bond Comb",
    "custom_bond_vs_holdings": "Custom Bond vs Holdings",
    "cad_cheap_vs_usd": "CAD Cheap vs USD",
    "cad_rich_vs_usd": "CAD Rich vs USD",
    "executable_cr01_vs_holdings": "CR01 vs Holdings",
    "executable_cr01_decent_bid_offer_vs_holdings": "CR01 Decent Bid Offer",
    "all_combos_vs_holdings": "All Combos vs Holdings",
}

# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class PairSummary:
    """Container with pair analytics metrics."""
    Bond_1: str
    Bond_2: str
    last_value: float
    average_value: float
    vs_average: float
    z_score: Optional[float]
    percentile: float
    cusip_1: str
    cusip_2: str

@dataclass
class CR01PairSummary:
    """Container for CR01-style pair analytics (universe/holdings naming)."""
    universe_name: str
    holdings_name: str
    last_value: float
    average_value: float
    vs_average: float
    z_score: Optional[float]
    percentile: float
    universe_cusip: str
    holdings_cusip: str

# ============================================================================
# SHARED UTILITY FUNCTIONS
# ============================================================================

def ensure_ascii(value: Optional[str]) -> str:
    """
    Convert text to ASCII by removing or replacing unsupported characters.

    Args:
        value: Input string that may contain non-ASCII characters.

    Returns:
        ASCII-safe string representation.
    """
    if value is None:
        return ""
    sanitized = value.encode("ascii", errors="ignore").decode("ascii")
    if sanitized:
        return sanitized
    return value.encode("ascii", errors="replace").decode("ascii")

def normalize_cusip(cusip: str) -> str:
    """
    Normalize a CUSIP to 9-character uppercase format.
    
    Handles CUSIPs that may already be normalized or may have trailing descriptors.
    
    Args:
        cusip: CUSIP string to normalize.
    
    Returns:
        Normalized 9-character uppercase CUSIP.
    """
    if pd.isna(cusip) or cusip == '':
        return ''
    cusip_str = str(cusip).strip().upper()
    # Remove trailing " CORP" if present
    cusip_str = cusip_str.replace(" CORP", "")
    # Take first 9 characters
    return cusip_str[:9]

def filter_recent_cusips(data: pd.DataFrame, percent: float) -> pd.DataFrame:
    """
    Filter to CUSIPs present in the most recent percent of dates.

    Args:
        data: BQL DataFrame with Date and CUSIP columns.
        percent: Percentage of most recent dates to consider (default 0.75).

    Returns:
        Filtered DataFrame containing only CUSIPs present in recent dates.
    """
    unique_dates = sorted(data["Date"].unique())
    num_recent_dates = max(1, int(len(unique_dates) * percent))
    recent_dates = set(unique_dates[-num_recent_dates:])
    
    # Filter to recent dates first
    recent_data = data[data["Date"].isin(recent_dates)].copy()
    
    # Find CUSIPs that appear in all recent dates (or at least most of them)
    # For efficiency, we'll use CUSIPs that have data in at least 90% of recent dates
    cusip_date_counts = recent_data.groupby("CUSIP")["Date"].nunique()
    min_dates_required = max(1, int(num_recent_dates * 0.9))
    valid_cusips = cusip_date_counts[cusip_date_counts >= min_dates_required].index.tolist()
    
    # Return all data for valid CUSIPs (not just recent dates)
    return data[data["CUSIP"].isin(valid_cusips)].copy()

def build_name_lookup(data: pd.DataFrame) -> Dict[str, str]:
    """
    Generate a mapping from CUSIP to security name.

    Args:
        data: BQL DataFrame containing `CUSIP` and `Name`.

    Returns:
        Dictionary mapping each CUSIP to its most recent non-null name.
    """
    name_series = (
        data.dropna(subset=["Name"])
        .sort_values("Date")
        .groupby("CUSIP")["Name"]
        .last()
    )
    return {
        cusip: ensure_ascii(name)
        for cusip, name in name_series.to_dict().items()
    }

def compute_pair_stats_vectorized(
    cusip_1_values: np.ndarray,
    cusip_2_values: np.ndarray,
) -> Optional[Tuple[float, float, float, Optional[float], float]]:
    """
    Compute pair statistics using vectorized operations.

    Args:
        cusip_1_values: Array of values for first CUSIP (aligned by date).
        cusip_2_values: Array of values for second CUSIP (aligned by date).

    Returns:
        Tuple of (last_value, average_value, vs_average, z_score, percentile)
        or None if insufficient data.
    """
    # Compute spreads
    spreads = cusip_1_values - cusip_2_values
    
    # Filter out NaN values
    valid_mask = ~(np.isnan(spreads))
    if valid_mask.sum() < 2:  # Need at least 2 data points
        return None
    
    valid_spreads = spreads[valid_mask]
    
    # Compute statistics
    last_value = float(valid_spreads[-1])
    average_value = float(np.mean(valid_spreads))
    vs_average = last_value - average_value
    
    # Compute Z score
    if len(valid_spreads) > 1:
        spread_std = float(np.std(valid_spreads, ddof=1))
        if spread_std > 0:
            z_score = vs_average / spread_std
        else:
            z_score = None
    else:
        z_score = None
    
    # Compute percentile (rank of last value)
    percentile = float((valid_spreads <= last_value).sum() / len(valid_spreads) * 100.0)
    
    return (last_value, average_value, vs_average, z_score, percentile)

# ============================================================================
# DATA LOADING FUNCTIONS
# ============================================================================

def get_cad_cusips(historical_path: Path) -> Set[str]:
    """Get CAD CUSIPs from the last date in historical_bond_details.parquet."""
    historical_df = pd.read_parquet(historical_path)
    last_date = historical_df["Date"].max()
    last_date_df = historical_df[historical_df["Date"] == last_date]
    cad_cusips = set(last_date_df[last_date_df["Currency"] == "CAD"]["CUSIP"].unique())
    return cad_cusips

def get_cad_cusips_with_yrs_cvn(historical_path: Path) -> Tuple[Set[str], Dict[str, float]]:
    """Get CAD CUSIPs and Yrs (Cvn) mapping from last date."""
    historical_df = pd.read_parquet(historical_path)
    last_date = historical_df["Date"].max()
    last_date_df = historical_df[historical_df["Date"] == last_date].copy()
    cad_df = last_date_df[last_date_df["Currency"] == "CAD"].copy()
    cad_cusips = set(cad_df["CUSIP"].unique())
    
    valid_df = cad_df[cad_df["Yrs (Cvn)"].notna()].copy()
    valid_df["Yrs (Cvn)"] = pd.to_numeric(valid_df["Yrs (Cvn)"], errors="coerce")
    valid_df = valid_df[valid_df["Yrs (Cvn)"].notna()].copy()
    yrs_cvn_mapping = dict(zip(valid_df["CUSIP"], valid_df["Yrs (Cvn)"].astype(float)))
    
    return cad_cusips, yrs_cvn_mapping

def get_cad_cusips_with_ticker(historical_path: Path) -> Tuple[Set[str], Dict[str, str]]:
    """Get CAD CUSIPs and Ticker mapping from last date."""
    historical_df = pd.read_parquet(historical_path)
    last_date = historical_df["Date"].max()
    last_date_df = historical_df[historical_df["Date"] == last_date].copy()
    cad_df = last_date_df[last_date_df["Currency"] == "CAD"].copy()
    cad_cusips = set(cad_df["CUSIP"].unique())
    
    valid_df = cad_df[cad_df["Ticker"].notna()].copy()
    ticker_mapping = dict(zip(valid_df["CUSIP"], valid_df["Ticker"].astype(str)))
    
    return cad_cusips, ticker_mapping

def get_cad_cusips_with_custom_sector(historical_path: Path) -> Tuple[Set[str], Dict[str, str]]:
    """Get CAD CUSIPs and Custom_Sector mapping from last date."""
    historical_df = pd.read_parquet(historical_path)
    last_date = historical_df["Date"].max()
    last_date_df = historical_df[historical_df["Date"] == last_date].copy()
    cad_df = last_date_df[last_date_df["Currency"] == "CAD"].copy()
    cad_cusips = set(cad_df["CUSIP"].unique())
    
    valid_df = cad_df[cad_df["Custom_Sector"].notna()].copy()
    sector_mapping = dict(zip(valid_df["CUSIP"], valid_df["Custom_Sector"].astype(str)))
    
    return cad_cusips, sector_mapping

def get_currency_ticker_sector_mappings(historical_path: Path) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str], Dict[str, float]]:
    """Get Currency, Ticker, Custom_Sector, and Yrs (Cvn) mappings from last date."""
    historical_df = pd.read_parquet(historical_path)
    last_date = historical_df["Date"].max()
    last_date_df = historical_df[historical_df["Date"] == last_date].copy()
    
    valid_df = last_date_df[
        last_date_df["Currency"].notna() &
        last_date_df["Ticker"].notna() &
        last_date_df["Custom_Sector"].notna() &
        last_date_df["Yrs (Cvn)"].notna()
    ].copy()
    
    currency_mapping = dict(zip(valid_df["CUSIP"], valid_df["Currency"]))
    ticker_mapping = dict(zip(valid_df["CUSIP"], valid_df["Ticker"]))
    sector_mapping = dict(zip(valid_df["CUSIP"], valid_df["Custom_Sector"]))
    yrs_cvn_mapping = dict(zip(valid_df["CUSIP"], valid_df["Yrs (Cvn)"].astype(float)))
    
    return currency_mapping, ticker_mapping, sector_mapping, yrs_cvn_mapping

def find_target_bond_cusips(
    universe_path: Path,
    cad_cusips: Set[str],
    ticker: Optional[str] = None,
    security_name: Optional[str] = None,
) -> Tuple[List[str], Dict[str, str]]:
    """Find CUSIP(s) for target bond(s) by searching Ticker or Security name."""
    universe_df = pd.read_parquet(universe_path)
    universe_df = universe_df[universe_df["CUSIP"].isin(cad_cusips)].copy()
    
    matches = None
    if ticker is not None and ticker.strip():
        matches = universe_df[universe_df["Ticker"] == ticker].copy()
        if len(matches) == 0 and security_name is not None and security_name.strip():
            matches = universe_df[universe_df["Security"] == security_name].copy()
    elif security_name is not None and security_name.strip():
        matches = universe_df[universe_df["Security"] == security_name].copy()
    
    if matches is None or len(matches) == 0:
        raise ValueError("No target bond found")
    
    target_cusips = [str(cusip) for cusip in matches["CUSIP"].tolist()]
    cusip_to_security = dict(zip(matches["CUSIP"], matches["Security"]))
    return target_cusips, cusip_to_security

def load_cr01_holdings(
    runs_today_path: Path,
    portfolio_path: Path,
    include_bid_offer_filter: bool = False,
) -> List[str]:
    """Load CR01 holdings CUSIPs from runs_today.csv matched with portfolio."""
    runs_df = pd.read_csv(runs_today_path)
    runs_df["CR01 @ Tight Bid"] = pd.to_numeric(runs_df["CR01 @ Tight Bid"], errors="coerce")
    
    filter_condition = (runs_df["CR01 @ Tight Bid"] > FILTERS["cr01_tight_bid_threshold"]) & runs_df["CR01 @ Tight Bid"].notna()
    
    if include_bid_offer_filter:
        runs_df["Bid/Offer>3mm"] = pd.to_numeric(runs_df["Bid/Offer>3mm"], errors="coerce")
        filter_condition = filter_condition & (runs_df["Bid/Offer>3mm"] < FILTERS["bid_offer_threshold"]) & runs_df["Bid/Offer>3mm"].notna()
    
    filtered_runs = runs_df[filter_condition].copy()
    runs_cusips = {normalize_cusip(c) for c in filtered_runs["CUSIP"].dropna().unique() if normalize_cusip(c)}
    
    portfolio_df = pd.read_parquet(portfolio_path)
    last_date = portfolio_df["Date"].max()
    last_date_df = portfolio_df[portfolio_df["Date"] == last_date].copy()
    portfolio_cusips = {normalize_cusip(c) for c in last_date_df["CUSIP"].dropna().unique() if normalize_cusip(c)}
    
    matched_cusips = runs_cusips & portfolio_cusips
    return sorted(matched_cusips)

def load_cr01_universe(
    runs_today_path: Path,
    bql_path: Path,
    include_bid_offer_filter: bool = False,
) -> List[str]:
    """Load CR01 universe CUSIPs from runs_today.csv matched with BQL."""
    runs_df = pd.read_csv(runs_today_path)
    runs_df["CR01 @ Wide Offer"] = pd.to_numeric(runs_df["CR01 @ Wide Offer"], errors="coerce")
    
    filter_condition = (runs_df["CR01 @ Wide Offer"] > FILTERS["cr01_wide_offer_threshold"]) & runs_df["CR01 @ Wide Offer"].notna()
    
    if include_bid_offer_filter:
        runs_df["Bid/Offer>3mm"] = pd.to_numeric(runs_df["Bid/Offer>3mm"], errors="coerce")
        filter_condition = filter_condition & (runs_df["Bid/Offer>3mm"] < FILTERS["bid_offer_threshold"]) & runs_df["Bid/Offer>3mm"].notna()
    
    filtered_runs = runs_df[filter_condition].copy()
    runs_cusips = {normalize_cusip(c) for c in filtered_runs["CUSIP"].dropna().unique() if normalize_cusip(c)}
    
    bql_df = pd.read_parquet(bql_path)
    bql_cusips = {normalize_cusip(c) for c in bql_df["CUSIP"].dropna().unique() if normalize_cusip(c)}
    
    matched_cusips = runs_cusips & bql_cusips
    return sorted(matched_cusips)

def load_all_portfolio_cusips(portfolio_path: Path) -> List[str]:
    """Load all CUSIPs from portfolio (last date only)."""
    portfolio_df = pd.read_parquet(portfolio_path)
    last_date = portfolio_df["Date"].max()
    last_date_df = portfolio_df[portfolio_df["Date"] == last_date].copy()
    portfolio_cusips = {normalize_cusip(c) for c in last_date_df["CUSIP"].dropna().unique() if normalize_cusip(c)}
    return sorted(portfolio_cusips)

# ============================================================================
# ANALYSIS FUNCTIONS
# ============================================================================

def run_all_comb_analysis(
    bql_path: Path,
    historical_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run all combinations analysis (all CAD CUSIPs)."""
    try:
        validation_log.write(f"[all_comb] Starting analysis...\n")
        cad_cusips = get_cad_cusips(historical_path)
        validation_log.write(f"[all_comb] Found {len(cad_cusips)} CAD CUSIPs\n")
        
        if not cad_cusips:
            raise ValueError("No CAD CUSIPs found in historical data")
        
        data = pd.read_parquet(bql_path)
        data = data[data["CUSIP"].isin(cad_cusips)].copy()
        data = data.dropna()
        
        if data.empty:
            raise ValueError("No BQL data found for CAD CUSIPs")
        
        filtered_data = filter_recent_cusips(data, FILTERS["recent_date_percent"])
        validation_log.write(f"[all_comb] After filtering: {len(filtered_data)} rows, {filtered_data['CUSIP'].nunique()} CUSIPs\n")
        
        if filtered_data.empty:
            raise ValueError("No data remaining after filtering for recent CUSIPs")
        
        name_lookup = build_name_lookup(filtered_data)
        wide_values = filtered_data.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        cusips = wide_values.columns.tolist()
        values_array = wide_values.values
        summaries: List[PairSummary] = []
        
        num_pairs = len(cusips) * (len(cusips) - 1) // 2
        batch_size = max(1000, num_pairs // 100)
        processed = 0
        
        for i, j in combinations(range(len(cusips)), 2):
            cusip_1_vals = values_array[:, i]
            cusip_2_vals = values_array[:, j]
            both_valid = ~(np.isnan(cusip_1_vals) | np.isnan(cusip_2_vals))
            
            if both_valid.sum() < 2:
                continue
            
            stats = compute_pair_stats_vectorized(cusip_1_vals[both_valid], cusip_2_vals[both_valid])
            if stats is None or stats[3] is None:
                continue
            
            summaries.append(PairSummary(
                Bond_1=name_lookup.get(cusips[i], cusips[i]),
                Bond_2=name_lookup.get(cusips[j], cusips[j]),
                last_value=stats[0],
                average_value=stats[1],
                vs_average=stats[2],
                z_score=stats[3],
                percentile=stats[4],
                cusip_1=cusips[i],
                cusip_2=cusips[j],
            ))
            
            processed += 1
            if processed % batch_size == 0:
                validation_log.write(f"[all_comb] Processed {processed:,} / {num_pairs:,} pairs ({processed*100/num_pairs:.1f}%)\n")
        
        results_df = pd.DataFrame([{
            "Bond_1": p.Bond_1, "Bond_2": p.Bond_2, "Last": p.last_value,
            "Avg": p.average_value, "vs Avg": p.vs_average, "Z Score": p.z_score,
            "Percentile": p.percentile, "cusip_1": p.cusip_1, "cusip_2": p.cusip_2,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["Bond_1"] = results_df["Bond_1"].map(ensure_ascii)
        results_df["Bond_2"] = results_df["Bond_2"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[all_comb] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[all_comb] ERROR: {str(e)}\n")
        return pd.DataFrame()

def run_term_comb_analysis(
    bql_path: Path,
    historical_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run term combinations analysis (Yrs Cvn diff <= 0.8)."""
    try:
        validation_log.write(f"[term_comb] Starting analysis...\n")
        cad_cusips, yrs_cvn_mapping = get_cad_cusips_with_yrs_cvn(historical_path)
        validation_log.write(f"[term_comb] Found {len(cad_cusips)} CAD CUSIPs, {len(yrs_cvn_mapping)} with Yrs (Cvn) data\n")
        
        if not cad_cusips or not yrs_cvn_mapping:
            raise ValueError("No CAD CUSIPs with valid Yrs (Cvn) data found")
        
        data = pd.read_parquet(bql_path)
        valid_cusips = set(yrs_cvn_mapping.keys())
        data = data[data["CUSIP"].isin(valid_cusips)].copy()
        data = data.dropna()
        
        if data.empty:
            raise ValueError("No BQL data found for CAD CUSIPs with valid Yrs (Cvn) data")
        
        filtered_data = filter_recent_cusips(data, FILTERS["recent_date_percent"])
        validation_log.write(f"[term_comb] After filtering: {len(filtered_data)} rows, {filtered_data['CUSIP'].nunique()} CUSIPs\n")
        
        if filtered_data.empty:
            raise ValueError("No data remaining after filtering for recent CUSIPs")
        
        name_lookup = build_name_lookup(filtered_data)
        wide_values = filtered_data.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        cusips = [c for c in wide_values.columns.tolist() if c in yrs_cvn_mapping]
        if not cusips:
            raise ValueError("No CUSIPs remaining after Yrs (Cvn) filtering")
        
        valid_pairs: List[Tuple[int, int]] = []
        for i, j in combinations(range(len(cusips)), 2):
            cusip_1 = cusips[i]
            cusip_2 = cusips[j]
            yrs_cvn_1 = yrs_cvn_mapping.get(cusip_1)
            yrs_cvn_2 = yrs_cvn_mapping.get(cusip_2)
            
            if yrs_cvn_1 is None or yrs_cvn_2 is None:
                continue
            
            if abs(yrs_cvn_1 - yrs_cvn_2) <= FILTERS["max_yrs_cvn_diff"]:
                valid_pairs.append((i, j))
        
        validation_log.write(f"[term_comb] Found {len(valid_pairs):,} valid pairs (out of {len(cusips) * (len(cusips) - 1) // 2:,} possible)\n")
        
        if not valid_pairs:
            raise ValueError(f"No pairs found with Yrs (Cvn) difference <= {FILTERS['max_yrs_cvn_diff']}")
        
        values_array = wide_values[cusips].values
        summaries: List[PairSummary] = []
        
        batch_size = max(1000, len(valid_pairs) // 100)
        processed = 0
        
        for i, j in valid_pairs:
            cusip_1_vals = values_array[:, i]
            cusip_2_vals = values_array[:, j]
            both_valid = ~(np.isnan(cusip_1_vals) | np.isnan(cusip_2_vals))
            
            if both_valid.sum() < 2:
                continue
            
            stats = compute_pair_stats_vectorized(cusip_1_vals[both_valid], cusip_2_vals[both_valid])
            if stats is None or stats[3] is None:
                continue
            
            summaries.append(PairSummary(
                Bond_1=name_lookup.get(cusips[i], cusips[i]),
                Bond_2=name_lookup.get(cusips[j], cusips[j]),
                last_value=stats[0],
                average_value=stats[1],
                vs_average=stats[2],
                z_score=stats[3],
                percentile=stats[4],
                cusip_1=cusips[i],
                cusip_2=cusips[j],
            ))
            
            processed += 1
            if processed % batch_size == 0:
                validation_log.write(f"[term_comb] Processed {processed:,} / {len(valid_pairs):,} pairs ({processed*100/len(valid_pairs):.1f}%)\n")
        
        results_df = pd.DataFrame([{
            "Bond_1": p.Bond_1, "Bond_2": p.Bond_2, "Last": p.last_value,
            "Avg": p.average_value, "vs Avg": p.vs_average, "Z Score": p.z_score,
            "Percentile": p.percentile, "cusip_1": p.cusip_1, "cusip_2": p.cusip_2,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["Bond_1"] = results_df["Bond_1"].map(ensure_ascii)
        results_df["Bond_2"] = results_df["Bond_2"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[term_comb] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[term_comb] ERROR: {str(e)}\n")
        return pd.DataFrame()

def run_ticker_comb_analysis(
    bql_path: Path,
    historical_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run ticker combinations analysis (matching Ticker values)."""
    try:
        validation_log.write(f"[ticker_comb] Starting analysis...\n")
        cad_cusips, ticker_mapping = get_cad_cusips_with_ticker(historical_path)
        validation_log.write(f"[ticker_comb] Found {len(cad_cusips)} CAD CUSIPs, {len(ticker_mapping)} with Ticker data\n")
        
        if not cad_cusips or not ticker_mapping:
            raise ValueError("No CAD CUSIPs with valid Ticker data found")
        
        data = pd.read_parquet(bql_path)
        valid_cusips = set(ticker_mapping.keys())
        data = data[data["CUSIP"].isin(valid_cusips)].copy()
        data = data.dropna()
        
        if data.empty:
            raise ValueError("No BQL data found for CAD CUSIPs with valid Ticker data")
        
        filtered_data = filter_recent_cusips(data, FILTERS["recent_date_percent"])
        validation_log.write(f"[ticker_comb] After filtering: {len(filtered_data)} rows, {filtered_data['CUSIP'].nunique()} CUSIPs\n")
        
        if filtered_data.empty:
            raise ValueError("No data remaining after filtering for recent CUSIPs")
        
        name_lookup = build_name_lookup(filtered_data)
        wide_values = filtered_data.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        cusips = [c for c in wide_values.columns.tolist() if c in ticker_mapping]
        if not cusips:
            raise ValueError("No CUSIPs remaining after Ticker filtering")
        
        valid_pairs: List[Tuple[int, int]] = []
        for i, j in combinations(range(len(cusips)), 2):
            cusip_1 = cusips[i]
            cusip_2 = cusips[j]
            ticker_1 = ticker_mapping.get(cusip_1)
            ticker_2 = ticker_mapping.get(cusip_2)
            
            if ticker_1 is None or ticker_2 is None:
                continue
            
            if ticker_1 == ticker_2:
                valid_pairs.append((i, j))
        
        validation_log.write(f"[ticker_comb] Found {len(valid_pairs):,} valid pairs (out of {len(cusips) * (len(cusips) - 1) // 2:,} possible)\n")
        
        if not valid_pairs:
            raise ValueError("No pairs found with matching Ticker values")
        
        values_array = wide_values[cusips].values
        summaries: List[PairSummary] = []
        
        batch_size = max(1000, len(valid_pairs) // 100)
        processed = 0
        
        for i, j in valid_pairs:
            cusip_1_vals = values_array[:, i]
            cusip_2_vals = values_array[:, j]
            both_valid = ~(np.isnan(cusip_1_vals) | np.isnan(cusip_2_vals))
            
            if both_valid.sum() < 2:
                continue
            
            stats = compute_pair_stats_vectorized(cusip_1_vals[both_valid], cusip_2_vals[both_valid])
            if stats is None or stats[3] is None:
                continue
            
            summaries.append(PairSummary(
                Bond_1=name_lookup.get(cusips[i], cusips[i]),
                Bond_2=name_lookup.get(cusips[j], cusips[j]),
                last_value=stats[0],
                average_value=stats[1],
                vs_average=stats[2],
                z_score=stats[3],
                percentile=stats[4],
                cusip_1=cusips[i],
                cusip_2=cusips[j],
            ))
            
            processed += 1
            if processed % batch_size == 0:
                validation_log.write(f"[ticker_comb] Processed {processed:,} / {len(valid_pairs):,} pairs ({processed*100/len(valid_pairs):.1f}%)\n")
        
        results_df = pd.DataFrame([{
            "Bond_1": p.Bond_1, "Bond_2": p.Bond_2, "Last": p.last_value,
            "Avg": p.average_value, "vs Avg": p.vs_average, "Z Score": p.z_score,
            "Percentile": p.percentile, "cusip_1": p.cusip_1, "cusip_2": p.cusip_2,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["Bond_1"] = results_df["Bond_1"].map(ensure_ascii)
        results_df["Bond_2"] = results_df["Bond_2"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[ticker_comb] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[ticker_comb] ERROR: {str(e)}\n")
        return pd.DataFrame()

def run_custom_sector_analysis(
    bql_path: Path,
    historical_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run custom sector combinations analysis (matching Custom_Sector values)."""
    try:
        validation_log.write(f"[custom_sector] Starting analysis...\n")
        cad_cusips, sector_mapping = get_cad_cusips_with_custom_sector(historical_path)
        validation_log.write(f"[custom_sector] Found {len(cad_cusips)} CAD CUSIPs, {len(sector_mapping)} with Custom_Sector data\n")
        
        if not cad_cusips or not sector_mapping:
            raise ValueError("No CAD CUSIPs with valid Custom_Sector data found")
        
        data = pd.read_parquet(bql_path)
        valid_cusips = set(sector_mapping.keys())
        data = data[data["CUSIP"].isin(valid_cusips)].copy()
        data = data.dropna()
        
        if data.empty:
            raise ValueError("No BQL data found for CAD CUSIPs with valid Custom_Sector data")
        
        filtered_data = filter_recent_cusips(data, FILTERS["recent_date_percent"])
        validation_log.write(f"[custom_sector] After filtering: {len(filtered_data)} rows, {filtered_data['CUSIP'].nunique()} CUSIPs\n")
        
        if filtered_data.empty:
            raise ValueError("No data remaining after filtering for recent CUSIPs")
        
        name_lookup = build_name_lookup(filtered_data)
        wide_values = filtered_data.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        cusips = [c for c in wide_values.columns.tolist() if c in sector_mapping]
        if not cusips:
            raise ValueError("No CUSIPs remaining after Custom_Sector filtering")
        
        valid_pairs: List[Tuple[int, int]] = []
        for i, j in combinations(range(len(cusips)), 2):
            cusip_1 = cusips[i]
            cusip_2 = cusips[j]
            sector_1 = sector_mapping.get(cusip_1)
            sector_2 = sector_mapping.get(cusip_2)
            
            if sector_1 is None or sector_2 is None:
                continue
            
            if sector_1 == sector_2:
                valid_pairs.append((i, j))
        
        validation_log.write(f"[custom_sector] Found {len(valid_pairs):,} valid pairs (out of {len(cusips) * (len(cusips) - 1) // 2:,} possible)\n")
        
        if not valid_pairs:
            raise ValueError("No pairs found with matching Custom_Sector values")
        
        values_array = wide_values[cusips].values
        summaries: List[PairSummary] = []
        
        batch_size = max(1000, len(valid_pairs) // 100)
        processed = 0
        
        for i, j in valid_pairs:
            cusip_1_vals = values_array[:, i]
            cusip_2_vals = values_array[:, j]
            both_valid = ~(np.isnan(cusip_1_vals) | np.isnan(cusip_2_vals))
            
            if both_valid.sum() < 2:
                continue
            
            stats = compute_pair_stats_vectorized(cusip_1_vals[both_valid], cusip_2_vals[both_valid])
            if stats is None or stats[3] is None:
                continue
            
            summaries.append(PairSummary(
                Bond_1=name_lookup.get(cusips[i], cusips[i]),
                Bond_2=name_lookup.get(cusips[j], cusips[j]),
                last_value=stats[0],
                average_value=stats[1],
                vs_average=stats[2],
                z_score=stats[3],
                percentile=stats[4],
                cusip_1=cusips[i],
                cusip_2=cusips[j],
            ))
            
            processed += 1
            if processed % batch_size == 0:
                validation_log.write(f"[custom_sector] Processed {processed:,} / {len(valid_pairs):,} pairs ({processed*100/len(valid_pairs):.1f}%)\n")
        
        results_df = pd.DataFrame([{
            "Bond_1": p.Bond_1, "Bond_2": p.Bond_2, "Last": p.last_value,
            "Avg": p.average_value, "vs Avg": p.vs_average, "Z Score": p.z_score,
            "Percentile": p.percentile, "cusip_1": p.cusip_1, "cusip_2": p.cusip_2,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["Bond_1"] = results_df["Bond_1"].map(ensure_ascii)
        results_df["Bond_2"] = results_df["Bond_2"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[custom_sector] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[custom_sector] ERROR: {str(e)}\n")
        return pd.DataFrame()

def run_custom_bond_comb_analysis(
    bql_path: Path,
    historical_path: Path,
    universe_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run custom bond combinations analysis (target bond vs all CAD CUSIPs)."""
    try:
        validation_log.write(f"[custom_bond_comb] Starting analysis...\n")
        cad_cusips = get_cad_cusips(historical_path)
        validation_log.write(f"[custom_bond_comb] Found {len(cad_cusips)} CAD CUSIPs\n")
        
        if not cad_cusips:
            raise ValueError("No CAD CUSIPs found in historical data")
        
        target_cusips, cusip_to_security = find_target_bond_cusips(
            universe_path, cad_cusips, TARGET_BOND["ticker"], TARGET_BOND["security"]
        )
        validation_log.write(f"[custom_bond_comb] Found {len(target_cusips)} target bond(s)\n")
        
        data = pd.read_parquet(bql_path)
        data = data[data["CUSIP"].isin(cad_cusips)].copy()
        data = data.dropna()
        
        if data.empty:
            raise ValueError("No BQL data found for CAD CUSIPs")
        
        filtered_data = filter_recent_cusips(data, FILTERS["recent_date_percent"])
        validation_log.write(f"[custom_bond_comb] After filtering: {len(filtered_data)} rows, {filtered_data['CUSIP'].nunique()} CUSIPs\n")
        
        if filtered_data.empty:
            raise ValueError("No data remaining after filtering for recent CUSIPs")
        
        target_cusips_in_data = [c for c in target_cusips if c in filtered_data["CUSIP"].values]
        if not target_cusips_in_data:
            raise ValueError(f"Target bond(s) not present in recent dates")
        target_cusips = target_cusips_in_data
        
        name_lookup = build_name_lookup(filtered_data)
        wide_values = filtered_data.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        cusips = wide_values.columns.tolist()
        values_array = wide_values.values
        summaries: List[PairSummary] = []
        
        for target_cusip in target_cusips:
            target_idx = cusips.index(target_cusip)
            target_values = values_array[:, target_idx]
            target_security_name = cusip_to_security.get(target_cusip, "")
            target_bond_name = name_lookup.get(target_cusip, target_security_name)
            
            for i, cusip_1 in enumerate(cusips):
                cusip_1_vals = values_array[:, i]
                both_valid = ~(np.isnan(cusip_1_vals) | np.isnan(target_values))
                
                if both_valid.sum() < 2:
                    continue
                
                stats = compute_pair_stats_vectorized(cusip_1_vals[both_valid], target_values[both_valid])
                if stats is None or stats[3] is None:
                    continue
                
                summaries.append(PairSummary(
                    Bond_1=name_lookup.get(cusip_1, cusip_1),
                    Bond_2=target_bond_name,
                    last_value=stats[0],
                    average_value=stats[1],
                    vs_average=stats[2],
                    z_score=stats[3],
                    percentile=stats[4],
                    cusip_1=cusip_1,
                    cusip_2=target_cusip,
                ))
        
        results_df = pd.DataFrame([{
            "Bond_1": p.Bond_1, "Bond_2": p.Bond_2, "Last": p.last_value,
            "Avg": p.average_value, "vs Avg": p.vs_average, "Z Score": p.z_score,
            "Percentile": p.percentile, "cusip_1": p.cusip_1, "cusip_2": p.cusip_2,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["Bond_1"] = results_df["Bond_1"].map(ensure_ascii)
        results_df["Bond_2"] = results_df["Bond_2"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[custom_bond_comb] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[custom_bond_comb] ERROR: {str(e)}\n")
        return pd.DataFrame()

def run_custom_bond_vs_holdings_analysis(
    bql_path: Path,
    historical_path: Path,
    universe_path: Path,
    runs_today_path: Path,
    portfolio_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run custom bond vs holdings analysis (target bond vs CR01 holdings)."""
    try:
        validation_log.write(f"[custom_bond_vs_holdings] Starting analysis...\n")
        cad_cusips = get_cad_cusips(historical_path)
        validation_log.write(f"[custom_bond_vs_holdings] Found {len(cad_cusips)} CAD CUSIPs\n")
        
        if not cad_cusips:
            raise ValueError("No CAD CUSIPs found in historical data")
        
        target_cusips, cusip_to_security = find_target_bond_cusips(
            universe_path, cad_cusips, TARGET_BOND["ticker"], TARGET_BOND["security"]
        )
        validation_log.write(f"[custom_bond_vs_holdings] Found {len(target_cusips)} target bond(s)\n")
        
        holdings_cusips_raw = load_cr01_holdings(runs_today_path, portfolio_path)
        validation_log.write(f"[custom_bond_vs_holdings] Found {len(holdings_cusips_raw)} CR01 holdings CUSIPs\n")
        
        bql_data = pd.read_parquet(bql_path)
        combined_cusips = set(target_cusips) | set(holdings_cusips_raw)
        filtered_bql = bql_data[bql_data["CUSIP"].isin(combined_cusips)].copy()
        filtered_bql = filtered_bql.dropna()
        
        if filtered_bql.empty:
            raise ValueError("No BQL data found for target bonds and holdings")
        
        filtered_bql = filter_recent_cusips(filtered_bql, FILTERS["recent_date_percent"])
        validation_log.write(f"[custom_bond_vs_holdings] After filtering: {len(filtered_bql)} rows, {filtered_bql['CUSIP'].nunique()} CUSIPs\n")
        
        if filtered_bql.empty:
            raise ValueError("No data remaining after filtering for recent CUSIPs")
        
        target_cusips_in_data = [c for c in target_cusips if c in filtered_bql["CUSIP"].values]
        if not target_cusips_in_data:
            raise ValueError(f"Target bond(s) not present in recent dates")
        target_cusips = target_cusips_in_data
        
        holdings_cusips = [c for c in holdings_cusips_raw if c in filtered_bql["CUSIP"].values]
        if not holdings_cusips:
            raise ValueError("No holdings CUSIPs found in recent BQL data")
        
        name_lookup = build_name_lookup(filtered_bql)
        wide_values = filtered_bql.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        values_array = wide_values.values
        cusip_to_index = {cusip: i for i, cusip in enumerate(wide_values.columns)}
        summaries: List[PairSummary] = []
        
        for target_cusip in target_cusips:
            target_idx = cusip_to_index[target_cusip]
            target_values = values_array[:, target_idx]
            target_security_name = cusip_to_security.get(target_cusip, "")
            target_bond_name = name_lookup.get(target_cusip, target_security_name)
            
            for holdings_cusip in holdings_cusips:
                holdings_idx = cusip_to_index[holdings_cusip]
                holdings_values = values_array[:, holdings_idx]
                both_valid = ~(np.isnan(target_values) | np.isnan(holdings_values))
                
                if both_valid.sum() < 2:
                    continue
                
                stats = compute_pair_stats_vectorized(target_values[both_valid], holdings_values[both_valid])
                if stats is None or stats[3] is None:
                    continue
                
                summaries.append(PairSummary(
                    Bond_1=target_bond_name,
                    Bond_2=name_lookup.get(holdings_cusip, holdings_cusip),
                    last_value=stats[0],
                    average_value=stats[1],
                    vs_average=stats[2],
                    z_score=stats[3],
                    percentile=stats[4],
                    cusip_1=target_cusip,
                    cusip_2=holdings_cusip,
                ))
        
        results_df = pd.DataFrame([{
            "Bond_1": p.Bond_1, "Bond_2": p.Bond_2, "Last": p.last_value,
            "Avg": p.average_value, "vs Avg": p.vs_average, "Z Score": p.z_score,
            "Percentile": p.percentile, "cusip_1": p.cusip_1, "cusip_2": p.cusip_2,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["Bond_1"] = results_df["Bond_1"].map(ensure_ascii)
        results_df["Bond_2"] = results_df["Bond_2"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[custom_bond_vs_holdings] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[custom_bond_vs_holdings] ERROR: {str(e)}\n")
        return pd.DataFrame()

def run_cad_cheap_vs_usd_analysis(
    bql_path: Path,
    historical_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run CAD cheap vs USD analysis (CAD/USD pairs with matching criteria)."""
    try:
        validation_log.write(f"[cad_cheap_vs_usd] Starting analysis...\n")
        data = pd.read_parquet(bql_path)
        data = data.dropna()
        
        if data.empty:
            raise ValueError("No BQL data found")
        
        filtered_data = filter_recent_cusips(data, FILTERS["recent_date_percent"])
        validation_log.write(f"[cad_cheap_vs_usd] After filtering: {len(filtered_data)} rows, {filtered_data['CUSIP'].nunique()} CUSIPs\n")
        
        if filtered_data.empty:
            raise ValueError("No data remaining after filtering for recent CUSIPs")
        
        name_lookup = build_name_lookup(filtered_data)
        wide_values = filtered_data.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        currency_mapping, ticker_mapping, sector_mapping, yrs_cvn_mapping = get_currency_ticker_sector_mappings(historical_path)
        validation_log.write(f"[cad_cheap_vs_usd] Found {len(currency_mapping)} CUSIPs with valid mappings\n")
        
        cusips = wide_values.columns.tolist()
        valid_cusips = [c for c in cusips if c in currency_mapping and c in ticker_mapping and c in sector_mapping and c in yrs_cvn_mapping]
        
        cad_cusips = [c for c in valid_cusips if currency_mapping.get(c) == "CAD"]
        usd_cusips = [c for c in valid_cusips if currency_mapping.get(c) == "USD"]
        validation_log.write(f"[cad_cheap_vs_usd] Found {len(cad_cusips)} CAD and {len(usd_cusips)} USD CUSIPs\n")
        
        if not cad_cusips or not usd_cusips:
            raise ValueError("Need both CAD and USD CUSIPs")
        
        cusip_to_idx = {cusip: i for i, cusip in enumerate(cusips)}
        valid_pairs: List[Tuple[int, int]] = []
        
        for cusip_1 in cad_cusips:
            idx_1 = cusip_to_idx[cusip_1]
            ticker_1 = ticker_mapping.get(cusip_1)
            sector_1 = sector_mapping.get(cusip_1)
            yrs_cvn_1 = yrs_cvn_mapping.get(cusip_1)
            
            if ticker_1 is None or sector_1 is None or yrs_cvn_1 is None:
                continue
            
            for cusip_2 in usd_cusips:
                idx_2 = cusip_to_idx[cusip_2]
                ticker_2 = ticker_mapping.get(cusip_2)
                sector_2 = sector_mapping.get(cusip_2)
                yrs_cvn_2 = yrs_cvn_mapping.get(cusip_2)
                
                if ticker_2 is None or sector_2 is None or yrs_cvn_2 is None:
                    continue
                
                if (ticker_1 == ticker_2 and sector_1 == sector_2 and abs(yrs_cvn_1 - yrs_cvn_2) <= FILTERS["max_yrs_cvn_diff_cad_usd"]):
                    valid_pairs.append((idx_1, idx_2))
        
        validation_log.write(f"[cad_cheap_vs_usd] Found {len(valid_pairs):,} valid pairs\n")
        
        if not valid_pairs:
            raise ValueError("No pairs found with matching criteria")
        
        values_array = wide_values.values
        summaries: List[PairSummary] = []
        
        for idx_1, idx_2 in valid_pairs:
            cusip_1_vals = values_array[:, idx_1]
            cusip_2_vals = values_array[:, idx_2]
            both_valid = ~(np.isnan(cusip_1_vals) | np.isnan(cusip_2_vals))
            
            if both_valid.sum() < 2:
                continue
            
            stats = compute_pair_stats_vectorized(cusip_1_vals[both_valid], cusip_2_vals[both_valid])
            if stats is None or stats[3] is None:
                continue
            
            summaries.append(PairSummary(
                Bond_1=name_lookup.get(cusips[idx_1], cusips[idx_1]),
                Bond_2=name_lookup.get(cusips[idx_2], cusips[idx_2]),
                last_value=stats[0],
                average_value=stats[1],
                vs_average=stats[2],
                z_score=stats[3],
                percentile=stats[4],
                cusip_1=cusips[idx_1],
                cusip_2=cusips[idx_2],
            ))
        
        results_df = pd.DataFrame([{
            "Bond_1": p.Bond_1, "Bond_2": p.Bond_2, "Last": p.last_value,
            "Avg": p.average_value, "vs Avg": p.vs_average, "Z Score": p.z_score,
            "Percentile": p.percentile, "cusip_1": p.cusip_1, "cusip_2": p.cusip_2,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["Bond_1"] = results_df["Bond_1"].map(ensure_ascii)
        results_df["Bond_2"] = results_df["Bond_2"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[cad_cheap_vs_usd] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[cad_cheap_vs_usd] ERROR: {str(e)}\n")
        return pd.DataFrame()

def run_cad_rich_vs_usd_analysis(
    bql_path: Path,
    historical_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run CAD rich vs USD analysis (USD/CAD pairs with matching criteria)."""
    try:
        validation_log.write(f"[cad_rich_vs_usd] Starting analysis...\n")
        data = pd.read_parquet(bql_path)
        data = data.dropna()
        
        if data.empty:
            raise ValueError("No BQL data found")
        
        filtered_data = filter_recent_cusips(data, FILTERS["recent_date_percent"])
        validation_log.write(f"[cad_rich_vs_usd] After filtering: {len(filtered_data)} rows, {filtered_data['CUSIP'].nunique()} CUSIPs\n")
        
        if filtered_data.empty:
            raise ValueError("No data remaining after filtering for recent CUSIPs")
        
        name_lookup = build_name_lookup(filtered_data)
        wide_values = filtered_data.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        currency_mapping, ticker_mapping, sector_mapping, yrs_cvn_mapping = get_currency_ticker_sector_mappings(historical_path)
        validation_log.write(f"[cad_rich_vs_usd] Found {len(currency_mapping)} CUSIPs with valid mappings\n")
        
        cusips = wide_values.columns.tolist()
        valid_cusips = [c for c in cusips if c in currency_mapping and c in ticker_mapping and c in sector_mapping and c in yrs_cvn_mapping]
        
        usd_cusips = [c for c in valid_cusips if currency_mapping.get(c) == "USD"]
        cad_cusips = [c for c in valid_cusips if currency_mapping.get(c) == "CAD"]
        validation_log.write(f"[cad_rich_vs_usd] Found {len(usd_cusips)} USD and {len(cad_cusips)} CAD CUSIPs\n")
        
        if not usd_cusips or not cad_cusips:
            raise ValueError("Need both USD and CAD CUSIPs")
        
        cusip_to_idx = {cusip: i for i, cusip in enumerate(cusips)}
        valid_pairs: List[Tuple[int, int]] = []
        
        for cusip_1 in usd_cusips:
            idx_1 = cusip_to_idx[cusip_1]
            ticker_1 = ticker_mapping.get(cusip_1)
            sector_1 = sector_mapping.get(cusip_1)
            yrs_cvn_1 = yrs_cvn_mapping.get(cusip_1)
            
            if ticker_1 is None or sector_1 is None or yrs_cvn_1 is None:
                continue
            
            for cusip_2 in cad_cusips:
                idx_2 = cusip_to_idx[cusip_2]
                ticker_2 = ticker_mapping.get(cusip_2)
                sector_2 = sector_mapping.get(cusip_2)
                yrs_cvn_2 = yrs_cvn_mapping.get(cusip_2)
                
                if ticker_2 is None or sector_2 is None or yrs_cvn_2 is None:
                    continue
                
                if (ticker_1 == ticker_2 and sector_1 == sector_2 and abs(yrs_cvn_1 - yrs_cvn_2) <= FILTERS["max_yrs_cvn_diff_cad_usd"]):
                    valid_pairs.append((idx_1, idx_2))
        
        validation_log.write(f"[cad_rich_vs_usd] Found {len(valid_pairs):,} valid pairs\n")
        
        if not valid_pairs:
            raise ValueError("No pairs found with matching criteria")
        
        values_array = wide_values.values
        summaries: List[PairSummary] = []
        
        for idx_1, idx_2 in valid_pairs:
            cusip_1_vals = values_array[:, idx_1]
            cusip_2_vals = values_array[:, idx_2]
            both_valid = ~(np.isnan(cusip_1_vals) | np.isnan(cusip_2_vals))
            
            if both_valid.sum() < 2:
                continue
            
            stats = compute_pair_stats_vectorized(cusip_1_vals[both_valid], cusip_2_vals[both_valid])
            if stats is None or stats[3] is None:
                continue
            
            summaries.append(PairSummary(
                Bond_1=name_lookup.get(cusips[idx_1], cusips[idx_1]),
                Bond_2=name_lookup.get(cusips[idx_2], cusips[idx_2]),
                last_value=stats[0],
                average_value=stats[1],
                vs_average=stats[2],
                z_score=stats[3],
                percentile=stats[4],
                cusip_1=cusips[idx_1],
                cusip_2=cusips[idx_2],
            ))
        
        results_df = pd.DataFrame([{
            "Bond_1": p.Bond_1, "Bond_2": p.Bond_2, "Last": p.last_value,
            "Avg": p.average_value, "vs Avg": p.vs_average, "Z Score": p.z_score,
            "Percentile": p.percentile, "cusip_1": p.cusip_1, "cusip_2": p.cusip_2,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["Bond_1"] = results_df["Bond_1"].map(ensure_ascii)
        results_df["Bond_2"] = results_df["Bond_2"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[cad_rich_vs_usd] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[cad_rich_vs_usd] ERROR: {str(e)}\n")
        return pd.DataFrame()

def compute_cr01_pair_summary(
    wide_values: pd.DataFrame,
    universe_cusip: str,
    holdings_cusip: str,
    name_lookup: Dict[str, str],
) -> Optional[CR01PairSummary]:
    """Compute summary statistics for a single universe/holdings pair."""
    if universe_cusip not in wide_values or holdings_cusip not in wide_values:
        return None
    
    pair_frame = wide_values[[universe_cusip, holdings_cusip]].dropna()
    if pair_frame.empty:
        return None
    
    if universe_cusip == holdings_cusip:
        spreads = pd.Series(0.0, index=pair_frame.index)
    else:
        spreads = pair_frame[universe_cusip] - pair_frame[holdings_cusip]
    
    last_value = float(spreads.iloc[-1]) if len(spreads) > 0 else 0.0
    average_value = float(spreads.mean())
    vs_average = last_value - average_value
    spread_array = spreads.to_numpy(dtype=float, copy=False)
    
    if spread_array.size > 1:
        spread_std = float(np.std(spread_array, ddof=1))
    else:
        spread_std = math.nan
    
    if not math.isnan(spread_std) and spread_std > 0:
        z_score = vs_average / spread_std
    else:
        z_score = None
    
    percentile = float(spreads.rank(pct=True).iloc[-1] * 100.0)
    
    return CR01PairSummary(
        universe_name=name_lookup.get(universe_cusip, universe_cusip),
        holdings_name=name_lookup.get(holdings_cusip, holdings_cusip),
        last_value=last_value,
        average_value=average_value,
        vs_average=vs_average,
        z_score=z_score,
        percentile=percentile,
        universe_cusip=universe_cusip,
        holdings_cusip=holdings_cusip,
    )

def run_executable_cr01_vs_holdings_analysis(
    bql_path: Path,
    runs_today_path: Path,
    portfolio_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run executable CR01 vs holdings analysis."""
    try:
        validation_log.write(f"[executable_cr01_vs_holdings] Starting analysis...\n")
        holdings_cusips_raw = load_cr01_holdings(runs_today_path, portfolio_path, include_bid_offer_filter=False)
        validation_log.write(f"[executable_cr01_vs_holdings] Found {len(holdings_cusips_raw)} CR01 holdings CUSIPs\n")
        
        universe_cusips_raw = load_cr01_universe(runs_today_path, bql_path, include_bid_offer_filter=False)
        validation_log.write(f"[executable_cr01_vs_holdings] Found {len(universe_cusips_raw)} CR01 universe CUSIPs\n")
        
        combined_cusips = sorted(set(universe_cusips_raw) | set(holdings_cusips_raw))
        
        data = pd.read_parquet(bql_path)
        filtered = data[data["CUSIP"].isin(combined_cusips)].copy()
        filtered = filtered.dropna(subset=["Value"])
        
        if filtered.empty:
            raise ValueError("No BQL data found for CR01 CUSIPs")
        
        name_lookup = build_name_lookup(filtered)
        wide_values = filtered.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        summaries: List[CR01PairSummary] = []
        for universe_cusip, holdings_cusip in product(universe_cusips_raw, holdings_cusips_raw):
            summary = compute_cr01_pair_summary(wide_values, universe_cusip, holdings_cusip, name_lookup)
            if summary is not None:
                summaries.append(summary)
        
        results_df = pd.DataFrame([{
            "universe_name": p.universe_name,
            "holdings_name": p.holdings_name,
            "Last": p.last_value,
            "Avg": p.average_value,
            "vs Avg": p.vs_average,
            "Z Score": p.z_score,
            "Percentile": p.percentile,
            "universe_cusip": p.universe_cusip,
            "holdings_cusip": p.holdings_cusip,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["universe_name"] = results_df["universe_name"].map(ensure_ascii)
        results_df["holdings_name"] = results_df["holdings_name"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[executable_cr01_vs_holdings] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[executable_cr01_vs_holdings] ERROR: {str(e)}\n")
        return pd.DataFrame()

def run_executable_cr01_decent_bid_offer_vs_holdings_analysis(
    bql_path: Path,
    runs_today_path: Path,
    portfolio_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run executable CR01 decent bid offer vs holdings analysis."""
    try:
        validation_log.write(f"[executable_cr01_decent_bid_offer_vs_holdings] Starting analysis...\n")
        holdings_cusips_raw = load_cr01_holdings(runs_today_path, portfolio_path, include_bid_offer_filter=True)
        validation_log.write(f"[executable_cr01_decent_bid_offer_vs_holdings] Found {len(holdings_cusips_raw)} CR01 holdings CUSIPs\n")
        
        universe_cusips_raw = load_cr01_universe(runs_today_path, bql_path, include_bid_offer_filter=True)
        validation_log.write(f"[executable_cr01_decent_bid_offer_vs_holdings] Found {len(universe_cusips_raw)} CR01 universe CUSIPs\n")
        
        combined_cusips = sorted(set(universe_cusips_raw) | set(holdings_cusips_raw))
        
        data = pd.read_parquet(bql_path)
        filtered = data[data["CUSIP"].isin(combined_cusips)].copy()
        filtered = filtered.dropna(subset=["Value"])
        
        if filtered.empty:
            raise ValueError("No BQL data found for CR01 CUSIPs")
        
        name_lookup = build_name_lookup(filtered)
        wide_values = filtered.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        summaries: List[CR01PairSummary] = []
        for universe_cusip, holdings_cusip in product(universe_cusips_raw, holdings_cusips_raw):
            summary = compute_cr01_pair_summary(wide_values, universe_cusip, holdings_cusip, name_lookup)
            if summary is not None:
                summaries.append(summary)
        
        results_df = pd.DataFrame([{
            "universe_name": p.universe_name,
            "holdings_name": p.holdings_name,
            "Last": p.last_value,
            "Avg": p.average_value,
            "vs Avg": p.vs_average,
            "Z Score": p.z_score,
            "Percentile": p.percentile,
            "universe_cusip": p.universe_cusip,
            "holdings_cusip": p.holdings_cusip,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["universe_name"] = results_df["universe_name"].map(ensure_ascii)
        results_df["holdings_name"] = results_df["holdings_name"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[executable_cr01_decent_bid_offer_vs_holdings] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[executable_cr01_decent_bid_offer_vs_holdings] ERROR: {str(e)}\n")
        return pd.DataFrame()

def run_all_combos_vs_holdings_analysis(
    bql_path: Path,
    portfolio_path: Path,
    output_dir: Path,
    validation_log: io.StringIO,
) -> pd.DataFrame:
    """Run all combos vs holdings analysis (all portfolio CUSIPs vs all portfolio CUSIPs)."""
    try:
        validation_log.write(f"[all_combos_vs_holdings] Starting analysis...\n")
        holdings_cusips_raw = load_all_portfolio_cusips(portfolio_path)
        validation_log.write(f"[all_combos_vs_holdings] Found {len(holdings_cusips_raw)} portfolio CUSIPs\n")
        
        universe_cusips_raw = load_all_portfolio_cusips(portfolio_path)
        combined_cusips = sorted(set(universe_cusips_raw) | set(holdings_cusips_raw))
        
        data = pd.read_parquet(bql_path)
        filtered = data[data["CUSIP"].isin(combined_cusips)].copy()
        filtered = filtered.dropna(subset=["Value"])
        
        if filtered.empty:
            raise ValueError("No BQL data found for portfolio CUSIPs")
        
        name_lookup = build_name_lookup(filtered)
        wide_values = filtered.pivot_table(index="Date", columns="CUSIP", values="Value", aggfunc="last").sort_index()
        
        summaries: List[CR01PairSummary] = []
        for universe_cusip, holdings_cusip in product(universe_cusips_raw, holdings_cusips_raw):
            summary = compute_cr01_pair_summary(wide_values, universe_cusip, holdings_cusip, name_lookup)
            if summary is not None:
                summaries.append(summary)
        
        results_df = pd.DataFrame([{
            "universe_name": p.universe_name,
            "holdings_name": p.holdings_name,
            "Last": p.last_value,
            "Avg": p.average_value,
            "vs Avg": p.vs_average,
            "Z Score": p.z_score,
            "Percentile": p.percentile,
            "universe_cusip": p.universe_cusip,
            "holdings_cusip": p.holdings_cusip,
        } for p in summaries])
        
        results_df = results_df.sort_values("Z Score", ascending=False, na_position="last")
        results_df["universe_name"] = results_df["universe_name"].map(ensure_ascii)
        results_df["holdings_name"] = results_df["holdings_name"].map(ensure_ascii)
        
        # Format numeric columns to 1 decimal place
        results_df = format_numeric_columns(results_df)
        
        validation_log.write(f"[all_combos_vs_holdings] Success: {len(results_df)} pairs generated\n")
        
        return results_df
    except Exception as e:
        validation_log.write(f"[all_combos_vs_holdings] ERROR: {str(e)}\n")
        return pd.DataFrame()

# ============================================================================
# OUTPUT FORMATTING FUNCTIONS
# ============================================================================

def format_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Format all numeric columns to 1 decimal place."""
    df_formatted = df.copy()
    for col in df_formatted.columns:
        if pd.api.types.is_numeric_dtype(df_formatted[col]):
            df_formatted[col] = df_formatted[col].round(1)
    return df_formatted

def write_excel_file(output_dir: Path, results: Dict[str, pd.DataFrame], bql_path: Path) -> None:
    """
    Write all results to a single Excel file with formatted tables on separate sheets.
    
    Args:
        output_dir: Directory to write Excel file to.
        results: Dictionary mapping analysis names to DataFrames.
        bql_path: Path to BQL parquet file for date range info.
    """
    output_path = output_dir / "comb.xlsx"
    
    # Read date range from bql.parquet for metadata
    bql_df = pd.read_parquet(bql_path)
    min_date = bql_df["Date"].min()
    max_date = bql_df["Date"].max()
    date_range_str = f"{min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}"
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write each non-empty DataFrame to a separate sheet
        for analysis_name, df in results.items():
            if df.empty:
                continue  # Skip empty DataFrames
            
            # Get tab name (max 31 characters for Excel)
            tab_name = EXCEL_TAB_NAMES.get(analysis_name, analysis_name[:31])
            
            # Format DataFrame (numeric columns to 1 decimal, ASCII-safe strings)
            df_formatted = format_numeric_columns(df.copy())
            for col in df_formatted.columns:
                if df_formatted[col].dtype == 'object':
                    df_formatted[col] = df_formatted[col].map(ensure_ascii)
            
            # Write to Excel sheet
            df_formatted.to_excel(writer, sheet_name=tab_name, index=False)
            
            # Get the worksheet to format it
            worksheet = writer.sheets[tab_name]
            
            # Auto-fit column widths
            for idx, col in enumerate(df_formatted.columns, start=1):
                column_letter = get_column_letter(idx)
                # Calculate max width (content + padding)
                max_length = max(
                    len(str(col)),  # Header length
                    df_formatted[col].astype(str).map(len).max() if len(df_formatted) > 0 else 0
                )
                # Set width with some padding (min 10, max 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Create Excel table with filters and banded rows
            # Table name must be unique and valid (no spaces, starts with letter)
            # Use full analysis_name but sanitize it (replace spaces/special chars, ensure starts with letter)
            sanitized_name = analysis_name.replace(' ', '_').replace('-', '_').replace('.', '_')
            # Remove any invalid characters and ensure it starts with a letter
            sanitized_name = ''.join(c if c.isalnum() or c == '_' else '' for c in sanitized_name)
            if sanitized_name and not sanitized_name[0].isalpha():
                sanitized_name = 'T_' + sanitized_name
            table_name = f"Table_{sanitized_name}"[:255]  # Excel table name max length is 255
            table_range = f"A1:{get_column_letter(len(df_formatted.columns))}{len(df_formatted) + 1}"
            table = Table(displayName=table_name, ref=table_range)
            
            # Apply table style with banded rows and filters
            style = TableStyleInfo(
                name="TableStyleMedium9",  # Medium style with banded rows
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,  # Banded rows
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            
            # Add table to worksheet
            worksheet.add_table(table)
    
    print(f"\nExcel file written to: {output_path}")
    print(f"Date Range: {date_range_str}")
    print(f"Total sheets: {sum(1 for df in results.values() if not df.empty)}")

def format_table_section(title: str, df: pd.DataFrame, top_n: int) -> str:
    """Format a table section for comb.txt with nice formatting."""
    if df.empty:
        return f"\n{'='*80}\n{title}\n{'='*80}\nNo data available.\n"
    
    top_df = df.head(top_n).copy()
    # Format numeric columns to 1 decimal place
    top_df = format_numeric_columns(top_df)
    # Ensure ASCII-safe output for txt file
    for col in top_df.columns:
        if top_df[col].dtype == 'object':
            top_df[col] = top_df[col].map(ensure_ascii)
    table_str = top_df.to_string(index=False)
    
    section = f"\n{'='*80}\n{title}\n{'='*80}\n"
    section += f"Total pairs: {len(df):,}\n"
    section += f"Showing top {min(top_n, len(df))} pairs:\n\n"
    section += table_str
    section += f"\n\n"
    
    return section

def write_comb_txt(output_dir: Path, results: Dict[str, pd.DataFrame], bql_path: Path) -> None:
    """Write formatted tables to comb.txt."""
    output_path = output_dir / "comb.txt"
    
    # Read date range from bql.parquet
    bql_df = pd.read_parquet(bql_path)
    min_date = bql_df["Date"].min()
    max_date = bql_df["Date"].max()
    date_range_str = f"{min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}"
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("PAIR ANALYTICS RESULTS\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Date Range: {date_range_str}\n")
        f.write("="*80 + "\n\n")
        
        for analysis_name, df in results.items():
            if not df.empty:
                title = TABLE_TITLES.get(analysis_name, analysis_name)
                f.write(format_table_section(title, df, DISPLAY["top_n_pairs"]))
        
        f.write("\n" + "="*80 + "\n")
        f.write("END OF REPORT\n")
    
    print(f"\nFormatted tables written to: {output_path}")

def write_validation_txt(output_dir: Path, validation_log: io.StringIO) -> None:
    """Write validation messages to comb_validation.txt."""
    output_path = output_dir / "comb_validation.txt"
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("PAIR ANALYTICS VALIDATION LOG\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("="*80 + "\n\n")
        f.write(validation_log.getvalue())
        f.write("\n" + "="*80 + "\n")
        f.write("END OF VALIDATION LOG\n")
    
    print(f"Validation log written to: {output_path}")

# ============================================================================
# MAIN EXECUTION FUNCTION
# ============================================================================

def main() -> None:
    """Main execution function that runs all analyses."""
    validation_log = io.StringIO()
    results: Dict[str, pd.DataFrame] = {}
    
    validation_log.write("Starting pair analytics execution...\n")
    validation_log.write(f"Configuration:\n")
    validation_log.write(f"  Recent date percent: {FILTERS['recent_date_percent']}\n")
    validation_log.write(f"  Top N pairs: {DISPLAY['top_n_pairs']}\n")
    validation_log.write(f"  Target bond ticker: {TARGET_BOND['ticker']}\n")
    validation_log.write(f"  Target bond security: {TARGET_BOND['security']}\n")
    validation_log.write(f"  CR01 tight bid threshold: {FILTERS['cr01_tight_bid_threshold']}\n")
    validation_log.write(f"  CR01 wide offer threshold: {FILTERS['cr01_wide_offer_threshold']}\n")
    validation_log.write(f"  Bid/Offer threshold: {FILTERS['bid_offer_threshold']}\n\n")
    
    # Run all analyses
    print("Running all combinations analysis...")
    results["all_comb"] = run_all_comb_analysis(
        BQL_PARQUET_PATH, HISTORICAL_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    print("Running term combinations analysis...")
    results["term_comb"] = run_term_comb_analysis(
        BQL_PARQUET_PATH, HISTORICAL_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    print("Running ticker combinations analysis...")
    results["ticker_comb"] = run_ticker_comb_analysis(
        BQL_PARQUET_PATH, HISTORICAL_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    print("Running custom sector analysis...")
    results["custom_sector"] = run_custom_sector_analysis(
        BQL_PARQUET_PATH, HISTORICAL_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    print("Running custom bond combinations analysis...")
    results["custom_bond_comb"] = run_custom_bond_comb_analysis(
        BQL_PARQUET_PATH, HISTORICAL_PARQUET_PATH, UNIVERSE_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    print("Running custom bond vs holdings analysis...")
    results["custom_bond_vs_holdings"] = run_custom_bond_vs_holdings_analysis(
        BQL_PARQUET_PATH, HISTORICAL_PARQUET_PATH, UNIVERSE_PARQUET_PATH,
        RUNS_TODAY_CSV_PATH, PORTFOLIO_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    print("Running CAD cheap vs USD analysis...")
    results["cad_cheap_vs_usd"] = run_cad_cheap_vs_usd_analysis(
        BQL_PARQUET_PATH, HISTORICAL_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    print("Running CAD rich vs USD analysis...")
    results["cad_rich_vs_usd"] = run_cad_rich_vs_usd_analysis(
        BQL_PARQUET_PATH, HISTORICAL_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    print("Running executable CR01 vs holdings analysis...")
    results["executable_cr01_vs_holdings"] = run_executable_cr01_vs_holdings_analysis(
        BQL_PARQUET_PATH, RUNS_TODAY_CSV_PATH, PORTFOLIO_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    print("Running executable CR01 decent bid offer vs holdings analysis...")
    results["executable_cr01_decent_bid_offer_vs_holdings"] = run_executable_cr01_decent_bid_offer_vs_holdings_analysis(
        BQL_PARQUET_PATH, RUNS_TODAY_CSV_PATH, PORTFOLIO_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    print("Running all combos vs holdings analysis...")
    results["all_combos_vs_holdings"] = run_all_combos_vs_holdings_analysis(
        BQL_PARQUET_PATH, PORTFOLIO_PARQUET_PATH, OUTPUT_DIR, validation_log
    )
    
    # Display all tables to console
    print("\n" + "="*80)
    print("PAIR ANALYTICS RESULTS")
    print("="*80)
    for analysis_name, df in results.items():
        if not df.empty:
            title = TABLE_TITLES.get(analysis_name, analysis_name)
            print(f"\n{title}")
            print("-"*80)
            top_df = df.head(DISPLAY["top_n_pairs"]).copy()
            # Format numeric columns to 1 decimal place
            top_df = format_numeric_columns(top_df)
            # Ensure ASCII-safe output for console
            for col in top_df.columns:
                if top_df[col].dtype == 'object':
                    top_df[col] = top_df[col].map(ensure_ascii)
            print(top_df.to_string(index=False))
            print(f"\nTotal pairs: {len(df):,}")
    
    # Write output files
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_excel_file(OUTPUT_DIR, results, BQL_PARQUET_PATH)
    write_comb_txt(OUTPUT_DIR, results, BQL_PARQUET_PATH)
    write_validation_txt(OUTPUT_DIR, validation_log)
    
    validation_log.close()
    print("\nAll analyses completed!")

if __name__ == "__main__":
    main()

